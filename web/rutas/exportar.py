"""
exportar.py — Exportación a PDF y XLSX

PDF: ReportLab (ya instalado, sin dependencias Windows)
XLSX: openpyxl puro (ya instalado, sin macros VBA)
"""
import io
import os
import shutil
import tempfile
import threading
from datetime import datetime
from typing import List, Dict

from fastapi import APIRouter, Request, Depends, Form
from fastapi.responses import StreamingResponse, JSONResponse

from web.auth import require_login
from web.database import guardar_cotizacion_db, cargar_config
from web.rutas.carrito import get_carrito

router = APIRouter(prefix="/api/exportar", tags=["exportar"])

# Lock para serializar accesos concurrentes a Excel/COM
_EXCEL_LOCK = threading.Lock()


def _num(v) -> float:
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


# ─────────────────────────────────────────────
# Helpers de texto / moneda  (módulo-nivel)
# ─────────────────────────────────────────────

def _fecha_larga(fecha_str: str) -> str:
    MESES = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    try:
        d, m, y = fecha_str.split("/")
        return f"San Martín de Porres, {int(d)} de {MESES[int(m)-1]} del {y}"
    except Exception:
        return fecha_str


def _entero_a_letras(n: int) -> str:
    if n == 0:
        return "CERO"
    UNIDADES = [
        "", "UN", "DOS", "TRES", "CUATRO", "CINCO", "SEIS", "SIETE", "OCHO", "NUEVE",
        "DIEZ", "ONCE", "DOCE", "TRECE", "CATORCE", "QUINCE",
        "DIECISÉIS", "DIECISIETE", "DIECIOCHO", "DIECINUEVE",
    ]
    DECENAS = [
        "", "DIEZ", "VEINTE", "TREINTA", "CUARENTA", "CINCUENTA",
        "SESENTA", "SETENTA", "OCHENTA", "NOVENTA",
    ]
    CENTENAS = [
        "", "CIENTO", "DOSCIENTOS", "TRESCIENTOS", "CUATROCIENTOS", "QUINIENTOS",
        "SEISCIENTOS", "SETECIENTOS", "OCHOCIENTOS", "NOVECIENTOS",
    ]
    VEINTI = [
        "", "VEINTIUNO", "VEINTIDÓS", "VEINTITRÉS", "VEINTICUATRO",
        "VEINTICINCO", "VEINTISÉIS", "VEINTISIETE", "VEINTIOCHO", "VEINTINUEVE",
    ]
    partes = []
    if n >= 1000:
        miles = n // 1000
        partes.append("MIL" if miles == 1 else _entero_a_letras(miles) + " MIL")
        n = n % 1000
    if n >= 100:
        c = n // 100
        resto = n % 100
        partes.append("CIEN" if (c == 1 and resto == 0) else CENTENAS[c])
        n = n % 100
    if n >= 20:
        d, u = n // 10, n % 10
        if d == 2:
            partes.append("VEINTE" if u == 0 else VEINTI[u])
        else:
            partes.append(DECENAS[d] if u == 0 else DECENAS[d] + " Y " + UNIDADES[u])
    elif n > 0:
        partes.append(UNIDADES[n])
    return " ".join(partes)


def _monto_en_letras(monto: float, usd: bool = False) -> str:
    moneda_str = "DÓLARES" if usd else "SOLES"
    entero = int(monto)
    centavos = round((monto - entero) * 100)
    return f"{_entero_a_letras(entero)} {centavos:02d}/100 {moneda_str}"


# ─────────────────────────────────────────────
# Export PDF
# ─────────────────────────────────────────────

def _generar_pdf_reportlab(
    carrito: List[Dict],
    cliente: str,
    atencion: str,
    moneda: str,
    proyecto: str,
    fecha: str,
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar: float = 3.8,
    validez: str = "30 días",
    encabezado_tabla: str = "",
    cotizacion_id: int = 0,
) -> bytes:
    """Genera PDF con WeasyPrint usando cotizacion_pdf.html — sin navegador, funciona en Linux/cloud."""
    import base64
    from jinja2 import Environment, FileSystemLoader
    from weasyprint import HTML as WeasyprintHTML

    es_usd = (moneda == "DOLARES")
    sym    = "$" if es_usd else "S/."

    def precio_display(p: float) -> float:
        return round(p / dolar, 2) if es_usd else p

    # ── Número de cotización ──
    if cotizacion_id:
        try:
            año = datetime.strptime(fecha, "%d/%m/%Y").year
        except Exception:
            año = datetime.now().year
        numero_cotizacion = f"COT-{año}-{cotizacion_id:05d}"
    else:
        try:
            numero_cotizacion = datetime.strptime(fecha, "%d/%m/%Y").strftime("%d-%m-%Y")
        except Exception:
            numero_cotizacion = fecha

    # ── Productos formateados para el template ──
    productos = []
    subtotal_soles = 0.0
    for item in carrito:
        pu_s = _num(item.get("precio_unitario", 0))
        cant = _num(item.get("cantidad", 1))
        pt_s = pu_s * cant
        subtotal_soles += pt_s
        productos.append({
            "descripcion":     item.get("descripcion", ""),
            "unidad":          item.get("unidad", "UND"),
            "cantidad":        int(cant),
            "precio_unitario": f"{precio_display(pu_s):,.2f}",
            "total":           f"{precio_display(pt_s):,.2f}",
        })

    subtotal = precio_display(subtotal_soles)
    igv      = round(subtotal * 0.18, 2)
    total    = round(subtotal + igv, 2)
    moneda_str = "DÓLARES AMERICANOS" if es_usd else "SOLES PERUANOS"

    # ── Logo en base64 (WeasyPrint no resuelve rutas relativas en string HTML) ──
    logo_src = None
    logo_path = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "static", "IMAGEN_LOGO_AROLUZEIRL_BARRITA.png")
    )
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as _f:
            logo_src = "data:image/png;base64," + base64.b64encode(_f.read()).decode()

    # ── Renderizar template HTML ──
    template_dir = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "templates")
    )
    env  = Environment(loader=FileSystemLoader(template_dir))
    tmpl = env.get_template("cotizacion_pdf.html")
    html_str = tmpl.render(
        numero_cotizacion = numero_cotizacion,
        fecha_completa    = _fecha_larga(fecha),
        razon_social      = cliente_nombre or cliente or "—",
        ruc               = cliente_ruc or "—",
        atencion          = atencion or "—",
        correo            = atencion_email or "—",
        proyecto          = proyecto or "—",
        ubicacion         = cliente_ubicacion or "—",
        encabezado_tabla  = encabezado_tabla.strip().upper(),
        productos         = productos,
        monto_letras      = _monto_en_letras(total, es_usd),
        subtotal          = f"{subtotal:,.2f}",
        igv               = f"{igv:,.2f}",
        total             = f"{total:,.2f}",
        sym_label         = sym,
        moneda            = moneda_str,
        validez           = validez.upper(),
        logo_src          = logo_src,
    )

    # ── Convertir HTML → PDF con WeasyPrint (sin navegador) ──
    return WeasyprintHTML(string=html_str).write_pdf()


# ─────────────────────────────────────────────
# PDF desde plantilla xlsm (Excel COM)
# ─────────────────────────────────────────────

def _generar_pdf_desde_excel(
    carrito: List[Dict],
    cliente: str,
    atencion: str,
    moneda: str,
    proyecto: str,
    fecha: str,
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar: float = 3.8,
    validez: str = "30 días",
) -> bytes:
    """Genera PDF usando la plantilla xlsm original vía xlwings/COM.

    Requiere: Windows, Microsoft Excel instalado, plantilla configurada.
    Lanza excepción si cualquier condición no se cumple → _generar_pdf() cae a ReportLab.
    """
    import sys
    import xlwings as xw
    import glob as _glob

    cfg = cargar_config()
    plantilla_path = cfg.get("rutas", {}).get("plantilla_excel", "")
    if not plantilla_path or not os.path.exists(plantilla_path):
        raise FileNotFoundError("Plantilla xlsm no configurada o no encontrada")

    es_usd = (moneda == "DOLARES")

    tmp_dir = tempfile.mkdtemp()
    tmp_xlsm = os.path.join(tmp_dir, "cotizacion_tmp.xlsm")
    shutil.copy2(plantilla_path, tmp_xlsm)

    with _EXCEL_LOCK:
        # visible=True: garantiza que ws.activate() y las macros VBA funcionen igual
        # que en el desktop (os.startfile abre Excel visiblemente)
        app = xw.App(visible=True, add_book=False)
        try:
            wb = app.books.open(tmp_xlsm)
            ws = wb.sheets["COTIZACIÓN"]

            # ── Cabecera ──
            ws["M5"].value = cliente_nombre or cliente
            ws["M6"].value = atencion
            ws["M8"].value = "DÓLARES" if es_usd else "SOLES"
            ws["B10"].value = proyecto
            ws["G8"].value  = cliente_ruc or ""
            ws["G9"].value  = atencion_email or ""
            ws["G10"].value = cliente_ubicacion or ""

            # ── Productos ──
            FILA_INICIO = 17
            MAX_FILAS_TEMPLATE = 14   # filas 17-30 pre-numeradas en plantilla

            for i, item in enumerate(carrito):
                fila = FILA_INICIO + i
                pu_s = _num(item.get("precio_unitario", 0))
                cant = _num(item.get("cantidad", 1))
                pu   = round(pu_s / dolar, 2) if es_usd else pu_s

                if i >= MAX_FILAS_TEMPLATE:
                    # Insertar fila antes de la sección de totales y copiar formato
                    ws.range(f"{fila}:{fila}").insert("down")
                    ws.range(f"{fila-1}:{fila-1}").copy()
                    ws.range(f"{fila}:{fila}").paste(paste="formats")

                ws[f"A{fila}"].value = i + 1
                ws[f"B{fila}"].value = item.get("descripcion", "")
                ws[f"F{fila}"].value = item.get("unidad", "UND")
                ws[f"G{fila}"].value = int(cant)
                ws[f"H{fila}"].value = pu
                ws[f"I{fila}"].formula = f"=G{fila}*H{fila}"

            # ── Limpiar filas vacías sobrantes ──
            if len(carrito) < MAX_FILAS_TEMPLATE:
                ultima_datos    = FILA_INICIO + len(carrito) - 1
                ultima_template = FILA_INICIO + MAX_FILAS_TEMPLATE - 1  # = 30
                sobrantes = ultima_template - ultima_datos
                if sobrantes > 0:
                    rango = f"{ultima_datos + 1}:{ultima_template}"
                    ws.range(rango).delete("up")

            # ── Calcular ──
            wb.app.calculate()

            # ── ACTIVAR COTIZACIÓN justo antes de la macro (fix idéntico al Tkinter) ──
            # La macro lee ActiveSheet.Range("A5") para el nombre del archivo y usa
            # ActiveSheet.ExportAsFixedFormat → COTIZACIÓN debe ser la hoja activa.
            ws.activate()
            print(f"[exportar] Hoja activa antes de macro: {wb.sheets.active.name}", file=sys.stderr)

            # Construir ruta de salida con \ al final (la VBA lo exige)
            tmp_dir_win = tmp_dir.replace("/", "\\")
            if not tmp_dir_win.endswith("\\"):
                tmp_dir_win += "\\"

            wb.app.api.Run("GUARDARPDF", tmp_dir_win)
            print(f"[exportar] Macro GUARDARPDF ejecutada en: {tmp_dir_win}", file=sys.stderr)

            # Buscar el PDF que generó la macro
            pdf_files = _glob.glob(os.path.join(tmp_dir, "*.pdf"))
            if not pdf_files:
                raise RuntimeError("La macro GUARDARPDF no generó ningún PDF")
            tmp_pdf = sorted(pdf_files, key=os.path.getmtime)[-1]
            print(f"[exportar] PDF generado: {os.path.basename(tmp_pdf)}", file=sys.stderr)

            with open(tmp_pdf, "rb") as f:
                return f.read()

        except Exception as e:
            print(f"[exportar] ERROR en _generar_pdf_desde_excel: {type(e).__name__}: {e}", file=sys.stderr)
            raise

        finally:
            try:
                wb.close()
            except Exception:
                pass
            try:
                app.quit()
            except Exception:
                pass
            shutil.rmtree(tmp_dir, ignore_errors=True)


def _find_browser() -> str | None:
    """Devuelve la ruta al ejecutable de Edge o Chrome si está disponible en Windows."""
    username = os.environ.get("USERNAME", "")
    candidates = [
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        rf"C:\Users\{username}\AppData\Local\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files\Google\Chrome\Application\chrome.exe",
        r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
        rf"C:\Users\{username}\AppData\Local\Google\Chrome\Application\chrome.exe",
    ]
    return next((p for p in candidates if os.path.exists(p)), None)


def _construir_html_cotizacion(
    carrito: List[Dict],
    cliente: str,
    atencion: str,
    moneda: str,
    proyecto: str,
    fecha: str,
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar: float = 3.8,
    validez: str = "30 días",
    encabezado_tabla: str = "",
    cotizacion_id: int = 0,
) -> str:
    """Lee cotizacion_aroluz.html, conecta todos los {{ }} y devuelve HTML listo para PDF."""
    import re

    _base = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))
    template_path = os.path.join(_base, "cotizacion_aroluz.html")
    with open(template_path, "r", encoding="utf-8") as f:
        html = f.read()

    es_usd = (moneda == "DOLARES")
    sym = "$" if es_usd else "S/"

    def precio_display(p: float) -> float:
        return round(p / dolar, 2) if es_usd else p

    # ── Totales ──
    subtotal_soles = sum(
        _num(item.get("precio_unitario", 0)) * _num(item.get("cantidad", 1))
        for item in carrito
    )
    subtotal = precio_display(subtotal_soles)
    igv = round(subtotal * 0.18, 2)
    total = round(subtotal + igv, 2)

    # ── Fecha ──
    MESES = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
    ]
    try:
        d, m, y = fecha.split("/")
        fecha_larga = f"{int(d)} de {MESES[int(m)-1]} del {y}"
        _fecha_cot = f"{d}-{m}-{y}"
    except Exception:
        fecha_larga = fecha
        _fecha_cot = fecha

    if cotizacion_id:
        try:
            _año_cot = datetime.strptime(fecha, "%d/%m/%Y").year
        except Exception:
            _año_cot = datetime.now().year
        numero_cotizacion = f"COT-{_año_cot}-{cotizacion_id:05d}"
    else:
        _partes_cot = [p.strip() for p in [cliente_nombre or cliente, proyecto, atencion] if p and p.strip()]
        numero_cotizacion = "  ".join(_partes_cot + [_fecha_cot])

    # Monto en letras — el HTML tiene "{{ monto_letras }} SOLES" literal
    monto_letras_str = _monto_en_letras(total, es_usd)  # ya incluye "SOLES" o "DÓLARES"
    moneda_str = "DÓLARES AMERICANOS" if es_usd else "SOLES PERUANOS"

    # ── 0. Reemplazar texto-logo por imagen PNG (data URI, no modifica el archivo HTML) ──
    _logo_path = os.path.join(_base, "web", "static", "IMAGEN_LOGO_AROLUZEIRL_BARRITA.png")
    if os.path.exists(_logo_path):
        import base64 as _b64
        with open(_logo_path, "rb") as _lf:
            _logo_b64 = _b64.b64encode(_lf.read()).decode()
        _img_tag = (
            f'<img src="data:image/png;base64,{_logo_b64}" '
            f'style="height:50px; width:auto; display:block; margin-left:25px;">'
        )
        # Bloque de texto-logo que aparece 2 veces (pág. 1 y pág. 2)
        _old_logo = (
            '<div class="logo-text"><span style="font-style:italic;">AROLUZ</span>'
            ' <span style="font-size:14pt; font-weight:400;">E.I.R.L.</span></div>\n'
            '        <div style="height:2px; background:#ffffff; margin:2px 0;"></div>'
        )
        html = html.replace(_old_logo, _img_tag)
        # Quitar el comentario de ayuda que quedó en pág. 1
        html = re.sub(
            r'<!--\s*Si tienes el logo como imagen.*?-->\s*',
            '',
            html,
            flags=re.DOTALL,
        )

    # ── 1. Limpiar bloques de ejemplo del HTML ──
    html = re.sub(r'<!--\s*BLOQUE DINÁMICO.*?-->', '', html, flags=re.DOTALL)
    html = re.sub(r'<!-- FILAS DE EJEMPLO \(eliminar en producción\) -->', '', html)

    # ── 2. Eliminar filas estáticas vacías (1-15) ──
    html = re.sub(
        r'<tr><td class="num">\d+</td><td class="desc"></td>'
        r'<td class="und">UND</td><td class="cant"></td>'
        r'<td class="punit"></td><td class="total">-</td></tr>\s*',
        '',
        html,
    )

    # ── 3. Generar filas reales de productos ──
    product_rows = ""
    for i, item in enumerate(carrito):
        pu_s = _num(item.get("precio_unitario", 0))
        cant = _num(item.get("cantidad", 1))
        pt_s = pu_s * cant
        pu_d = precio_display(pu_s)
        pt_d = precio_display(pt_s)
        und = item.get("unidad", "UND")
        desc = item.get("descripcion", "")
        product_rows += (
            f'<tr>'
            f'<td class="num">{i + 1}</td>'
            f'<td class="desc">{desc}</td>'
            f'<td class="und">{und}</td>'
            f'<td class="cant">{int(cant)}</td>'
            f'<td class="punit">{pu_d:,.2f}</td>'
            f'<td class="total">{pt_d:,.2f}</td>'
            f'</tr>\n'
        )
    html = html.replace(
        '<!-- FILA MONTO EN LETRAS + SUBTOTAL -->',
        product_rows + '<!-- FILA MONTO EN LETRAS + SUBTOTAL -->',
    )

    # ── 4. Reemplazar {{ placeholders }} ──
    # Caso especial: " SOLES" está hardcoded después del placeholder en el HTML
    html = html.replace('{{ monto_letras }} SOLES', monto_letras_str)

    replacements = {
        '{{ numero_cotizacion }}': numero_cotizacion,
        '{{ ciudad }}':            'San Martín de Porres',
        '{{ fecha_larga }}':       fecha_larga,
        '{{ razon_social }}':      cliente_nombre or cliente or '-',
        '{{ ruc }}':               cliente_ruc or '-',
        '{{ atencion }}':          atencion or '-',
        '{{ correo }}':            atencion_email or '-',
        '{{ proyecto }}':          proyecto or '-',
        '{{ ubicacion }}':         cliente_ubicacion or '-',
        '{{ encabezado_tabla }}':  encabezado_tabla.strip().upper(),
        '{{ subtotal }}':          f'{subtotal:,.2f}',
        '{{ igv }}':               f'{igv:,.2f}',
        '{{ total }}':             f'{total:,.2f}',
        '{{ moneda }}':            moneda_str,
    }
    for placeholder, value in replacements.items():
        html = html.replace(placeholder, value)

    # Ajustar etiquetas de moneda si es USD
    if es_usd:
        html = html.replace('SUBTOTAL S/.', 'SUBTOTAL $')
        html = html.replace('TOTAL S/.', 'TOTAL $')

    # Validez real en pág. 2 (el HTML tiene hardcoded "3 DÍAS")
    html = html.replace('>3 DÍAS<', f'>{validez.upper()}<')

    # ── CSS inyectado en memoria (no se toca el archivo HTML) ──
    # @page con los mismos márgenes que el .page div → márgenes seguros en TODAS las páginas
    # incluido el overflow. Se anula el padding del .page para evitar doble margen.
    # thead { display: table-header-group } → cabecera de tabla repite en cada página.
    html = html.replace(
        '</style>',
        '''
  /* ── PRINT FIX (generado en Python, no modifica el HTML) ── */
  @page { margin: 10mm 12mm; size: A4; }
  .page  { padding: 0 !important; width: 100% !important;
            margin: 0 !important; min-height: auto !important; }
  thead  { display: table-header-group; }
  tfoot  { display: table-footer-group; }
  tr     { page-break-inside: avoid; }
  table  { page-break-inside: auto;  }
  .cotizacion-header .numero { font-weight: bold; font-size: 9pt; letter-spacing: 0.5px; }
  </style>''',
        1,
    )

    return html


def _generar_pdf_html(
    carrito: List[Dict],
    cliente: str,
    atencion: str,
    moneda: str,
    proyecto: str,
    fecha: str,
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar: float = 3.8,
    validez: str = "30 días",
    encabezado_tabla: str = "",
    cotizacion_id: int = 0,
) -> bytes:
    """Fallback: genera PDF con Playwright usando cotizacion_pdf.html.

    Sólo se usa si WeasyPrint no está disponible (Windows local sin GTK+).
    Requiere: pip install playwright && playwright install chromium
    """
    import base64
    import concurrent.futures
    from jinja2 import Environment, FileSystemLoader

    # Renderizar la misma plantilla que usa WeasyPrint
    es_usd = (moneda == "DOLARES")
    sym    = "$" if es_usd else "S/."

    def precio_display(p: float) -> float:
        return round(p / dolar, 2) if es_usd else p

    if cotizacion_id:
        try:
            año = datetime.strptime(fecha, "%d/%m/%Y").year
        except Exception:
            año = datetime.now().year
        numero_cotizacion = f"COT-{año}-{cotizacion_id:05d}"
    else:
        try:
            numero_cotizacion = datetime.strptime(fecha, "%d/%m/%Y").strftime("%d-%m-%Y")
        except Exception:
            numero_cotizacion = fecha

    productos = []
    subtotal_soles = 0.0
    for item in carrito:
        pu_s = _num(item.get("precio_unitario", 0))
        cant = _num(item.get("cantidad", 1))
        pt_s = pu_s * cant
        subtotal_soles += pt_s
        productos.append({
            "descripcion":     item.get("descripcion", ""),
            "unidad":          item.get("unidad", "UND"),
            "cantidad":        int(cant),
            "precio_unitario": f"{precio_display(pu_s):,.2f}",
            "total":           f"{precio_display(pt_s):,.2f}",
        })

    subtotal = precio_display(subtotal_soles)
    igv      = round(subtotal * 0.18, 2)
    total    = round(subtotal + igv, 2)
    moneda_str = "DÓLARES AMERICANOS" if es_usd else "SOLES PERUANOS"

    logo_src = None
    logo_path = os.path.normpath(
        os.path.join(os.path.dirname(__file__), "..", "static", "IMAGEN_LOGO_AROLUZEIRL_BARRITA.png")
    )
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as _f:
            logo_src = "data:image/png;base64," + base64.b64encode(_f.read()).decode()

    template_dir = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", "templates"))
    env  = Environment(loader=FileSystemLoader(template_dir))
    html = env.get_template("cotizacion_pdf.html").render(
        numero_cotizacion = numero_cotizacion,
        fecha_completa    = _fecha_larga(fecha),
        razon_social      = cliente_nombre or cliente or "—",
        ruc               = cliente_ruc or "—",
        atencion          = atencion or "—",
        correo            = atencion_email or "—",
        proyecto          = proyecto or "—",
        ubicacion         = cliente_ubicacion or "—",
        encabezado_tabla  = encabezado_tabla.strip().upper(),
        productos         = productos,
        monto_letras      = _monto_en_letras(total, es_usd),
        subtotal          = f"{subtotal:,.2f}",
        igv               = f"{igv:,.2f}",
        total             = f"{total:,.2f}",
        sym_label         = sym,
        moneda            = moneda_str,
        validez           = validez.upper(),
        logo_src          = logo_src,
    )

    def _run_playwright() -> bytes:
        from playwright.sync_api import sync_playwright  # type: ignore

        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page()
            page.set_content(html, wait_until="networkidle")
            pdf_bytes = page.pdf(
                format="A4",
                print_background=True,
                display_header_footer=False,
                margin={"top": "0", "right": "0", "bottom": "0", "left": "0"},
            )
            browser.close()
        return pdf_bytes

    with concurrent.futures.ThreadPoolExecutor(max_workers=1) as pool:
        return pool.submit(_run_playwright).result(timeout=60)


def _generar_pdf(
    carrito: List[Dict],
    cliente: str,
    atencion: str,
    moneda: str,
    proyecto: str,
    fecha: str,
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar: float = 3.8,
    validez: str = "30 días",
    encabezado_tabla: str = "",
    cotizacion_id: int = 0,
) -> bytes:
    """Genera PDF usando cotizacion_pdf.html.

    Orden de intentos:
    1. WeasyPrint  — funciona en Linux/cloud (Fly.io). Requiere libpango/libcairo.
    2. Playwright  — fallback para Windows local (si playwright está instalado).
    """
    import sys

    # 1. WeasyPrint (Linux / Fly.io)
    try:
        return _generar_pdf_reportlab(
            carrito, cliente, atencion, moneda, proyecto, fecha,
            cliente_nombre, cliente_ruc, cliente_ubicacion, atencion_email,
            dolar, validez, encabezado_tabla=encabezado_tabla, cotizacion_id=cotizacion_id,
        )
    except Exception as e:
        print(f"[exportar] WeasyPrint falló ({type(e).__name__}: {e}) — intentando Playwright", file=sys.stderr)

    # 2. Playwright (Windows local)
    return _generar_pdf_html(
        carrito, cliente, atencion, moneda, proyecto, fecha,
        cliente_nombre, cliente_ruc, cliente_ubicacion, atencion_email,
        dolar, validez, encabezado_tabla, cotizacion_id,
    )


# ─────────────────────────────────────────────
# Export XLSX
# ─────────────────────────────────────────────

def _generar_xlsx(
    carrito: List[Dict],
    cliente: str,
    atencion: str,
    moneda: str,
    proyecto: str,
    fecha: str,
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar: float = 3.8,
    validez: str = "30 días",
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotización"

    # ── Helpers de moneda ──
    es_usd = (moneda == "DOLARES")
    sym = "$" if es_usd else "S/"

    def precio_display(p: float) -> float:
        return round(p / dolar, 2) if es_usd else p

    COLOR_HEADER = "1A3A5C"
    COLOR_ALT    = "F0F4F8"
    COLOR_TOTAL  = "E8F0FE"

    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def h_cell(ws, row, col, value, bold=False, bg=None, align="left", number_format=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(bold=bold, color="FFFFFF" if bg == COLOR_HEADER else "000000", size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
        if bg:
            cell.fill = PatternFill("solid", fgColor=bg)
        if number_format:
            cell.number_format = number_format
        cell.border = border
        return cell

    # Título
    ws.merge_cells("A1:E1")
    cell_titulo = ws["A1"]
    cell_titulo.value = "AROLUZ — COTIZACIÓN"
    cell_titulo.font = Font(bold=True, size=14, color=COLOR_HEADER)
    cell_titulo.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Info cliente — filas dinámicas
    info_filas = []
    cli_txt = f"Cliente: {cliente}"
    if cliente_nombre:
        cli_txt += f" — {cliente_nombre}"
    info_filas.append((cli_txt, f"Proyecto: {proyecto}"))
    if cliente_ruc:
        info_filas.append((f"RUC: {cliente_ruc}", f"Moneda: {moneda}    Fecha: {fecha}"))
    else:
        info_filas.append(("", f"Moneda: {moneda}    Fecha: {fecha}"))
    if cliente_ubicacion:
        info_filas.append((f"Dirección: {cliente_ubicacion}", ""))
    if atencion:
        info_filas.append((f"Atención: {atencion}", ""))
    if atencion_email:
        info_filas.append((f"Email: {atencion_email}", ""))

    for idx, (izq, der) in enumerate(info_filas):
        fila = idx + 2
        if izq:
            ws.merge_cells(f"A{fila}:B{fila}")
            ws[f"A{fila}"] = izq
            ws[f"A{fila}"].font = Font(size=9)
        if der:
            ws.merge_cells(f"C{fila}:E{fila}")
            ws[f"C{fila}"] = der
            ws[f"C{fila}"].font = Font(size=9)

    header_row    = len(info_filas) + 2
    data_start_row = header_row + 1

    # Encabezado tabla
    encabezados = ["DESCRIPCIÓN", "CANT.", "UND", f"P.U. ({sym})", f"P.T. ({sym})"]
    for col, txt in enumerate(encabezados, 1):
        h_cell(ws, header_row, col, txt, bold=True, bg=COLOR_HEADER, align="center")
    ws.row_dimensions[header_row].height = 20

    # Datos
    total = 0.0
    for i, item in enumerate(carrito):
        row     = i + data_start_row
        pu_s    = _num(item.get("precio_unitario", 0))
        cant    = int(_num(item.get("cantidad", 1)))
        pu      = precio_display(pu_s)
        pt      = pu * cant
        total  += pt
        und    = item.get("unidad", "UND")
        desc   = item.get("descripcion", "")

        bg = None if i % 2 == 0 else COLOR_ALT

        ws.cell(row=row, column=1, value=desc).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=1).border = border
        if bg:
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=bg)

        h_cell(ws, row, 2, cant, align="center", bg=bg)
        h_cell(ws, row, 3, und,  align="center", bg=bg)
        h_cell(ws, row, 4, pu,   align="right",  bg=bg, number_format='#,##0.00')
        h_cell(ws, row, 5, pt,   align="right",  bg=bg, number_format='#,##0.00')
        ws.row_dimensions[row].height = 30

    # Fila total
    total_row = len(carrito) + data_start_row
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=COLOR_TOTAL)
    ws.cell(row=total_row, column=1).border = border

    tc = ws.cell(row=total_row, column=4, value="TOTAL")
    tc.font = Font(bold=True, size=11, color=COLOR_HEADER)
    tc.alignment = Alignment(horizontal="right")
    tc.fill = PatternFill("solid", fgColor=COLOR_TOTAL)
    tc.border = border

    tv = ws.cell(row=total_row, column=5, value=total)
    tv.font = Font(bold=True, size=11, color=COLOR_HEADER)
    tv.alignment = Alignment(horizontal="right")
    tv.fill = PatternFill("solid", fgColor=COLOR_TOTAL)
    tv.number_format = '#,##0.00'
    tv.border = border

    # Anchos de columna
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────
# Endpoints
# ─────────────────────────────────────────────

@router.post("/pdf")
async def exportar_pdf(
    request: Request,
    usuario: dict = Depends(require_login),
    cliente: str = Form(""),
    cliente_nombre: str = Form(""),
    cliente_ruc: str = Form(""),
    cliente_ubicacion: str = Form(""),
    atencion: str = Form(""),
    atencion_email: str = Form(""),
    moneda: str = Form("SOLES"),
    proyecto: str = Form(""),
    validez: str = Form("30 días"),
    encabezado_tabla: str = Form(""),
):
    carrito = get_carrito(usuario["u"])
    if not carrito:
        return JSONResponse({"ok": False, "error": "El carrito está vacío"}, status_code=400)

    cfg   = cargar_config()
    dolar = float(cfg.get("valores_defecto", {}).get("dolar", 3.8))

    fecha = datetime.now().strftime("%d/%m/%Y")

    # Guardar en historial PRIMERO para obtener el ID que irá en el PDF
    historial_id = 0
    try:
        historial_id = guardar_cotizacion_db(
            username=usuario["u"],
            cliente=cliente, atencion=atencion,
            proyecto=proyecto, moneda=moneda,
            items=carrito,
            cliente_nombre=cliente_nombre, cliente_ruc=cliente_ruc,
            cliente_ubicacion=cliente_ubicacion, atencion_email=atencion_email,
            dolar_rate=dolar, validez=validez, encabezado_tabla=encabezado_tabla,
        )
    except Exception:
        pass  # fallo silencioso — la descarga no se interrumpe

    try:
        pdf_bytes = _generar_pdf(
            carrito, cliente, atencion, moneda, proyecto, fecha,
            cliente_nombre, cliente_ruc, cliente_ubicacion, atencion_email,
            dolar=dolar, validez=validez, encabezado_tabla=encabezado_tabla,
            cotizacion_id=historial_id,
        )
    except Exception as e:
        import sys
        print(f"[exportar_pdf] Error generando PDF: {type(e).__name__}: {e}", file=sys.stderr)
        return JSONResponse({"ok": False, "error": f"Error generando PDF: {e}"}, status_code=500)

    try:
        _fecha_doc = datetime.strptime(fecha, "%d/%m/%Y").strftime("%d-%m-%Y")
    except Exception:
        _fecha_doc = datetime.now().strftime("%d-%m-%Y")
    numero_cot = f"COT-{datetime.now().year}-{historial_id:05d}" if historial_id else ""
    _partes = [p.strip() for p in [numero_cot, cliente_nombre or cliente, proyecto, atencion, _fecha_doc] if p and p.strip()]
    _base = " ".join(_partes) if _partes else "COTIZACIÓN"
    for _c in r'\/:*?"<>|':
        _base = _base.replace(_c, "")
    nombre_archivo = _base + ".pdf"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="{nombre_archivo}"',
            "X-Historial-ID": str(historial_id),
        },
    )


