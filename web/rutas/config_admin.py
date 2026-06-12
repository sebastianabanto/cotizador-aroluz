# -*- coding: utf-8 -*-
"""Configuración (ADMIN): precios, catálogo, clientes/atenciones, contactos, usuarios — extraído de web/main.py (refactor jun 2026)."""
import io
import json
import os
import shutil as _shutil
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import (
    HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse, FileResponse,
)

from web.auth import (
    require_login, require_admin, require_asistencias, get_session,
    verificar_usuario, set_session_cookie, clear_session_cookie,
)
from web.database import (
    cargar_config, guardar_config, obtener_catalogo,
    importar_catalogo_desde_excel, importar_catalogo_desde_json,
    exportar_contactos_xlsx, importar_contactos_desde_xlsx,
    agregar_cliente, agregar_atencion,
    eliminar_cliente, eliminar_atencion,
    editar_cliente, editar_atencion,
    obtener_cliente, obtener_atenciones_de_cliente,
    crear_usuario, cambiar_password, listar_usuarios,
    editar_usuario, toggle_activo_usuario, eliminar_usuario,
    get_kpis_proyectos, get_proyectos_con_stats, set_proyecto_estado,
    update_proyecto_direccion, update_proyecto_numero_oc, update_proyecto_contacto,
    update_proyecto_notas, renombrar_proyecto,
    add_adjunto, list_adjuntos, get_adjunto_filepath, delete_adjunto,
    crear_proyecto, eliminar_proyecto,
    get_oc_items, add_oc_item, update_oc_item, delete_oc_item,
    ESTADOS_KANBAN, ESTADO_LABELS, ADJUNTOS_DIR,
)
from web.changelog import VERSIONES as _CHANGELOG
from web.limits import limiter
from web.plantillas import templates, ctx, _permiso_usuario
from web.validators import validar_ruc

router = APIRouter()

# ─────────────────────────────────────────────

@router.get("/configuracion", response_class=HTMLResponse)
async def configuracion_page(request: Request, usuario: dict = Depends(require_admin)):
    config = cargar_config()
    catalogo = obtener_catalogo()
    return templates.TemplateResponse(
        "configuracion.html",
        ctx(request, usuario, config=config, catalogo=catalogo),
    )


@router.post("/configuracion/guardar")
async def guardar_configuracion(
    request: Request,
    usuario: dict = Depends(require_admin),
    ganancia: str = Form("30"),
    galvanizado: str = Form("GO"),
    espesor_producto: str = Form("1.5"),
    espesor_tapa: str = Form("1.5"),
    go_12: float = Form(150.0),
    go_15: float = Form(180.0),
    go_20: float = Form(220.0),
    gc_12: float = Form(140.0),
    gc_15: float = Form(170.0),
    gc_20: float = Form(210.0),
    dolar: float = Form(3.8),
    usd_kg_productos: float = Form(1.0),
    usd_kg_cajas: float = Form(3.0),
    # Factores de ganancia por producto — Sin comisión (30%)
    fg30_B: float = Form(0.70),
    fg30_CH: float = Form(0.50),
    fg30_CVE: float = Form(0.50),
    fg30_CVI: float = Form(0.50),
    fg30_T: float = Form(0.60),
    fg30_C: float = Form(0.70),
    fg30_R: float = Form(0.20),
    fg30_CP: float = Form(0.50),
    # Factores de ganancia por producto — Con comisión (35%)
    fg35_B: float = Form(0.65),
    fg35_CH: float = Form(0.45),
    fg35_CVE: float = Form(0.45),
    fg35_CVI: float = Form(0.45),
    fg35_T: float = Form(0.55),
    fg35_C: float = Form(0.65),
    fg35_R: float = Form(0.15),
    fg35_CP: float = Form(0.475),
):
    campos_precio = [go_12, go_15, go_20, gc_12, gc_15, gc_20, dolar, usd_kg_productos, usd_kg_cajas]
    if any(v < 0 for v in campos_precio):
        return JSONResponse({"ok": False, "error": "Los precios no pueden ser negativos"}, status_code=422)
    precios_plancha = [go_12, go_15, go_20, gc_12, gc_15, gc_20]
    if any(v == 0 for v in precios_plancha) or dolar == 0:
        return JSONResponse({"ok": False, "error": "Los precios de plancha y el tipo de cambio no pueden ser cero"}, status_code=422)
    factores = [fg30_B, fg30_CH, fg30_CVE, fg30_CVI, fg30_T, fg30_C, fg30_R, fg30_CP,
                fg35_B, fg35_CH, fg35_CVE, fg35_CVI, fg35_T, fg35_C, fg35_R, fg35_CP]
    if any(v <= 0 for v in factores):
        return JSONResponse({"ok": False, "error": "Los factores de ganancia deben ser mayores que cero"}, status_code=422)

    config = cargar_config()
    config["valores_defecto"].update({
        "ganancia": ganancia,
        "galvanizado": galvanizado,
        "espesor_producto": espesor_producto,
        "espesor_tapa": espesor_tapa,
        "precios_go": {"1.2": go_12, "1.5": go_15, "2.0": go_20},
        "precios_gc": {"1.2": gc_12, "1.5": gc_15, "2.0": gc_20},
        "dolar": dolar,
        "usd_kg_productos": usd_kg_productos,
        "usd_kg_cajas": usd_kg_cajas,
    })
    config["factores_ganancia"] = {
        "30": {"B": fg30_B, "CH": fg30_CH, "CVE": fg30_CVE, "CVI": fg30_CVI,
               "T": fg30_T, "C": fg30_C, "R": fg30_R, "CP": fg30_CP},
        "35": {"B": fg35_B, "CH": fg35_CH, "CVE": fg35_CVE, "CVI": fg35_CVI,
               "T": fg35_T, "C": fg35_C, "R": fg35_R, "CP": fg35_CP},
    }
    ok = guardar_config(config)
    return JSONResponse({"ok": ok})


@router.post("/configuracion/permisos")
async def guardar_permisos(
    request: Request,
    usuario: dict = Depends(require_admin),
):
    body = await request.json()
    config = cargar_config()
    config["permisos_usuario"] = {
        "ver_historial": bool(body.get("ver_historial", True)),
        "ver_catalogo":  bool(body.get("ver_catalogo", True)),
    }
    ok = guardar_config(config)
    return JSONResponse({"ok": ok})


@router.post("/configuracion/importar_excel")
async def importar_excel(
    request: Request,
    usuario: dict = Depends(require_admin),
    ruta_excel: str = Form(""),
):
    # Si no se indicó ruta, usar la plantilla configurada
    if not ruta_excel:
        config = cargar_config()
        ruta_excel = config.get("rutas", {}).get("plantilla_excel", "")
    # Fallback: buscar la plantilla en la carpeta local
    if not ruta_excel:
        from pathlib import Path as _P
        candidato = _P(__file__).parent.parent.parent / "plantillas" / "COTIZACIÓN v1.2 12-07-2023.xlsm"
        if candidato.exists():
            ruta_excel = str(candidato)
    try:
        conteo = importar_catalogo_desde_excel(ruta_excel)
        return JSONResponse({"ok": True, "conteo": conteo})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@router.post("/configuracion/guardar_plantilla")
async def guardar_plantilla(
    request: Request,
    usuario: dict = Depends(require_admin),
    plantilla_excel: str = Form(""),
):
    config = cargar_config()
    if "rutas" not in config:
        config["rutas"] = {}
    config["rutas"]["plantilla_excel"] = plantilla_excel.strip()
    ok = guardar_config(config)
    return JSONResponse({"ok": ok})


@router.post("/configuracion/cliente/agregar")
async def cfg_agregar_cliente(
    usuario: dict = Depends(require_admin),
    codigo: str = Form(...),
    nombre: str = Form(""),
    ruc: str = Form(""),
    ubicacion: str = Form(""),
    abreviacion: str = Form(""),
):
    err_ruc = validar_ruc(ruc)
    if err_ruc:
        return JSONResponse({"ok": False, "error": err_ruc}, status_code=422)
    ok = agregar_cliente(codigo, nombre, ruc, ubicacion, abreviacion)
    return JSONResponse({"ok": ok})


@router.post("/configuracion/atencion/agregar")
async def cfg_agregar_atencion(
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    codigo_empresa: str = Form(...),
    email: str = Form(""),
):
    ok = agregar_atencion(nombre, codigo_empresa, email)
    return JSONResponse({"ok": ok})


@router.post("/api/cliente/nuevo")
async def api_nuevo_cliente(
    usuario: dict = Depends(require_login),
    codigo: str = Form(...),
    nombre: str = Form(""),
    ruc: str = Form(""),
    ubicacion: str = Form(""),
    abreviacion: str = Form(""),
):
    err_ruc = validar_ruc(ruc)
    if err_ruc:
        return JSONResponse({"ok": False, "error": err_ruc}, status_code=422)
    # Si ya existe ese código, añadir sufijo numérico
    from web.database import obtener_catalogo as _cat
    codigos_existentes = {c["codigo"] for c in _cat()["clientes"]}
    cod = codigo
    if cod in codigos_existentes:
        i = 2
        while f"{cod}{i}" in codigos_existentes:
            i += 1
        cod = f"{cod}{i}"
    ok = agregar_cliente(cod, nombre, ruc, ubicacion, abreviacion)
    return JSONResponse({"ok": ok, "codigo": cod})


@router.post("/api/atencion/nueva")
async def api_nueva_atencion(
    usuario: dict = Depends(require_login),
    nombre: str = Form(...),
    codigo_empresa: str = Form(...),
    email: str = Form(""),
):
    ok = agregar_atencion(nombre, codigo_empresa, email)
    return JSONResponse({"ok": ok})


@router.put("/configuracion/cliente/{codigo}")
async def cfg_editar_cliente(
    codigo: str,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    ruc: str = Form(""),
    ubicacion: str = Form(""),
    abreviacion: str = Form(""),
):
    err_ruc = validar_ruc(ruc)
    if err_ruc:
        return JSONResponse({"ok": False, "error": err_ruc}, status_code=422)
    ok = editar_cliente(codigo, nombre, ruc, ubicacion, abreviacion)
    return JSONResponse({"ok": ok})


@router.put("/configuracion/atencion/{nombre_actual}")
async def cfg_editar_atencion(
    nombre_actual: str,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    codigo_empresa: str = Form(...),
    email: str = Form(""),
):
    ok = editar_atencion(nombre_actual, nombre, codigo_empresa, email)
    return JSONResponse({"ok": ok})


@router.delete("/configuracion/cliente/{codigo}")
async def cfg_eliminar_cliente(codigo: str, usuario: dict = Depends(require_admin)):
    eliminar_cliente(codigo)
    return JSONResponse({"ok": True})


@router.delete("/configuracion/atencion/{nombre}")
async def cfg_eliminar_atencion(nombre: str, usuario: dict = Depends(require_admin)):
    eliminar_atencion(nombre)
    return JSONResponse({"ok": True})


@router.get("/configuracion/catalogo", response_class=HTMLResponse)
async def configuracion_catalogo_page(request: Request, usuario: dict = Depends(require_admin)):
    ruta_json = Path(__file__).resolve().parent.parent.parent / "catalogo_productos.json"
    try:
        with open(ruta_json, encoding="utf-8") as f:
            datos = json.load(f)
    except Exception:
        datos = {"version": "1.0", "categorias": []}
    return templates.TemplateResponse(
        "cotizacion/configuracion_catalogo.html",
        ctx(request, usuario, catalogo=datos),
    )


def _leer_catalogo_json() -> dict:
    ruta = Path(__file__).resolve().parent.parent.parent / "catalogo_productos.json"
    if not ruta.exists():
        return {"version": "1.0", "categorias": []}
    with open(ruta, encoding="utf-8") as f:
        return json.load(f)


def _guardar_catalogo_json(datos: dict) -> bool:
    try:
        from datetime import date
        datos["ultima_actualizacion"] = date.today().isoformat()
        ruta = Path(__file__).resolve().parent.parent.parent / "catalogo_productos.json"
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


@router.post("/api/catalogo/categoria")
async def api_agregar_categoria(
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    icono: str = Form("📦"),
):
    datos = _leer_catalogo_json()
    if any(c["nombre"] == nombre for c in datos["categorias"]):
        return JSONResponse({"ok": False, "error": "La categoría ya existe"})
    datos["categorias"].append({"nombre": nombre, "icono": icono, "subcategorias": []})
    ok = _guardar_catalogo_json(datos)
    return JSONResponse({"ok": ok})


@router.put("/api/catalogo/categoria/{cat_idx}")
async def api_editar_categoria(
    cat_idx: int,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    icono: str = Form("📦"),
):
    datos = _leer_catalogo_json()
    if cat_idx < 0 or cat_idx >= len(datos["categorias"]):
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    datos["categorias"][cat_idx].update({"nombre": nombre, "icono": icono})
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.delete("/api/catalogo/categoria/{cat_idx}")
async def api_eliminar_categoria(cat_idx: int, usuario: dict = Depends(require_admin)):
    datos = _leer_catalogo_json()
    if cat_idx < 0 or cat_idx >= len(datos["categorias"]):
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    datos["categorias"].pop(cat_idx)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.post("/api/catalogo/subcategoria")
async def api_agregar_subcategoria(
    usuario: dict = Depends(require_admin),
    cat_idx: int = Form(...),
    nombre: str = Form(...),
    icono: str = Form("📋"),
):
    datos = _leer_catalogo_json()
    try:
        cat = datos["categorias"][cat_idx]
    except IndexError:
        return JSONResponse({"ok": False, "error": "Categoría no encontrada"}, status_code=404)
    if any(s["nombre"] == nombre for s in cat["subcategorias"]):
        return JSONResponse({"ok": False, "error": "La subcategoría ya existe"})
    cat["subcategorias"].append({"nombre": nombre, "icono": icono, "productos": []})
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.put("/api/catalogo/subcategoria/{cat_idx}/{sub_idx}")
async def api_editar_subcategoria(
    cat_idx: int, sub_idx: int,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
):
    datos = _leer_catalogo_json()
    try:
        datos["categorias"][cat_idx]["subcategorias"][sub_idx]["nombre"] = nombre
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.delete("/api/catalogo/subcategoria/{cat_idx}/{sub_idx}")
async def api_eliminar_subcategoria(
    cat_idx: int, sub_idx: int, usuario: dict = Depends(require_admin)
):
    datos = _leer_catalogo_json()
    try:
        datos["categorias"][cat_idx]["subcategorias"].pop(sub_idx)
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.post("/api/catalogo/producto")
async def api_agregar_producto(
    usuario: dict = Depends(require_admin),
    cat_idx: int = Form(...),
    sub_idx: int = Form(...),
    descripcion: str = Form(...),
    unidad: str = Form("UND"),
    precio: float = Form(...),
    presentacion: str = Form(""),
):
    datos = _leer_catalogo_json()
    try:
        sub = datos["categorias"][cat_idx]["subcategorias"][sub_idx]
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    sub["productos"].append({
        "descripcion": descripcion.strip(),
        "unidad": unidad.strip(),
        "precio": round(precio, 4),
        "presentacion": presentacion.strip(),
    })
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.put("/api/catalogo/producto/{cat_idx}/{sub_idx}/{prod_idx}")
async def api_editar_producto(
    cat_idx: int, sub_idx: int, prod_idx: int,
    usuario: dict = Depends(require_admin),
    descripcion: str = Form(...),
    unidad: str = Form("UND"),
    precio: float = Form(...),
    presentacion: str = Form(""),
):
    datos = _leer_catalogo_json()
    try:
        prod = datos["categorias"][cat_idx]["subcategorias"][sub_idx]["productos"][prod_idx]
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    prod.update({
        "descripcion": descripcion.strip(),
        "unidad": unidad.strip(),
        "precio": round(precio, 4),
        "presentacion": presentacion.strip(),
    })
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.delete("/api/catalogo/producto/{cat_idx}/{sub_idx}/{prod_idx}")
async def api_eliminar_producto(
    cat_idx: int, sub_idx: int, prod_idx: int,
    usuario: dict = Depends(require_admin),
):
    datos = _leer_catalogo_json()
    try:
        datos["categorias"][cat_idx]["subcategorias"][sub_idx]["productos"].pop(prod_idx)
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@router.get("/configuracion/clientes", response_class=HTMLResponse)
async def clientes_page(request: Request, usuario: dict = Depends(require_admin)):
    catalogo = obtener_catalogo()
    return templates.TemplateResponse(
        "cotizacion/clientes.html",
        ctx(request, usuario, catalogo=catalogo),
    )


@router.get("/api/configuracion/cliente/{codigo}")
async def api_get_cliente(codigo: str, usuario: dict = Depends(require_login)):
    cliente = obtener_cliente(codigo)
    if not cliente:
        raise HTTPException(404, "Cliente no encontrado")
    atenciones = obtener_atenciones_de_cliente(codigo)
    return JSONResponse({"ok": True, "cliente": cliente, "atenciones": atenciones})


@router.post("/configuracion/importar_json")
async def importar_json_endpoint(
    usuario: dict = Depends(require_admin),
    ruta_json: str = Form(""),
):
    if not ruta_json:
        ruta_json = str(Path(__file__).resolve().parent.parent / "data" / "catalogo_contactos.json")
    try:
        conteo = importar_catalogo_desde_json(ruta_json)
        return JSONResponse({"ok": True, "conteo": conteo})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@router.get("/configuracion/contactos/exportar")
async def exportar_contactos(usuario: dict = Depends(require_admin)):
    """Descarga un .xlsx con todas las hojas de clientes, atenciones y monedas."""
    contenido = exportar_contactos_xlsx()
    headers = {
        "Content-Disposition": 'attachment; filename="contactos_aroluz.xlsx"',
    }
    return StreamingResponse(
        iter([contenido]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@router.post("/configuracion/contactos/importar")
async def importar_contactos(
    usuario: dict = Depends(require_admin),
    archivo: UploadFile = File(...),
):
    """Importa/actualiza clientes, atenciones y monedas desde un .xlsx subido."""
    try:
        contenido = await archivo.read()
        conteo = importar_contactos_desde_xlsx(contenido)
        return JSONResponse({
            "ok": True,
            "conteo": conteo,
            "mensaje": (
                f"Importado: {conteo['clientes']} clientes, "
                f"{conteo['atenciones']} atenciones, "
                f"{conteo['monedas']} monedas"
            ),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@router.get("/configuracion/catalogo/descargar")
async def descargar_catalogo_xlsx(usuario: dict = Depends(require_admin)):
    """Descarga el catálogo de productos como Excel (.xlsx) editable."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ruta_json = Path(__file__).resolve().parent.parent.parent / "catalogo_productos.json"
    if not ruta_json.exists():
        raise HTTPException(404, "catalogo_productos.json no encontrado")
    with open(ruta_json, encoding="utf-8") as f:
        datos = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo"

    # Encabezados
    headers = ["Categoría", "Subcategoría", "Descripción", "Precio (S/)", "Unidad", "Presentación"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas de datos
    row = 2
    cat_fill = PatternFill("solid", fgColor="D6E4F0")
    sub_fill = PatternFill("solid", fgColor="EBF5FB")
    for cat in datos.get("categorias", []):
        for sub in cat.get("subcategorias", []):
            for prod in sub.get("productos", []):
                ws.cell(row=row, column=1, value=cat["nombre"]).fill = cat_fill
                ws.cell(row=row, column=2, value=sub["nombre"]).fill = sub_fill
                ws.cell(row=row, column=3, value=prod.get("descripcion", ""))
                ws.cell(row=row, column=4, value=prod.get("precio", 0))
                ws.cell(row=row, column=5, value=prod.get("unidad", "UND"))
                ws.cell(row=row, column=6, value=prod.get("presentacion", ""))
                row += 1

    # Anchos de columna
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="catalogo_productos.xlsx"'},
    )


@router.post("/configuracion/catalogo/subir")
async def subir_catalogo_xlsx(
    usuario: dict = Depends(require_admin),
    archivo: UploadFile = File(...),
):
    """Importa el catálogo desde un .xlsx y reconstruye catalogo_productos.json."""
    import io
    import openpyxl

    try:
        contenido = await archivo.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active

        # Leer filas (saltar encabezado en fila 1)
        categorias: dict = {}  # nombre_cat -> dict con subcategorias
        n_prod = 0
        n_invalidos = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre_cat, nombre_sub, desc, precio, unidad, presentacion = (
                (row[i] if i < len(row) else None) for i in range(6)
            )
            if not nombre_cat or not nombre_sub or not desc:
                continue  # fila vacía o incompleta
            nombre_cat = str(nombre_cat).strip()
            nombre_sub = str(nombre_sub).strip()
            desc       = str(desc).strip()
            try:
                precio = float(precio) if precio is not None else 0.0
            except (ValueError, TypeError):
                precio = 0.0
                n_invalidos += 1
            unidad     = str(unidad).strip() if unidad else "UND"
            presentacion = str(presentacion).strip() if presentacion else ""

            if nombre_cat not in categorias:
                categorias[nombre_cat] = {"nombre": nombre_cat, "icono": "📦", "subcategorias": {}}
            cat = categorias[nombre_cat]
            if nombre_sub not in cat["subcategorias"]:
                cat["subcategorias"][nombre_sub] = {"nombre": nombre_sub, "icono": "📋", "productos": []}
            cat["subcategorias"][nombre_sub]["productos"].append({
                "descripcion": desc,
                "unidad": unidad,
                "precio": round(precio, 4),
                "presentacion": presentacion,
            })
            n_prod += 1

        if n_prod == 0:
            return JSONResponse({"ok": False, "error": "No se encontraron productos en el Excel"}, status_code=400)

        # Convertir dicts anidados a listas
        resultado = {
            "version": "1.0",
            "ultima_actualizacion": __import__("datetime").date.today().isoformat(),
            "categorias": [
                {
                    "nombre": cat["nombre"],
                    "icono": cat["icono"],
                    "subcategorias": [
                        sub for sub in cat["subcategorias"].values()
                    ],
                }
                for cat in categorias.values()
            ],
        }

        ruta_json = Path(__file__).resolve().parent.parent.parent / "catalogo_productos.json"
        with open(ruta_json, "w", encoding="utf-8") as f:
            json.dump(resultado, f, ensure_ascii=False, indent=2)

        n_cat = len(resultado["categorias"])
        resp: dict = {
            "ok": True,
            "mensaje": f"Catálogo actualizado: {n_cat} categorías, {n_prod} productos",
        }
        if n_invalidos:
            resp["advertencias"] = f"{n_invalidos} fila(s) con precio inválido (se usó 0.0)"
        return JSONResponse(resp)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@router.post("/configuracion/usuario/crear")
@limiter.limit("10/minute")
async def cfg_crear_usuario(
    request: Request,
    usuario: dict = Depends(require_admin),
    username: str = Form(...),
    password: str = Form(...),
    nombre: str = Form(""),
    rol: str = Form("USER"),
    ver_asistencias: str = Form(""),
):
    if rol not in ("ADMIN", "USER"):
        rol = "USER"
    ok = crear_usuario(username, password, nombre, rol, ver_asistencias=bool(ver_asistencias))
    return JSONResponse({"ok": ok, "error": "El usuario ya existe" if not ok else None})


@router.post("/configuracion/usuario/cambiar_password")
@limiter.limit("5/minute")
async def cfg_cambiar_password(
    request: Request,
    usuario: dict = Depends(require_admin),
    password_actual: str = Form(...),
    password_nuevo: str = Form(...),
):
    from web.database import verificar_usuario as ver
    if not ver(usuario["u"], password_actual):
        return JSONResponse({"ok": False, "error": "Contraseña actual incorrecta"}, status_code=401)
    ok = cambiar_password(usuario["u"], password_nuevo)
    return JSONResponse({"ok": ok})


@router.put("/configuracion/usuario/{username}")
async def cfg_editar_usuario(
    username: str,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(""),
    rol: str = Form("USER"),
    ver_asistencias: str = Form(""),
):
    if rol not in ("ADMIN", "USER"):
        rol = "USER"
    ok = editar_usuario(username, nombre, rol, ver_asistencias=bool(ver_asistencias))
    return JSONResponse({"ok": ok})


@router.post("/configuracion/usuario/{username}/toggle_activo")
async def cfg_toggle_activo(username: str, usuario: dict = Depends(require_admin)):
    if username == usuario["u"]:
        return JSONResponse({"ok": False, "error": "No puedes desactivar tu propia cuenta"}, status_code=422)
    result = toggle_activo_usuario(username)
    return JSONResponse(result)


@router.delete("/configuracion/usuario/{username}")
async def cfg_eliminar_usuario(username: str, usuario: dict = Depends(require_admin)):
    if username == usuario["u"]:
        return JSONResponse({"ok": False, "error": "No puedes eliminar tu propia cuenta"}, status_code=422)
    ok = eliminar_usuario(username)
    return JSONResponse({"ok": ok})


@router.post("/configuracion/usuario/{username}/reset_password")
@limiter.limit("5/minute")
async def cfg_reset_password(
    request: Request,
    username: str,
    usuario: dict = Depends(require_admin),
    password_nuevo: str = Form(...),
):
    if not password_nuevo or len(password_nuevo) < 8:
        return JSONResponse({"ok": False, "error": "La contraseña debe tener al menos 8 caracteres"}, status_code=422)
    ok = cambiar_password(username, password_nuevo)
    return JSONResponse({"ok": ok})

