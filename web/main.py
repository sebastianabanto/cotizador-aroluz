"""
main.py — Aplicación FastAPI principal de AROLUZ Cotizador Web

Incluye:
- Páginas HTML (login, cotización, carrito, catálogo, configuración)
- API endpoints (cotizar, carrito, exportar)
- Autenticación con sesiones firmadas
"""
import json
import os
import sys
from pathlib import Path

# Asegurar que el directorio raíz del proyecto esté en el path
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from fastapi import FastAPI, Request, Depends, Form, HTTPException, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

from web.auth import require_login, require_admin, require_asistencias, get_session, verificar_usuario, set_session_cookie, clear_session_cookie
from web.database import (
    cargar_config, guardar_config, obtener_catalogo,
    importar_catalogo_desde_excel, importar_catalogo_desde_json,
    exportar_contactos_xlsx, importar_contactos_desde_xlsx,
    agregar_cliente, agregar_atencion,
    eliminar_cliente, eliminar_atencion,
    editar_cliente, editar_atencion,
    obtener_cliente, obtener_atenciones_de_cliente,
    crear_usuario, cambiar_password, listar_usuarios,
    editar_usuario, toggle_activo_usuario, eliminar_usuario,
    get_kpis_proyectos, get_proyectos_con_stats, set_proyecto_estado,
    update_proyecto_direccion, update_proyecto_numero_oc, update_proyecto_contacto, renombrar_proyecto,
    add_adjunto, list_adjuntos, get_adjunto_filepath, delete_adjunto,
    crear_proyecto, eliminar_proyecto,
    get_oc_items, add_oc_item, update_oc_item, delete_oc_item,
    ESTADOS_KANBAN, ESTADO_LABELS, ADJUNTOS_DIR,
)
from web.rutas import cotizar as rutas_cotizar
from web.rutas import carrito as rutas_carrito
from web.rutas import exportar as rutas_exportar
from web.rutas import historial as rutas_historial
from web.rutas.carrito import get_carrito
from asistencias.main import router as asistencias_router

# ─────────────────────────────────────────────
# Configurar app
# ─────────────────────────────────────────────

limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="AROLUZ Cotizador Web",
    description="Cotizador de bandejas porta cables AROLUZ",
    version="2.0.0",
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# Templates y archivos estáticos
TEMPLATES_DIR = Path(__file__).parent / "templates"
STATIC_DIR = Path(__file__).parent / "static"

templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

# Filtros Jinja2 adicionales
from urllib.parse import quote as _url_quote
import json as _json
from markupsafe import Markup as _Markup
templates.env.filters["urlencode_str"] = lambda s: _url_quote(str(s), safe="")
templates.env.filters.setdefault("tojson", lambda o: _Markup(_json.dumps(o, ensure_ascii=False)))

if STATIC_DIR.exists():
    app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

# Incluir routers de API
app.include_router(rutas_cotizar.router)
app.include_router(rutas_carrito.router)
app.include_router(rutas_exportar.router)
app.include_router(rutas_historial.router)
app.include_router(asistencias_router, prefix="/asistencias")

# Static de asistencias (CSS/imágenes propias del módulo)
_ASISTENCIAS_STATIC = Path(__file__).parent.parent / "asistencias" / "static"
if _ASISTENCIAS_STATIC.exists():
    app.mount("/asistencias/static",
              StaticFiles(directory=str(_ASISTENCIAS_STATIC)),
              name="asistencias_static")


# ─────────────────────────────────────────────
# Helper para contexto base de templates
# ─────────────────────────────────────────────

def ctx(request: Request, usuario: dict, **kwargs) -> dict:
    """Contexto base para todos los templates."""
    carrito = get_carrito(usuario["u"])
    total_carrito = sum(i["precio_unitario"] * i["cantidad"] for i in carrito)
    return {
        "request": request,
        "usuario": usuario,
        "n_carrito": len(carrito),
        "total_carrito": round(total_carrito, 2),
        **kwargs,
    }


# ─────────────────────────────────────────────
# Login / Logout
# ─────────────────────────────────────────────

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    session = get_session(request)
    if session:
        return RedirectResponse("/home", status_code=302)
    return templates.TemplateResponse("login.html", {"request": request, "error": None})


@app.post("/login")
@limiter.limit("5/minute")
async def login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
):
    user = verificar_usuario(username, password)
    if not user:
        return templates.TemplateResponse(
            "login.html",
            {"request": request, "error": "Usuario o contraseña incorrectos"},
            status_code=401,
        )
    response = RedirectResponse("/home", status_code=302)
    set_session_cookie(response, user["username"], user.get("nombre", ""), user.get("rol", "USER"), user.get("ver_asistencias", False))
    return response


@app.get("/logout")
async def logout(request: Request):
    response = RedirectResponse("/login", status_code=302)
    clear_session_cookie(response)
    return response


# ─────────────────────────────────────────────
# Redirect raíz
# ─────────────────────────────────────────────

@app.get("/")
async def root(request: Request):
    session = get_session(request)
    if session:
        return RedirectResponse("/home", status_code=302)
    return RedirectResponse("/login", status_code=302)


# ─────────────────────────────────────────────
# Página Home
# ─────────────────────────────────────────────

@app.get("/home", response_class=HTMLResponse)
async def home_page(request: Request, usuario: dict = Depends(require_login)):
    kpis = get_kpis_proyectos()
    proyectos = get_proyectos_con_stats()
    catalogo = obtener_catalogo()
    return templates.TemplateResponse(
        "home.html",
        ctx(
            request, usuario,
            active="home",
            kpis=kpis,
            proyectos=proyectos,
            ESTADOS_KANBAN=ESTADOS_KANBAN,
            ESTADO_LABELS=ESTADO_LABELS,
            catalogo=catalogo,
        ),
    )


# ─────────────────────────────────────────────
# API — Proyectos: estado, dirección, adjuntos
# ─────────────────────────────────────────────

@app.put("/api/proyecto/{nombre}/estado")
async def api_set_proyecto_estado(
    nombre: str,
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
        estado = body.get("estado", "")
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    if estado not in ESTADOS_KANBAN:
        return JSONResponse({"ok": False, "error": "Estado inválido"}, status_code=422)
    set_proyecto_estado(nombre, estado)
    return JSONResponse({"ok": True})


@app.patch("/api/proyecto/{nombre}/numero-oc")
async def api_update_numero_oc(
    nombre: str,
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
        numero_oc = body.get("numero_oc", "")
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    update_proyecto_numero_oc(nombre, numero_oc)
    return JSONResponse({"ok": True})


@app.patch("/api/proyecto/{nombre}/contacto")
async def api_update_contacto(
    nombre: str,
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
        contacto = body.get("contacto", "")
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    update_proyecto_contacto(nombre, contacto)
    return JSONResponse({"ok": True})


@app.patch("/api/proyecto/{nombre}/info")
async def api_update_proyecto_info(
    nombre: str,
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
        nuevo_nombre = body.get("nuevo_nombre", nombre)
        nuevo_cliente = body.get("nuevo_cliente", "")
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    ok = renombrar_proyecto(nombre, nuevo_nombre, nuevo_cliente)
    if not ok:
        return JSONResponse({"ok": False, "error": "Nombre duplicado o vacío"}, status_code=409)
    return JSONResponse({"ok": True, "nuevo_nombre": nuevo_nombre.strip()})


@app.patch("/api/proyecto/{nombre}/direccion")
async def api_update_direccion(
    nombre: str,
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
        direccion = body.get("direccion", "")
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    update_proyecto_direccion(nombre, direccion)
    return JSONResponse({"ok": True})


@app.get("/api/proyecto/{nombre}/adjuntos")
async def api_list_adjuntos(nombre: str, usuario: dict = Depends(require_login)):
    return JSONResponse({"ok": True, "adjuntos": list_adjuntos(nombre)})


@app.post("/api/proyecto/{nombre}/adjunto")
async def api_upload_adjunto(
    nombre: str,
    usuario: dict = Depends(require_login),
    archivo: UploadFile = File(...),
    categoria: str = Form("oc"),
):
    import shutil as _shutil
    ext = Path(archivo.filename or "").suffix.lower()
    if ext not in {".pdf", ".jpg", ".jpeg", ".png", ".webp"}:
        return JSONResponse(
            {"ok": False, "error": "Tipo no permitido. Use PDF, JPG o PNG."},
            status_code=422,
        )
    cat = categoria if categoria in ("oc", "ev") else "oc"
    slug = "".join(c if c.isalnum() or c in "-_. " else "_" for c in nombre)[:60]
    dest_dir = ADJUNTOS_DIR / slug
    dest_dir.mkdir(parents=True, exist_ok=True)

    from datetime import datetime as _dt2
    ts = _dt2.now().strftime("%Y%m%d_%H%M%S")
    safe_name = f"{ts}_{Path(archivo.filename or 'archivo').name}"
    dest = dest_dir / safe_name
    with dest.open("wb") as f:
        _shutil.copyfileobj(archivo.file, f)

    adj_id = add_adjunto(nombre, archivo.filename or safe_name, str(dest), archivo.content_type or "", cat)
    return JSONResponse({"ok": True, "id": adj_id, "filename": archivo.filename})


@app.get("/api/proyecto/{nombre}/adjunto/{adj_id}/descargar")
async def api_descargar_adjunto(
    nombre: str, adj_id: int, usuario: dict = Depends(require_login),
):
    info = get_adjunto_filepath(adj_id, nombre)
    if not info:
        raise HTTPException(404, "Adjunto no encontrado")
    p = Path(info["filepath"])
    if not p.exists():
        raise HTTPException(404, "Archivo no encontrado en disco")
    return StreamingResponse(
        open(p, "rb"),
        media_type=info.get("content_type") or "application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{info["filename"]}"'},
    )


@app.get("/api/proyecto/{nombre}/adjunto/{adj_id}/ver")
async def api_ver_adjunto_inline(
    nombre: str, adj_id: int, usuario: dict = Depends(require_login),
):
    """Sirve el adjunto inline (para visualizar en iframe sin forzar descarga)."""
    info = get_adjunto_filepath(adj_id, nombre)
    if not info:
        raise HTTPException(404, "Adjunto no encontrado")
    p = Path(info["filepath"])
    if not p.exists():
        raise HTTPException(404, "Archivo no encontrado en disco")
    return StreamingResponse(
        open(p, "rb"),
        media_type=info.get("content_type") or "application/pdf",
        headers={"Content-Disposition": f'inline; filename="{info["filename"]}"'},
    )


@app.delete("/api/proyecto/{nombre}/adjunto/{adj_id}")
async def api_delete_adjunto(
    nombre: str, adj_id: int, usuario: dict = Depends(require_login),
):
    filepath = delete_adjunto(adj_id)
    if filepath:
        try:
            Path(filepath).unlink(missing_ok=True)
        except Exception:
            pass
    return JSONResponse({"ok": True})


# ─────────────────────────────────────────────
# API — Proyectos: crear / eliminar (manual)
# ─────────────────────────────────────────────

@app.post("/api/proyecto")
async def api_crear_proyecto(
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
        nombre = (body.get("nombre") or "").strip()
        cliente = (body.get("cliente") or "").strip()
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    if not nombre:
        return JSONResponse({"ok": False, "error": "El nombre es requerido"}, status_code=422)
    ok = crear_proyecto(nombre, cliente)
    if not ok:
        return JSONResponse({"ok": False, "error": "Ya existe un proyecto con ese nombre"}, status_code=409)
    return JSONResponse({"ok": True})


@app.delete("/api/proyecto/{nombre}")
async def api_eliminar_proyecto(
    nombre: str,
    usuario: dict = Depends(require_admin),
):
    eliminar_proyecto(nombre)
    return JSONResponse({"ok": True})


# ─────────────────────────────────────────────
# API — OC Items
# ─────────────────────────────────────────────

@app.get("/api/proyecto/{nombre}/oc-items")
async def api_get_oc_items(nombre: str, usuario: dict = Depends(require_login)):
    try:
        items = get_oc_items(nombre)
        return JSONResponse({"ok": True, "items": items})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e), "items": []}, status_code=200)


@app.post("/api/proyecto/{nombre}/oc-items")
async def api_add_oc_item(
    nombre: str,
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    item_id = add_oc_item(
        proyecto=nombre,
        descripcion=body.get("descripcion", ""),
        unidad=body.get("unidad", "UND"),
        cantidad_pedida=float(body.get("cantidad_pedida", 0)),
        cantidad_despachada=float(body.get("cantidad_despachada", 0)),
        orden=int(body.get("orden", 0)),
    )
    return JSONResponse({"ok": True, "id": item_id})


@app.put("/api/proyecto/{nombre}/oc-item/{item_id}")
async def api_update_oc_item(
    nombre: str,
    item_id: int,
    request: Request,
    usuario: dict = Depends(require_login),
):
    try:
        body = await request.json()
    except Exception:
        return JSONResponse({"ok": False, "error": "JSON inválido"}, status_code=400)
    ok = update_oc_item(
        item_id=item_id,
        proyecto=nombre,
        descripcion=body.get("descripcion", ""),
        unidad=body.get("unidad", "UND"),
        cantidad_pedida=float(body.get("cantidad_pedida", 0)),
        cantidad_despachada=float(body.get("cantidad_despachada", 0)),
    )
    return JSONResponse({"ok": ok})


@app.delete("/api/proyecto/{nombre}/oc-item/{item_id}")
async def api_delete_oc_item(
    nombre: str,
    item_id: int,
    usuario: dict = Depends(require_login),
):
    ok = delete_oc_item(item_id, nombre)
    return JSONResponse({"ok": ok})


# ─────────────────────────────────────────────
# Página Cotización
# ─────────────────────────────────────────────

@app.get("/cotizar", response_class=HTMLResponse)
async def cotizar_page(request: Request, usuario: dict = Depends(require_login)):
    config = cargar_config()
    valores = config.get("valores_defecto", {})
    return templates.TemplateResponse(
        "cotizacion.html",
        ctx(request, usuario, config=valores),
    )


# ─────────────────────────────────────────────
# Página Carrito
# ─────────────────────────────────────────────

@app.get("/carrito", response_class=HTMLResponse)
async def carrito_page(request: Request, usuario: dict = Depends(require_login)):
    carrito = get_carrito(usuario["u"])
    catalogo = obtener_catalogo()
    config = cargar_config()
    total = sum(i["precio_unitario"] * i["cantidad"] for i in carrito)
    peso_total = sum(i["peso_unitario"] * i["cantidad"] for i in carrito)

    dolar = config.get("valores_defecto", {}).get("dolar", 3.8)
    return templates.TemplateResponse(
        "carrito.html",
        ctx(
            request,
            usuario,
            carrito=carrito,
            total=round(total, 2),
            peso_total=round(peso_total, 4),
            catalogo=catalogo,
            dolar=dolar,
        ),
    )


# ─────────────────────────────────────────────
# Página Historial
# ─────────────────────────────────────────────

@app.get("/historial", response_class=HTMLResponse)
async def historial_page(request: Request, usuario: dict = Depends(require_login)):
    es_admin = usuario.get("r") == "ADMIN"
    return templates.TemplateResponse(
        "historial.html",
        ctx(request, usuario, es_admin=es_admin),
    )


# ─────────────────────────────────────────────
# Página Catálogo
# ─────────────────────────────────────────────

@app.get("/catalogo", response_class=HTMLResponse)
async def catalogo_page(request: Request, usuario: dict = Depends(require_login)):
    ruta_json = Path(__file__).resolve().parent.parent / "catalogo_productos.json"
    try:
        with open(ruta_json, encoding="utf-8") as f:
            catalogo_productos = json.load(f)
    except Exception:
        catalogo_productos = {"categorias": []}
    return templates.TemplateResponse(
        "catalogo.html",
        ctx(request, usuario, catalogo_productos=catalogo_productos),
    )


# ─────────────────────────────────────────────
# Página Configuración
# ─────────────────────────────────────────────

@app.get("/configuracion", response_class=HTMLResponse)
async def configuracion_page(request: Request, usuario: dict = Depends(require_admin)):
    config = cargar_config()
    catalogo = obtener_catalogo()
    return templates.TemplateResponse(
        "configuracion.html",
        ctx(request, usuario, config=config, catalogo=catalogo),
    )


@app.post("/configuracion/guardar")
async def guardar_configuracion(
    request: Request,
    usuario: dict = Depends(require_admin),
    ganancia: str = Form("30"),
    galvanizado: str = Form("GO"),
    espesor_producto: str = Form("1.5"),
    espesor_tapa: str = Form("1.2"),
    go_12: float = Form(150.0),
    go_15: float = Form(180.0),
    go_20: float = Form(220.0),
    gc_12: float = Form(140.0),
    gc_15: float = Form(170.0),
    gc_20: float = Form(210.0),
    dolar: float = Form(3.8),
    usd_kg_productos: float = Form(1.0),
    usd_kg_cajas: float = Form(3.0),
):
    campos_precio = [go_12, go_15, go_20, gc_12, gc_15, gc_20, dolar, usd_kg_productos, usd_kg_cajas]
    if any(v < 0 for v in campos_precio):
        return JSONResponse({"ok": False, "error": "Los precios no pueden ser negativos"}, status_code=422)
    precios_plancha = [go_12, go_15, go_20, gc_12, gc_15, gc_20]
    if any(v == 0 for v in precios_plancha) or dolar == 0:
        return JSONResponse({"ok": False, "error": "Los precios de plancha y el tipo de cambio no pueden ser cero"}, status_code=422)

    config = cargar_config()
    config["valores_defecto"].update({
        "ganancia": ganancia,
        "galvanizado": galvanizado,
        "espesor_producto": espesor_producto,
        "espesor_tapa": espesor_tapa,
        "precios_go": {"1.2": go_12, "1.5": go_15, "2.0": go_20},
        "precios_gc": {"1.2": gc_12, "1.5": gc_15, "2.0": gc_20},
        "dolar": dolar,
        "usd_kg_productos": usd_kg_productos,
        "usd_kg_cajas": usd_kg_cajas,
    })
    ok = guardar_config(config)
    return JSONResponse({"ok": ok})


@app.post("/configuracion/importar_excel")
async def importar_excel(
    request: Request,
    usuario: dict = Depends(require_admin),
    ruta_excel: str = Form(""),
):
    # Si no se indicó ruta, usar la plantilla configurada
    if not ruta_excel:
        config = cargar_config()
        ruta_excel = config.get("rutas", {}).get("plantilla_excel", "")
    # Fallback: buscar la plantilla en la carpeta local
    if not ruta_excel:
        from pathlib import Path as _P
        candidato = _P(__file__).parent.parent / "plantillas" / "COTIZACIÓN v1.2 12-07-2023.xlsm"
        if candidato.exists():
            ruta_excel = str(candidato)
    try:
        conteo = importar_catalogo_desde_excel(ruta_excel)
        return JSONResponse({"ok": True, "conteo": conteo})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@app.post("/configuracion/guardar_plantilla")
async def guardar_plantilla(
    request: Request,
    usuario: dict = Depends(require_admin),
    plantilla_excel: str = Form(""),
):
    config = cargar_config()
    if "rutas" not in config:
        config["rutas"] = {}
    config["rutas"]["plantilla_excel"] = plantilla_excel.strip()
    ok = guardar_config(config)
    return JSONResponse({"ok": ok})


@app.post("/configuracion/cliente/agregar")
async def cfg_agregar_cliente(
    usuario: dict = Depends(require_admin),
    codigo: str = Form(...),
    nombre: str = Form(""),
    ruc: str = Form(""),
    ubicacion: str = Form(""),
):
    import re as _re
    if ruc and not _re.fullmatch(r'\d{11}', ruc):
        return JSONResponse({"ok": False, "error": "El RUC debe tener exactamente 11 dígitos numéricos"}, status_code=422)
    ok = agregar_cliente(codigo, nombre, ruc, ubicacion)
    return JSONResponse({"ok": ok})


@app.post("/configuracion/atencion/agregar")
async def cfg_agregar_atencion(
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    codigo_empresa: str = Form(...),
    email: str = Form(""),
):
    ok = agregar_atencion(nombre, codigo_empresa, email)
    return JSONResponse({"ok": ok})


@app.put("/configuracion/cliente/{codigo}")
async def cfg_editar_cliente(
    codigo: str,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    ruc: str = Form(""),
    ubicacion: str = Form(""),
):
    import re as _re
    if ruc and not _re.fullmatch(r'\d{11}', ruc):
        return JSONResponse({"ok": False, "error": "El RUC debe tener exactamente 11 dígitos numéricos"}, status_code=422)
    ok = editar_cliente(codigo, nombre, ruc, ubicacion)
    return JSONResponse({"ok": ok})


@app.put("/configuracion/atencion/{nombre_actual}")
async def cfg_editar_atencion(
    nombre_actual: str,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    codigo_empresa: str = Form(...),
    email: str = Form(""),
):
    ok = editar_atencion(nombre_actual, nombre, codigo_empresa, email)
    return JSONResponse({"ok": ok})


@app.delete("/configuracion/cliente/{codigo}")
async def cfg_eliminar_cliente(codigo: str, usuario: dict = Depends(require_admin)):
    eliminar_cliente(codigo)
    return JSONResponse({"ok": True})


@app.delete("/configuracion/atencion/{nombre}")
async def cfg_eliminar_atencion(nombre: str, usuario: dict = Depends(require_admin)):
    eliminar_atencion(nombre)
    return JSONResponse({"ok": True})


@app.get("/configuracion/catalogo", response_class=HTMLResponse)
async def configuracion_catalogo_page(request: Request, usuario: dict = Depends(require_admin)):
    ruta_json = Path(__file__).resolve().parent.parent / "catalogo_productos.json"
    try:
        with open(ruta_json, encoding="utf-8") as f:
            datos = json.load(f)
    except Exception:
        datos = {"version": "1.0", "categorias": []}
    return templates.TemplateResponse(
        "configuracion_catalogo.html",
        ctx(request, usuario, catalogo=datos),
    )


def _leer_catalogo_json() -> dict:
    ruta = Path(__file__).resolve().parent.parent / "catalogo_productos.json"
    if not ruta.exists():
        return {"version": "1.0", "categorias": []}
    with open(ruta, encoding="utf-8") as f:
        return json.load(f)


def _guardar_catalogo_json(datos: dict) -> bool:
    try:
        from datetime import date
        datos["ultima_actualizacion"] = date.today().isoformat()
        ruta = Path(__file__).resolve().parent.parent / "catalogo_productos.json"
        with open(ruta, "w", encoding="utf-8") as f:
            json.dump(datos, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False


@app.post("/api/catalogo/categoria")
async def api_agregar_categoria(
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    icono: str = Form("📦"),
):
    datos = _leer_catalogo_json()
    if any(c["nombre"] == nombre for c in datos["categorias"]):
        return JSONResponse({"ok": False, "error": "La categoría ya existe"})
    datos["categorias"].append({"nombre": nombre, "icono": icono, "subcategorias": []})
    ok = _guardar_catalogo_json(datos)
    return JSONResponse({"ok": ok})


@app.put("/api/catalogo/categoria/{cat_idx}")
async def api_editar_categoria(
    cat_idx: int,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
    icono: str = Form("📦"),
):
    datos = _leer_catalogo_json()
    if cat_idx < 0 or cat_idx >= len(datos["categorias"]):
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    datos["categorias"][cat_idx].update({"nombre": nombre, "icono": icono})
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.delete("/api/catalogo/categoria/{cat_idx}")
async def api_eliminar_categoria(cat_idx: int, usuario: dict = Depends(require_admin)):
    datos = _leer_catalogo_json()
    if cat_idx < 0 or cat_idx >= len(datos["categorias"]):
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    datos["categorias"].pop(cat_idx)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.post("/api/catalogo/subcategoria")
async def api_agregar_subcategoria(
    usuario: dict = Depends(require_admin),
    cat_idx: int = Form(...),
    nombre: str = Form(...),
    icono: str = Form("📋"),
):
    datos = _leer_catalogo_json()
    try:
        cat = datos["categorias"][cat_idx]
    except IndexError:
        return JSONResponse({"ok": False, "error": "Categoría no encontrada"}, status_code=404)
    if any(s["nombre"] == nombre for s in cat["subcategorias"]):
        return JSONResponse({"ok": False, "error": "La subcategoría ya existe"})
    cat["subcategorias"].append({"nombre": nombre, "icono": icono, "productos": []})
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.put("/api/catalogo/subcategoria/{cat_idx}/{sub_idx}")
async def api_editar_subcategoria(
    cat_idx: int, sub_idx: int,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(...),
):
    datos = _leer_catalogo_json()
    try:
        datos["categorias"][cat_idx]["subcategorias"][sub_idx]["nombre"] = nombre
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.delete("/api/catalogo/subcategoria/{cat_idx}/{sub_idx}")
async def api_eliminar_subcategoria(
    cat_idx: int, sub_idx: int, usuario: dict = Depends(require_admin)
):
    datos = _leer_catalogo_json()
    try:
        datos["categorias"][cat_idx]["subcategorias"].pop(sub_idx)
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.post("/api/catalogo/producto")
async def api_agregar_producto(
    usuario: dict = Depends(require_admin),
    cat_idx: int = Form(...),
    sub_idx: int = Form(...),
    descripcion: str = Form(...),
    unidad: str = Form("UND"),
    precio: float = Form(...),
    presentacion: str = Form(""),
):
    datos = _leer_catalogo_json()
    try:
        sub = datos["categorias"][cat_idx]["subcategorias"][sub_idx]
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    sub["productos"].append({
        "descripcion": descripcion.strip(),
        "unidad": unidad.strip(),
        "precio": round(precio, 4),
        "presentacion": presentacion.strip(),
    })
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.put("/api/catalogo/producto/{cat_idx}/{sub_idx}/{prod_idx}")
async def api_editar_producto(
    cat_idx: int, sub_idx: int, prod_idx: int,
    usuario: dict = Depends(require_admin),
    descripcion: str = Form(...),
    unidad: str = Form("UND"),
    precio: float = Form(...),
    presentacion: str = Form(""),
):
    datos = _leer_catalogo_json()
    try:
        prod = datos["categorias"][cat_idx]["subcategorias"][sub_idx]["productos"][prod_idx]
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    prod.update({
        "descripcion": descripcion.strip(),
        "unidad": unidad.strip(),
        "precio": round(precio, 4),
        "presentacion": presentacion.strip(),
    })
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.delete("/api/catalogo/producto/{cat_idx}/{sub_idx}/{prod_idx}")
async def api_eliminar_producto(
    cat_idx: int, sub_idx: int, prod_idx: int,
    usuario: dict = Depends(require_admin),
):
    datos = _leer_catalogo_json()
    try:
        datos["categorias"][cat_idx]["subcategorias"][sub_idx]["productos"].pop(prod_idx)
    except IndexError:
        return JSONResponse({"ok": False, "error": "No encontrado"}, status_code=404)
    return JSONResponse({"ok": _guardar_catalogo_json(datos)})


@app.get("/configuracion/clientes", response_class=HTMLResponse)
async def clientes_page(request: Request, usuario: dict = Depends(require_admin)):
    catalogo = obtener_catalogo()
    return templates.TemplateResponse(
        "clientes.html",
        ctx(request, usuario, catalogo=catalogo),
    )


@app.get("/api/configuracion/cliente/{codigo}")
async def api_get_cliente(codigo: str, usuario: dict = Depends(require_login)):
    cliente = obtener_cliente(codigo)
    if not cliente:
        raise HTTPException(404, "Cliente no encontrado")
    atenciones = obtener_atenciones_de_cliente(codigo)
    return JSONResponse({"ok": True, "cliente": cliente, "atenciones": atenciones})


@app.post("/configuracion/importar_json")
async def importar_json_endpoint(
    usuario: dict = Depends(require_admin),
    ruta_json: str = Form(""),
):
    if not ruta_json:
        ruta_json = str(Path(__file__).resolve().parent / "data" / "catalogo_contactos.json")
    try:
        conteo = importar_catalogo_desde_json(ruta_json)
        return JSONResponse({"ok": True, "conteo": conteo})
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@app.get("/configuracion/contactos/exportar")
async def exportar_contactos(usuario: dict = Depends(require_admin)):
    """Descarga un .xlsx con todas las hojas de clientes, atenciones y monedas."""
    contenido = exportar_contactos_xlsx()
    headers = {
        "Content-Disposition": 'attachment; filename="contactos_aroluz.xlsx"',
    }
    return StreamingResponse(
        iter([contenido]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.post("/configuracion/contactos/importar")
async def importar_contactos(
    usuario: dict = Depends(require_admin),
    archivo: UploadFile = File(...),
):
    """Importa/actualiza clientes, atenciones y monedas desde un .xlsx subido."""
    try:
        contenido = await archivo.read()
        conteo = importar_contactos_desde_xlsx(contenido)
        return JSONResponse({
            "ok": True,
            "conteo": conteo,
            "mensaje": (
                f"Importado: {conteo['clientes']} clientes, "
                f"{conteo['atenciones']} atenciones, "
                f"{conteo['monedas']} monedas"
            ),
        })
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@app.get("/configuracion/catalogo/descargar")
async def descargar_catalogo_xlsx(usuario: dict = Depends(require_admin)):
    """Descarga el catálogo de productos como Excel (.xlsx) editable."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    ruta_json = Path(__file__).resolve().parent.parent / "catalogo_productos.json"
    if not ruta_json.exists():
        raise HTTPException(404, "catalogo_productos.json no encontrado")
    with open(ruta_json, encoding="utf-8") as f:
        datos = json.load(f)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo"

    # Encabezados
    headers = ["Categoría", "Subcategoría", "Descripción", "Precio (S/)", "Unidad", "Presentación"]
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Filas de datos
    row = 2
    cat_fill = PatternFill("solid", fgColor="D6E4F0")
    sub_fill = PatternFill("solid", fgColor="EBF5FB")
    for cat in datos.get("categorias", []):
        for sub in cat.get("subcategorias", []):
            for prod in sub.get("productos", []):
                ws.cell(row=row, column=1, value=cat["nombre"]).fill = cat_fill
                ws.cell(row=row, column=2, value=sub["nombre"]).fill = sub_fill
                ws.cell(row=row, column=3, value=prod.get("descripcion", ""))
                ws.cell(row=row, column=4, value=prod.get("precio", 0))
                ws.cell(row=row, column=5, value=prod.get("unidad", "UND"))
                ws.cell(row=row, column=6, value=prod.get("presentacion", ""))
                row += 1

    # Anchos de columna
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="catalogo_productos.xlsx"'},
    )


@app.post("/configuracion/catalogo/subir")
async def subir_catalogo_xlsx(
    usuario: dict = Depends(require_admin),
    archivo: UploadFile = File(...),
):
    """Importa el catálogo desde un .xlsx y reconstruye catalogo_productos.json."""
    import io
    import openpyxl

    try:
        contenido = await archivo.read()
        wb = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        ws = wb.active

        # Leer filas (saltar encabezado en fila 1)
        categorias: dict = {}  # nombre_cat -> dict con subcategorias
        n_prod = 0
        n_invalidos = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            nombre_cat, nombre_sub, desc, precio, unidad, presentacion = (
                (row[i] if i < len(row) else None) for i in range(6)
            )
            if not nombre_cat or not nombre_sub or not desc:
                continue  # fila vacía o incompleta
            nombre_cat = str(nombre_cat).strip()
            nombre_sub = str(nombre_sub).strip()
            desc       = str(desc).strip()
            try:
                precio = float(precio) if precio is not None else 0.0
            except (ValueError, TypeError):
                precio = 0.0
                n_invalidos += 1
            unidad     = str(unidad).strip() if unidad else "UND"
            presentacion = str(presentacion).strip() if presentacion else ""

            if nombre_cat not in categorias:
                categorias[nombre_cat] = {"nombre": nombre_cat, "icono": "📦", "subcategorias": {}}
            cat = categorias[nombre_cat]
            if nombre_sub not in cat["subcategorias"]:
                cat["subcategorias"][nombre_sub] = {"nombre": nombre_sub, "icono": "📋", "productos": []}
            cat["subcategorias"][nombre_sub]["productos"].append({
                "descripcion": desc,
                "unidad": unidad,
                "precio": round(precio, 4),
                "presentacion": presentacion,
            })
            n_prod += 1

        if n_prod == 0:
            return JSONResponse({"ok": False, "error": "No se encontraron productos en el Excel"}, status_code=400)

        # Convertir dicts anidados a listas
        resultado = {
            "version": "1.0",
            "ultima_actualizacion": __import__("datetime").date.today().isoformat(),
            "categorias": [
                {
                    "nombre": cat["nombre"],
                    "icono": cat["icono"],
                    "subcategorias": [
                        sub for sub in cat["subcategorias"].values()
                    ],
                }
                for cat in categorias.values()
            ],
        }

        ruta_json = Path(__file__).resolve().parent.parent / "catalogo_productos.json"
        with open(ruta_json, "w", encoding="utf-8") as f:
            json.dump(resultado, f, ensure_ascii=False, indent=2)

        n_cat = len(resultado["categorias"])
        resp: dict = {
            "ok": True,
            "mensaje": f"Catálogo actualizado: {n_cat} categorías, {n_prod} productos",
        }
        if n_invalidos:
            resp["advertencias"] = f"{n_invalidos} fila(s) con precio inválido (se usó 0.0)"
        return JSONResponse(resp)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=400)


@app.post("/configuracion/usuario/crear")
async def cfg_crear_usuario(
    usuario: dict = Depends(require_admin),
    username: str = Form(...),
    password: str = Form(...),
    nombre: str = Form(""),
    rol: str = Form("USER"),
    ver_asistencias: str = Form(""),
):
    if rol not in ("ADMIN", "USER"):
        rol = "USER"
    ok = crear_usuario(username, password, nombre, rol, ver_asistencias=bool(ver_asistencias))
    return JSONResponse({"ok": ok, "error": "El usuario ya existe" if not ok else None})


@app.post("/configuracion/usuario/cambiar_password")
async def cfg_cambiar_password(
    usuario: dict = Depends(require_admin),
    password_actual: str = Form(...),
    password_nuevo: str = Form(...),
):
    from web.database import verificar_usuario as ver
    if not ver(usuario["u"], password_actual):
        return JSONResponse({"ok": False, "error": "Contraseña actual incorrecta"}, status_code=401)
    ok = cambiar_password(usuario["u"], password_nuevo)
    return JSONResponse({"ok": ok})


@app.put("/configuracion/usuario/{username}")
async def cfg_editar_usuario(
    username: str,
    usuario: dict = Depends(require_admin),
    nombre: str = Form(""),
    rol: str = Form("USER"),
    ver_asistencias: str = Form(""),
):
    if rol not in ("ADMIN", "USER"):
        rol = "USER"
    ok = editar_usuario(username, nombre, rol, ver_asistencias=bool(ver_asistencias))
    return JSONResponse({"ok": ok})


@app.post("/configuracion/usuario/{username}/toggle_activo")
async def cfg_toggle_activo(username: str, usuario: dict = Depends(require_admin)):
    if username == usuario["u"]:
        return JSONResponse({"ok": False, "error": "No puedes desactivar tu propia cuenta"}, status_code=422)
    result = toggle_activo_usuario(username)
    return JSONResponse(result)


@app.delete("/configuracion/usuario/{username}")
async def cfg_eliminar_usuario(username: str, usuario: dict = Depends(require_admin)):
    if username == usuario["u"]:
        return JSONResponse({"ok": False, "error": "No puedes eliminar tu propia cuenta"}, status_code=422)
    ok = eliminar_usuario(username)
    return JSONResponse({"ok": ok})


@app.post("/configuracion/usuario/{username}/reset_password")
async def cfg_reset_password(
    username: str,
    usuario: dict = Depends(require_admin),
    password_nuevo: str = Form(...),
):
    if not password_nuevo or len(password_nuevo) < 4:
        return JSONResponse({"ok": False, "error": "La contraseña debe tener al menos 4 caracteres"}, status_code=422)
    ok = cambiar_password(username, password_nuevo)
    return JSONResponse({"ok": ok})


# ─────────────────────────────────────────────
# Página Mi Configuración (accesible para todos los roles)
# Solo permite editar precios de plancha y cambiar contraseña
# ─────────────────────────────────────────────

@app.get("/mi-config", response_class=HTMLResponse)
async def mi_config_page(request: Request, usuario: dict = Depends(require_login)):
    config = cargar_config()
    return templates.TemplateResponse(
        "mi_config.html",
        ctx(request, usuario, config=config),
    )


@app.post("/mi-config/guardar_planchas")
async def mi_config_guardar_planchas(
    usuario: dict = Depends(require_login),
    go_12: float = Form(...),
    go_15: float = Form(...),
    go_20: float = Form(...),
    gc_12: float = Form(...),
    gc_15: float = Form(...),
    gc_20: float = Form(...),
):
    valores = [go_12, go_15, go_20, gc_12, gc_15, gc_20]
    if any(v <= 0 for v in valores):
        return JSONResponse({"ok": False, "error": "Los precios deben ser mayores a cero"}, status_code=422)
    config = cargar_config()
    config["valores_defecto"]["precios_go"] = {"1.2": go_12, "1.5": go_15, "2.0": go_20}
    config["valores_defecto"]["precios_gc"] = {"1.2": gc_12, "1.5": gc_15, "2.0": gc_20}
    ok = guardar_config(config)
    return JSONResponse({"ok": ok})


@app.post("/mi-config/cambiar_password")
async def mi_config_cambiar_password(
    usuario: dict = Depends(require_login),
    password_actual: str = Form(...),
    password_nuevo: str = Form(...),
    password_confirmar: str = Form(...),
):
    from web.database import verificar_usuario as ver
    if password_nuevo != password_confirmar:
        return JSONResponse({"ok": False, "error": "Las contraseñas nuevas no coinciden"}, status_code=422)
    if len(password_nuevo) < 4:
        return JSONResponse({"ok": False, "error": "La contraseña debe tener al menos 4 caracteres"}, status_code=422)
    if not ver(usuario["u"], password_actual):
        return JSONResponse({"ok": False, "error": "Contraseña actual incorrecta"}, status_code=401)
    ok = cambiar_password(usuario["u"], password_nuevo)
    return JSONResponse({"ok": ok})


# ─────────────────────────────────────────────
# Página Usuarios (solo ADMIN)
# ─────────────────────────────────────────────

@app.get("/usuarios", response_class=HTMLResponse)
async def usuarios_page(request: Request, usuario: dict = Depends(require_admin)):
    usuarios = listar_usuarios()
    return templates.TemplateResponse(
        "usuarios.html",
        ctx(request, usuario, usuarios=usuarios),
    )


# ─────────────────────────────────────────────
# Página Mi Cuenta (accesible para todos los roles)
# ─────────────────────────────────────────────

@app.get("/cuenta", response_class=HTMLResponse)
async def cuenta_page(request: Request, usuario: dict = Depends(require_login)):
    return templates.TemplateResponse(
        "cuenta.html",
        ctx(request, usuario),
    )


@app.post("/cuenta/cambiar_password")
async def cuenta_cambiar_password(
    request: Request,
    usuario: dict = Depends(require_login),
    password_actual: str = Form(...),
    password_nuevo: str = Form(...),
    password_confirmar: str = Form(...),
):
    from web.database import verificar_usuario as ver
    if password_nuevo != password_confirmar:
        return templates.TemplateResponse(
            "cuenta.html",
            ctx(request, usuario, error="Las contraseñas nuevas no coinciden"),
        )
    if not ver(usuario["u"], password_actual):
        return templates.TemplateResponse(
            "cuenta.html",
            ctx(request, usuario, error="Contraseña actual incorrecta"),
        )
    ok = cambiar_password(usuario["u"], password_nuevo)
    if ok:
        return templates.TemplateResponse(
            "cuenta.html",
            ctx(request, usuario, exito="Contraseña actualizada correctamente"),
        )
    return templates.TemplateResponse(
        "cuenta.html",
        ctx(request, usuario, error="Error al cambiar la contraseña"),
    )


# ─────────────────────────────────────────────
# Manejo de errores de autenticación
# ─────────────────────────────────────────────

@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if exc.status_code == 303 and "Location" in exc.headers:
        return RedirectResponse(exc.headers["Location"], status_code=303)
    return JSONResponse({"error": exc.detail}, status_code=exc.status_code)


# ─────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────

if __name__ == "__main__":
    import uvicorn
    import argparse

    parser = argparse.ArgumentParser(description="AROLUZ Cotizador Web")
    parser.add_argument("--host", default="0.0.0.0", help="Host (default: 0.0.0.0)")
    parser.add_argument("--port", type=int, default=8000, help="Puerto (default: 8000)")
    parser.add_argument("--reload", action="store_true", help="Auto-reload en desarrollo")
    args = parser.parse_args()

    print(f"\n{'='*50}")
    print(f"  AROLUZ Cotizador Web v2.0")
    print(f"  URL: http://localhost:{args.port}")
    print(f"  Red local: http://0.0.0.0:{args.port}")
    print(f"{'='*50}\n")

    uvicorn.run(
        "web.main:app",
        host=args.host,
        port=args.port,
        reload=args.reload,
    )
