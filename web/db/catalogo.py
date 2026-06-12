# -*- coding: utf-8 -*-
"""Catálogo: clientes, atenciones, monedas, import/export Excel-JSON — extraído de web/database.py (refactor jun 2026)."""
import hashlib
import json
import os
import re as _re
import secrets
import shutil
import sqlite3
from datetime import datetime as _dt
from pathlib import Path
from typing import Optional, Dict, List, Any

import bcrypt as _bcrypt

from web.db.core import (
    BASE_DIR, DB_PATH, CONFIG_PATH, _CONFIG_RAIZ, CONFIG_DEFECTO,
    _add_column_if_missing, _crear_usuario, _hash_password,
)

# ─────────────────────────────────────────────
# Catálogo (clientes, atenciones, monedas)
# ─────────────────────────────────────────────

def obtener_catalogo() -> Dict[str, List]:
    """Devuelve clientes, atenciones y monedas desde SQLite."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row

    clientes = [dict(r) for r in conn.execute(
        "SELECT codigo, nombre, ruc, ubicacion, abreviacion FROM clientes ORDER BY codigo"
    ).fetchall()]
    atenciones = [dict(r) for r in conn.execute(
        "SELECT nombre, codigo_empresa, email FROM atenciones ORDER BY nombre"
    ).fetchall()]
    monedas = [r["nombre"] for r in conn.execute("SELECT nombre FROM monedas ORDER BY nombre").fetchall()]

    conn.close()
    return {"clientes": clientes, "atenciones": atenciones, "monedas": monedas}


def obtener_cliente(codigo: str) -> Optional[Dict]:
    """Devuelve un cliente por su código, o None si no existe."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT codigo, nombre, ruc, ubicacion, abreviacion FROM clientes WHERE codigo=?", (codigo,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def obtener_atenciones_de_cliente(codigo_cliente: str) -> List[Dict]:
    """Devuelve las atenciones asociadas a un cliente."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT nombre, codigo_empresa, email FROM atenciones WHERE codigo_empresa=? ORDER BY nombre",
        (codigo_cliente,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def importar_catalogo_desde_excel(ruta_excel: str) -> Dict[str, int]:
    """
    Lee clientes, atenciones y monedas desde el Excel (.xlsm) y los guarda en SQLite.

    Estructura de las hojas:
      CLIENTES:  A=CODIGO  B=RAZÓN SOCIAL  C=RUC  D=UBICACIÓN
      ATENCIÓN:  A=CODIGO  B=NOMBRES       C=CORREO  D=CELULAR  E=RAZON SOCIAL (codigo_empresa)
      MONEDA:    A=nombre

    Usa INSERT OR REPLACE para actualizar registros existentes con los datos completos.
    Retorna conteo de registros procesados.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no está instalado")

    if not os.path.exists(ruta_excel):
        raise FileNotFoundError(f"Excel no encontrado: {ruta_excel}")

    # keep_vba=True permite abrir .xlsm sin errores
    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True, keep_vba=True)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conteo = {"clientes": 0, "atenciones": 0, "monedas": 0}

    def _s(val) -> str:
        return str(val).strip() if val is not None else ""

    # Clientes: A=codigo, B=razón social, C=RUC, D=ubicación
    if "CLIENTES" in wb.sheetnames:
        ws = wb["CLIENTES"]
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            codigo = _s(row[0])
            if not codigo:
                continue
            nombre = _s(row[1]) or codigo
            ruc = _s(row[2])
            ubicacion = _s(row[3])
            conn.execute(
                """INSERT INTO clientes (codigo, nombre, ruc, ubicacion)
                   VALUES (?, ?, ?, ?)
                   ON CONFLICT(codigo) DO UPDATE SET
                     nombre=excluded.nombre,
                     ruc=excluded.ruc,
                     ubicacion=excluded.ubicacion""",
                (codigo, nombre, ruc, ubicacion),
            )
            conteo["clientes"] += 1

    # Atenciones: A=codigo_interno, B=nombres, C=correo, D=celular, E=razon_social (codigo_empresa)
    if "ATENCIÓN" in wb.sheetnames:
        ws = wb["ATENCIÓN"]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            nombre = _s(row[1])        # col B — nombre completo de la persona
            email = _s(row[2])         # col C — correo
            codigo_emp = _s(row[4])    # col E — código empresa (vincula con CLIENTES)
            if not nombre or not codigo_emp:
                continue
            conn.execute(
                """INSERT INTO atenciones (nombre, codigo_empresa, email)
                   VALUES (?, ?, ?)
                   ON CONFLICT(nombre) DO UPDATE SET
                     codigo_empresa=excluded.codigo_empresa,
                     email=excluded.email""",
                (nombre, codigo_emp, email),
            )
            conteo["atenciones"] += 1

    # Monedas: A=nombre
    if "MONEDA" in wb.sheetnames:
        ws = wb["MONEDA"]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = _s(row[0])
            if val:
                conn.execute(
                    "INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (val,)
                )
                conteo["monedas"] += 1

    conn.commit()
    conn.close()
    wb.close()
    return conteo


def agregar_cliente(codigo: str, nombre: str = "", ruc: str = "", ubicacion: str = "", abreviacion: str = "") -> bool:
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "INSERT OR IGNORE INTO clientes (codigo, nombre, ruc, ubicacion, abreviacion) VALUES (?, ?, ?, ?, ?)",
            (codigo.strip(), nombre.strip() or codigo.strip(), ruc.strip(), ubicacion.strip(), abreviacion.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def agregar_atencion(nombre: str, codigo_empresa: str, email: str = "") -> bool:
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "INSERT OR IGNORE INTO atenciones (nombre, codigo_empresa, email) VALUES (?, ?, ?)",
            (nombre.strip(), codigo_empresa.strip(), email.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def eliminar_cliente(codigo: str) -> bool:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("DELETE FROM clientes WHERE codigo = ?", (codigo,))
    conn.commit()
    conn.close()
    return True


def eliminar_atencion(nombre: str) -> bool:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("DELETE FROM atenciones WHERE nombre = ?", (nombre,))
    conn.commit()
    conn.close()
    return True


def editar_cliente(codigo: str, nuevo_nombre: str, ruc: str = "", ubicacion: str = "", abreviacion: str = "") -> bool:
    """Actualiza nombre, RUC, dirección y abreviación de un cliente existente."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "UPDATE clientes SET nombre=?, ruc=?, ubicacion=?, abreviacion=? WHERE codigo=?",
            (nuevo_nombre.strip(), ruc.strip(), ubicacion.strip(), abreviacion.strip(), codigo.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def editar_atencion(nombre_actual: str, nuevo_nombre: str, nuevo_codigo_empresa: str, email: str = "") -> bool:
    """Actualiza nombre, código de empresa y email de una atención."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "UPDATE atenciones SET nombre=?, codigo_empresa=?, email=? WHERE nombre=?",
            (nuevo_nombre.strip(), nuevo_codigo_empresa.strip(), email.strip(), nombre_actual.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def exportar_contactos_xlsx() -> bytes:
    """
    Exporta clientes, atenciones y monedas a un .xlsx con 3 hojas.
    Retorna los bytes del archivo para streaming.

    Hojas:
      CLIENTES:   codigo · nombre · ruc · ubicacion
      ATENCIONES: nombre · codigo_empresa · email
      MONEDAS:    nombre
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("openpyxl no está instalado")

    from io import BytesIO

    catalogo = obtener_catalogo()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    bold = Font(bold=True)

    # Hoja CLIENTES
    ws_cli = wb.create_sheet("CLIENTES")
    headers_cli = ["codigo", "nombre", "ruc", "ubicacion", "abreviacion"]
    for col, h in enumerate(headers_cli, 1):
        cell = ws_cli.cell(row=1, column=col, value=h)
        cell.font = bold
    for row_idx, c in enumerate(catalogo["clientes"], 2):
        ws_cli.cell(row=row_idx, column=1, value=c.get("codigo", ""))
        ws_cli.cell(row=row_idx, column=2, value=c.get("nombre", ""))
        ws_cli.cell(row=row_idx, column=3, value=c.get("ruc", ""))
        ws_cli.cell(row=row_idx, column=4, value=c.get("ubicacion", ""))
        ws_cli.cell(row=row_idx, column=5, value=c.get("abreviacion", ""))

    # Hoja ATENCIONES
    ws_ate = wb.create_sheet("ATENCIONES")
    headers_ate = ["nombre", "codigo_empresa", "email"]
    for col, h in enumerate(headers_ate, 1):
        cell = ws_ate.cell(row=1, column=col, value=h)
        cell.font = bold
    for row_idx, a in enumerate(catalogo["atenciones"], 2):
        ws_ate.cell(row=row_idx, column=1, value=a.get("nombre", ""))
        ws_ate.cell(row=row_idx, column=2, value=a.get("codigo_empresa", ""))
        ws_ate.cell(row=row_idx, column=3, value=a.get("email", ""))

    # Hoja MONEDAS
    ws_mon = wb.create_sheet("MONEDAS")
    cell = ws_mon.cell(row=1, column=1, value="nombre")
    cell.font = bold
    for row_idx, m in enumerate(catalogo["monedas"], 2):
        ws_mon.cell(row=row_idx, column=1, value=m)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_contactos_desde_xlsx(contenido_bytes: bytes) -> Dict[str, int]:
    """
    Lee un .xlsx exportado por exportar_contactos_xlsx() e importa/actualiza
    clientes, atenciones y monedas en SQLite.

    Estructura esperada:
      Hoja CLIENTES:   col A=codigo, B=nombre, C=ruc, D=ubicacion
      Hoja ATENCIONES: col A=nombre, B=codigo_empresa, C=email
      Hoja MONEDAS:    col A=nombre

    Ignora filas donde la columna A esté vacía (encabezados u hojas vacías).
    Retorna conteo de filas procesadas por hoja.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no está instalado")

    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(contenido_bytes), read_only=True, data_only=True)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conteo = {"clientes": 0, "atenciones": 0, "monedas": 0}

    def _s(val) -> str:
        return str(val).strip() if val is not None else ""

    # Hoja CLIENTES
    if "CLIENTES" in wb.sheetnames:
        ws = wb["CLIENTES"]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            codigo = _s(row[0]) if len(row) > 0 else ""
            if not codigo:
                continue
            nombre = _s(row[1]) if len(row) > 1 else ""
            ruc = _s(row[2]) if len(row) > 2 else ""
            ubicacion = _s(row[3]) if len(row) > 3 else ""
            abreviacion = _s(row[4]) if len(row) > 4 else ""
            conn.execute(
                """INSERT INTO clientes (codigo, nombre, ruc, ubicacion, abreviacion)
                   VALUES (?, ?, ?, ?, ?)
                   ON CONFLICT(codigo) DO UPDATE SET
                     nombre=excluded.nombre,
                     ruc=excluded.ruc,
                     ubicacion=excluded.ubicacion,
                     abreviacion=excluded.abreviacion""",
                (codigo, nombre or codigo, ruc, ubicacion, abreviacion),
            )
            conteo["clientes"] += 1

    # Hoja ATENCIONES
    if "ATENCIONES" in wb.sheetnames:
        ws = wb["ATENCIONES"]
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            nombre = _s(row[0]) if len(row) > 0 else ""
            if not nombre:
                continue
            codigo_empresa = _s(row[1]) if len(row) > 1 else ""
            email = _s(row[2]) if len(row) > 2 else ""
            conn.execute(
                """INSERT INTO atenciones (nombre, codigo_empresa, email)
                   VALUES (?, ?, ?)
                   ON CONFLICT(nombre) DO UPDATE SET
                     codigo_empresa=excluded.codigo_empresa,
                     email=excluded.email""",
                (nombre, codigo_empresa, email),
            )
            conteo["atenciones"] += 1

    # Hoja MONEDAS
    if "MONEDAS" in wb.sheetnames:
        ws = wb["MONEDAS"]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = _s(row[0]) if len(row) > 0 else ""
            if val:
                conn.execute("INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (val,))
                conteo["monedas"] += 1

    conn.commit()
    conn.close()
    wb.close()
    return conteo


def importar_catalogo_desde_json(ruta: str) -> Dict[str, int]:
    """
    Lee clientes, atenciones y monedas desde un JSON y los inserta en SQLite (INSERT OR IGNORE).
    Retorna conteo de registros procesados.
    """
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta}")
    with open(ruta, encoding="utf-8") as f:
        data = json.load(f)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conteo = {"clientes": 0, "atenciones": 0, "monedas": 0}
    for c in data.get("clientes", []):
        conn.execute(
            "INSERT OR IGNORE INTO clientes (codigo, nombre) VALUES (?, ?)",
            (c["codigo"], c.get("nombre", c["codigo"])),
        )
        conteo["clientes"] += 1
    for a in data.get("atenciones", []):
        conn.execute(
            "INSERT OR IGNORE INTO atenciones (nombre, codigo_empresa) VALUES (?, ?)",
            (a["nombre"], a["codigo_empresa"]),
        )
        conteo["atenciones"] += 1
    for m in data.get("monedas", []):
        conn.execute(
            "INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (m,)
        )
        conteo["monedas"] += 1
    conn.commit()
    conn.close()
    return conteo

