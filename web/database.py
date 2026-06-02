"""
database.py — Base de datos SQLite para usuarios y catálogo

Gestiona:
- Tabla de usuarios (autenticación)
- Configuración del sistema (leída desde cotizador_config.json)
- Catálogo de clientes/atenciones (desde Excel o tabla SQLite)
"""
import hashlib
import json
import os
import re as _re
import secrets
import shutil
import sqlite3
from datetime import datetime as _dt
from pathlib import Path
from typing import Optional, Dict, List, Any

import bcrypt as _bcrypt

# Ruta de la base de datos
BASE_DIR = Path(__file__).parent
DB_PATH = BASE_DIR / "data" / "aroluz.db"
CONFIG_PATH = BASE_DIR / "data" / "cotizador_config.json"
_CONFIG_RAIZ = BASE_DIR.parent / "cotizador_config.json"

CONFIG_DEFECTO = {
    "rutas": {
        "carpeta_excel": "",
        "carpeta_pdfs": "",
    },
    "valores_defecto": {
        "ganancia": "30",
        "galvanizado": "GO",
        "espesor_producto": "1.5",
        "espesor_tapa": "1.5",
        "precios_go": {"1.2": 150.0, "1.5": 180.0, "2.0": 220.0},
        "precios_gc": {"1.2": 140.0, "1.5": 170.0, "2.0": 210.0},
        "dolar": 3.8,
        "usd_kg_productos": 1.0,
        "usd_kg_cajas": 3.0,
    },
    "interfaz": {
        "recordar_config": True,
        "recordar_medidas": True,
        "mostrar_validaciones": True,
    },
    "permisos_usuario": {
        "ver_historial": True,
        "ver_catalogo": True,
    },
    "factores_ganancia": {
        "30": {"B": 0.70, "CH": 0.50, "CVE": 0.50, "CVI": 0.50, "T": 0.60, "C": 0.70, "R": 0.20, "CP": 0.50},
        "35": {"B": 0.65, "CH": 0.45, "CVE": 0.45, "CVI": 0.45, "T": 0.55, "C": 0.65, "R": 0.15, "CP": 0.475},
    },
}


# ─────────────────────────────────────────────
# Inicialización
# ─────────────────────────────────────────────

def _backup_db():
    """Copia la DB a web/data/backups/ al iniciar. Conserva solo los últimos 7 backups."""
    if not DB_PATH.exists():
        return
    backup_dir = DB_PATH.parent / "backups"
    backup_dir.mkdir(exist_ok=True)
    ts = _dt.now().strftime("%Y%m%d_%H%M%S")
    dest = backup_dir / f"aroluz_{ts}.db"
    shutil.copy2(DB_PATH, dest)
    # Conservar solo los últimos 7 backups
    backups = sorted(backup_dir.glob("aroluz_*.db"))
    for old in backups[:-7]:
        old.unlink()


def _add_column_if_missing(conn, table: str, column: str, col_def: str):
    """Agrega una columna a una tabla si no existe (migración segura)."""
    try:
        conn.execute(f"ALTER TABLE {table} ADD COLUMN {column} {col_def}")
    except sqlite3.OperationalError:
        pass  # columna ya existe


def _migrate_cantidad_to_real(conn):
    """Migra carrito_items.cantidad de INTEGER a REAL si todavía es INTEGER."""
    rows = conn.execute("PRAGMA table_info(carrito_items)").fetchall()
    col = next((r for r in rows if r[1] == "cantidad"), None)
    if col is None or col[2].upper() == "REAL":
        return  # ya está bien
    # Recrear tabla con REAL
    conn.execute("ALTER TABLE carrito_items RENAME TO carrito_items_bak")
    conn.execute("""
        CREATE TABLE carrito_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            tipo TEXT NOT NULL,
            descripcion TEXT NOT NULL,
            precio_unitario REAL NOT NULL,
            peso_unitario REAL NOT NULL,
            cantidad REAL NOT NULL DEFAULT 1,
            unidad TEXT NOT NULL DEFAULT 'UND',
            tipo_galvanizado TEXT NOT NULL DEFAULT 'GO',
            porcentaje_ganancia TEXT NOT NULL DEFAULT '30',
            descripcion_calculada TEXT DEFAULT NULL
        )
    """)
    conn.execute("""
        INSERT INTO carrito_items
            (id, username, tipo, descripcion, precio_unitario, peso_unitario,
             cantidad, unidad, tipo_galvanizado, porcentaje_ganancia, descripcion_calculada)
        SELECT id, username, tipo, descripcion, precio_unitario, peso_unitario,
               CAST(cantidad AS REAL), unidad, tipo_galvanizado, porcentaje_ganancia,
               descripcion_calculada
        FROM carrito_items_bak
    """)
    conn.execute("DROP TABLE carrito_items_bak")


def init_db():
    """Crea las tablas necesarias si no existen."""
    _backup_db()
    DB_PATH.parent.mkdir(parents=True, exist_ok=True)
    # Migrar config de la raíz a web/data/ si aún no existe allí
    if not CONFIG_PATH.exists() and _CONFIG_RAIZ.exists():
        shutil.copy2(_CONFIG_RAIZ, CONFIG_PATH)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            nombre TEXT DEFAULT '',
            activo INTEGER DEFAULT 1
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS clientes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS atenciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL,
            codigo_empresa TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS monedas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT UNIQUE NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS carrito_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            tipo TEXT NOT NULL,
            descripcion TEXT NOT NULL,
            precio_unitario REAL NOT NULL,
            peso_unitario REAL NOT NULL,
            cantidad REAL NOT NULL DEFAULT 1,
            unidad TEXT NOT NULL DEFAULT 'UND',
            tipo_galvanizado TEXT NOT NULL DEFAULT 'GO',
            porcentaje_ganancia TEXT NOT NULL DEFAULT '30'
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS cotizaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            fecha TEXT NOT NULL,
            cliente TEXT NOT NULL DEFAULT '',
            atencion TEXT NOT NULL DEFAULT '',
            proyecto TEXT NOT NULL DEFAULT '',
            moneda TEXT NOT NULL DEFAULT 'SOLES',
            total_precio REAL NOT NULL DEFAULT 0,
            total_peso REAL NOT NULL DEFAULT 0
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS cotizacion_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cotizacion_id INTEGER NOT NULL REFERENCES cotizaciones(id) ON DELETE CASCADE,
            tipo TEXT NOT NULL,
            descripcion TEXT NOT NULL,
            precio_unitario REAL NOT NULL,
            peso_unitario REAL NOT NULL,
            cantidad INTEGER NOT NULL,
            unidad TEXT NOT NULL,
            tipo_galvanizado TEXT NOT NULL,
            porcentaje_ganancia TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS reportes_asistencia (
            id             INTEGER PRIMARY KEY AUTOINCREMENT,
            periodo        TEXT NOT NULL UNIQUE,
            periodo_label  TEXT NOT NULL,
            fecha_subida   TEXT NOT NULL,
            subido_por     TEXT NOT NULL,
            nombre_archivo TEXT NOT NULL,
            num_empleados  INTEGER NOT NULL,
            datos_json     TEXT NOT NULL
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS email_imap_config (
            id       INTEGER PRIMARY KEY DEFAULT 1,
            host     TEXT NOT NULL DEFAULT '',
            port     INTEGER NOT NULL DEFAULT 993,
            username TEXT NOT NULL DEFAULT '',
            password TEXT NOT NULL DEFAULT '',
            folder   TEXT NOT NULL DEFAULT 'INBOX',
            days_back INTEGER NOT NULL DEFAULT 30
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS email_importados (
            message_id    TEXT PRIMARY KEY,
            proyecto      TEXT NOT NULL DEFAULT '',
            importado_at  TEXT NOT NULL
        )
    """)

    conn.commit()

    # Migraciones: agregar columnas nuevas si no existen (bases de datos existentes)
    _add_column_if_missing(conn, "clientes", "ruc", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "clientes", "ubicacion", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "clientes", "abreviacion", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "atenciones", "email", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "cotizaciones", "cliente_nombre", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "cotizaciones", "cliente_ruc", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "cotizaciones", "cliente_ubicacion", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "cotizaciones", "atencion_email", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "cotizaciones", "dolar_rate", "REAL NOT NULL DEFAULT 3.8")
    _add_column_if_missing(conn, "cotizaciones", "validez", "TEXT NOT NULL DEFAULT '30 días'")
    _add_column_if_missing(conn, "cotizaciones", "encabezado_tabla", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "usuarios", "rol", "TEXT NOT NULL DEFAULT 'USER'")
    _add_column_if_missing(conn, "usuarios", "ver_asistencias", "INTEGER NOT NULL DEFAULT 0")
    _add_column_if_missing(conn, "carrito_items", "descripcion_calculada", "TEXT DEFAULT NULL")
    _add_column_if_missing(conn, "carrito_items", "posicion", "REAL DEFAULT NULL")
    _add_column_if_missing(conn, "carrito_items", "tapa_para_id", "INTEGER DEFAULT NULL")
    _add_column_if_missing(conn, "carrito_items", "precio_manual", "INTEGER NOT NULL DEFAULT 0")
    _add_column_if_missing(conn, "cotizacion_items", "precio_manual", "INTEGER NOT NULL DEFAULT 0")
    _add_column_if_missing(conn, "email_importados", "pdf_hash", "TEXT NOT NULL DEFAULT ''")
    _migrate_cantidad_to_real(conn)
    _add_column_if_missing(conn, "cotizaciones", "origen", "TEXT NOT NULL DEFAULT 'web'")
    # Promover al usuario admin a ADMIN si aún no lo es
    conn.execute("UPDATE usuarios SET rol='ADMIN' WHERE username='admin' AND rol='USER'")
    # ADMIN siempre tiene acceso a asistencias
    conn.execute("UPDATE usuarios SET ver_asistencias=1 WHERE rol='ADMIN'")
    conn.commit()

    # Crear usuario admin por defecto si no hay usuarios
    c.execute("SELECT COUNT(*) FROM usuarios")
    if c.fetchone()[0] == 0:
        _crear_usuario(conn, "admin", "aroluz2024", "Administrador", "ADMIN")
        print("Usuario admin creado. Cambiá la contraseña en /cuenta")

    # Seed inicial desde JSON si las tablas de catálogo están vacías
    c.execute("SELECT COUNT(*) FROM clientes")
    if c.fetchone()[0] == 0:
        _seed_desde_json(conn)
        conn.commit()

    # Poblar monedas por defecto si están vacías (fallback si JSON no tiene monedas)
    c.execute("SELECT COUNT(*) FROM monedas")
    if c.fetchone()[0] == 0:
        for m in ["SOLES", "DÓLARES"]:
            c.execute("INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (m,))
        conn.commit()

    # Índices para consultas frecuentes
    c.execute("CREATE INDEX IF NOT EXISTS idx_cotizaciones_username ON cotizaciones(username)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_cotizaciones_fecha ON cotizaciones(fecha DESC)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_cotizaciones_proyecto ON cotizaciones(proyecto)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_cotizacion_items_cotizacion_id ON cotizacion_items(cotizacion_id)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_carrito_items_username ON carrito_items(username)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_reportes_asistencia_periodo ON reportes_asistencia(periodo DESC)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_atenciones_codigo_empresa ON atenciones(codigo_empresa)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_email_importados_pdf_hash ON email_importados(pdf_hash)")
    c.execute("CREATE INDEX IF NOT EXISTS idx_email_importados_proyecto ON email_importados(proyecto)")
    conn.commit()

    conn.close()


def _seed_desde_json(conn):
    """Puebla clientes, atenciones y monedas desde catalogo_contactos.json (solo si están vacíos)."""
    json_path = os.path.join(os.path.dirname(__file__), "data", "catalogo_contactos.json")
    if not os.path.exists(json_path):
        return
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    for c in data.get("clientes", []):
        conn.execute(
            "INSERT OR IGNORE INTO clientes (codigo, nombre, ruc, ubicacion) VALUES (?, ?, ?, ?)",
            (c["codigo"], c.get("nombre", c["codigo"]), c.get("ruc", ""), c.get("ubicacion", "")),
        )
    for a in data.get("atenciones", []):
        conn.execute(
            "INSERT OR IGNORE INTO atenciones (nombre, codigo_empresa, email) VALUES (?, ?, ?)",
            (a["nombre"], a["codigo_empresa"], a.get("email", "")),
        )
    for m in data.get("monedas", []):
        conn.execute(
            "INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (m,)
        )


def _crear_usuario(conn, username: str, password: str, nombre: str = "", rol: str = "USER", ver_asistencias: bool = False):
    ph = _hash_password(password)
    va = 1 if (rol == "ADMIN" or ver_asistencias) else 0
    conn.execute(
        "INSERT OR IGNORE INTO usuarios (username, password_hash, nombre, rol, ver_asistencias) VALUES (?, ?, ?, ?, ?)",
        (username, ph, nombre, rol, va),
    )
    conn.commit()


def _hash_password(password: str) -> str:
    return _bcrypt.hashpw(password.encode(), _bcrypt.gensalt(rounds=12)).decode()


# ─────────────────────────────────────────────
# Usuarios
# ─────────────────────────────────────────────

def verificar_usuario(username: str, password: str) -> Optional[Dict]:
    """Devuelve el usuario si las credenciales son válidas, None si no.

    Soporta migración transparente SHA256 → bcrypt: en el primer login con hash
    SHA256 legado, re-hashea con bcrypt automáticamente.
    """
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        "SELECT id, username, password_hash, nombre, rol, ver_asistencias FROM usuarios WHERE username=? AND activo=1",
        (username,),
    )
    row = c.fetchone()
    if not row:
        conn.close()
        return None
    ph = row["password_hash"]
    if ph.startswith("$2"):
        # Hash bcrypt moderno
        ok = _bcrypt.checkpw(password.encode(), ph.encode())
    else:
        # Hash SHA256 legado (64 chars hex) — verificar y migrar a bcrypt
        ok = (hashlib.sha256(password.encode()).hexdigest() == ph)
        if ok:
            new_ph = _hash_password(password)
            conn.execute(
                "UPDATE usuarios SET password_hash=? WHERE username=?",
                (new_ph, username),
            )
            conn.commit()
    conn.close()
    if not ok:
        return None
    return {"id": row["id"], "username": row["username"], "nombre": row["nombre"], "rol": row["rol"], "ver_asistencias": bool(row["ver_asistencias"])}


def crear_usuario(username: str, password: str, nombre: str = "", rol: str = "USER", ver_asistencias: bool = False) -> bool:
    """Crea un nuevo usuario. Retorna True si fue exitoso."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        _crear_usuario(conn, username, password, nombre, rol, ver_asistencias)
        conn.close()
        return True
    except sqlite3.IntegrityError:
        return False


def cambiar_password(username: str, nueva_password: str) -> bool:
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        ph = _hash_password(nueva_password)
        conn.execute(
            "UPDATE usuarios SET password_hash=? WHERE username=?",
            (ph, username),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def listar_usuarios() -> List[Dict]:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("SELECT id, username, nombre, activo, rol, ver_asistencias FROM usuarios").fetchall()
    conn.close()
    return [{**dict(r), "ver_asistencias": bool(r["ver_asistencias"])} for r in rows]


def editar_usuario(username: str, nombre: str, rol: str, ver_asistencias: bool = False) -> bool:
    """Edita nombre, rol y permisos de un usuario existente."""
    if rol not in ("ADMIN", "USER"):
        rol = "USER"
    va = 1 if (rol == "ADMIN" or ver_asistencias) else 0
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "UPDATE usuarios SET nombre=?, rol=?, ver_asistencias=? WHERE username=?",
            (nombre, rol, va, username),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def toggle_activo_usuario(username: str) -> dict:
    """Alterna el estado activo/inactivo de un usuario."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        row = conn.execute("SELECT activo FROM usuarios WHERE username=?", (username,)).fetchone()
        if not row:
            conn.close()
            return {"ok": False, "error": "Usuario no encontrado"}
        nuevo = 0 if row[0] else 1
        conn.execute("UPDATE usuarios SET activo=? WHERE username=?", (nuevo, username))
        conn.commit()
        conn.close()
        return {"ok": True, "activo": bool(nuevo)}
    except Exception as e:
        return {"ok": False, "error": str(e)}


def eliminar_usuario(username: str) -> bool:
    """Elimina un usuario permanentemente."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute("DELETE FROM usuarios WHERE username=?", (username,))
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


# ─────────────────────────────────────────────
# Configuración
# ─────────────────────────────────────────────

def cargar_config() -> Dict:
    """Carga configuración desde cotizador_config.json, mergeando con defaults."""
    if CONFIG_PATH.exists():
        try:
            with open(CONFIG_PATH, encoding="utf-8") as f:
                cfg = json.load(f)
            resultado = _fusionar(CONFIG_DEFECTO, cfg)
            # Migración: corregir espesor_tapa si quedó en el valor antiguo por defecto
            vd = resultado.get("valores_defecto", {})
            if vd.get("espesor_tapa") == "1.2" and cfg.get("valores_defecto", {}).get("espesor_tapa") == "1.2":
                vd["espesor_tapa"] = "1.5"
                guardar_config(resultado)
            return resultado
        except Exception:
            pass
    return dict(CONFIG_DEFECTO)


def guardar_config(config: Dict) -> bool:
    """Guarda configuración en cotizador_config.json."""
    try:
        backup = CONFIG_PATH.parent / "cotizador_config_backup.json"
        if CONFIG_PATH.exists():
            import shutil
            shutil.copy2(CONFIG_PATH, backup)
        with open(CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
        return True
    except Exception:
        return False


def _fusionar(defecto: Dict, cargada: Dict) -> Dict:
    resultado = {**defecto}
    for k, v in cargada.items():
        if k in resultado and isinstance(v, dict):
            resultado[k] = {**resultado[k], **v}
        else:
            resultado[k] = v
    return resultado


# ─────────────────────────────────────────────
# Catálogo (clientes, atenciones, monedas)
# ─────────────────────────────────────────────

def obtener_catalogo() -> Dict[str, List]:
    """Devuelve clientes, atenciones y monedas desde SQLite."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row

    clientes = [dict(r) for r in conn.execute(
        "SELECT codigo, nombre, ruc, ubicacion, abreviacion FROM clientes ORDER BY codigo"
    ).fetchall()]
    atenciones = [dict(r) for r in conn.execute(
        "SELECT nombre, codigo_empresa, email FROM atenciones ORDER BY nombre"
    ).fetchall()]
    monedas = [r["nombre"] for r in conn.execute("SELECT nombre FROM monedas ORDER BY nombre").fetchall()]

    conn.close()
    return {"clientes": clientes, "atenciones": atenciones, "monedas": monedas}


def obtener_cliente(codigo: str) -> Optional[Dict]:
    """Devuelve un cliente por su código, o None si no existe."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT codigo, nombre, ruc, ubicacion, abreviacion FROM clientes WHERE codigo=?", (codigo,)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def obtener_atenciones_de_cliente(codigo_cliente: str) -> List[Dict]:
    """Devuelve las atenciones asociadas a un cliente."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT nombre, codigo_empresa, email FROM atenciones WHERE codigo_empresa=? ORDER BY nombre",
        (codigo_cliente,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def importar_catalogo_desde_excel(ruta_excel: str) -> Dict[str, int]:
    """
    Lee clientes, atenciones y monedas desde el Excel (.xlsm) y los guarda en SQLite.

    Estructura de las hojas:
      CLIENTES:  A=CODIGO  B=RAZÓN SOCIAL  C=RUC  D=UBICACIÓN
      ATENCIÓN:  A=CODIGO  B=NOMBRES       C=CORREO  D=CELULAR  E=RAZON SOCIAL (codigo_empresa)
      MONEDA:    A=nombre

    Usa INSERT OR REPLACE para actualizar registros existentes con los datos completos.
    Retorna conteo de registros procesados.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no está instalado")

    if not os.path.exists(ruta_excel):
        raise FileNotFoundError(f"Excel no encontrado: {ruta_excel}")

    # keep_vba=True permite abrir .xlsm sin errores
    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True, keep_vba=True)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conteo = {"clientes": 0, "atenciones": 0, "monedas": 0}

    def _s(val) -> str:
        return str(val).strip() if val is not None else ""

    # Clientes: A=codigo, B=razón social, C=RUC, D=ubicación
    if "CLIENTES" in wb.sheetnames:
        ws = wb["CLIENTES"]
        for row in ws.iter_rows(min_row=2, max_col=4, values_only=True):
            codigo = _s(row[0])
            if not codigo:
                continue
            nombre = _s(row[1]) or codigo
            ruc = _s(row[2])
            ubicacion = _s(row[3])
            conn.execute(
                """INSERT INTO clientes (codigo, nombre, ruc, ubicacion)
                   VALUES (?, ?, ?, ?)
                   ON CONFLICT(codigo) DO UPDATE SET
                     nombre=excluded.nombre,
                     ruc=excluded.ruc,
                     ubicacion=excluded.ubicacion""",
                (codigo, nombre, ruc, ubicacion),
            )
            conteo["clientes"] += 1

    # Atenciones: A=codigo_interno, B=nombres, C=correo, D=celular, E=razon_social (codigo_empresa)
    if "ATENCIÓN" in wb.sheetnames:
        ws = wb["ATENCIÓN"]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            nombre = _s(row[1])        # col B — nombre completo de la persona
            email = _s(row[2])         # col C — correo
            codigo_emp = _s(row[4])    # col E — código empresa (vincula con CLIENTES)
            if not nombre or not codigo_emp:
                continue
            conn.execute(
                """INSERT INTO atenciones (nombre, codigo_empresa, email)
                   VALUES (?, ?, ?)
                   ON CONFLICT(nombre) DO UPDATE SET
                     codigo_empresa=excluded.codigo_empresa,
                     email=excluded.email""",
                (nombre, codigo_emp, email),
            )
            conteo["atenciones"] += 1

    # Monedas: A=nombre
    if "MONEDA" in wb.sheetnames:
        ws = wb["MONEDA"]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = _s(row[0])
            if val:
                conn.execute(
                    "INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (val,)
                )
                conteo["monedas"] += 1

    conn.commit()
    conn.close()
    wb.close()
    return conteo


def agregar_cliente(codigo: str, nombre: str = "", ruc: str = "", ubicacion: str = "", abreviacion: str = "") -> bool:
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "INSERT OR IGNORE INTO clientes (codigo, nombre, ruc, ubicacion, abreviacion) VALUES (?, ?, ?, ?, ?)",
            (codigo.strip(), nombre.strip() or codigo.strip(), ruc.strip(), ubicacion.strip(), abreviacion.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def agregar_atencion(nombre: str, codigo_empresa: str, email: str = "") -> bool:
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "INSERT OR IGNORE INTO atenciones (nombre, codigo_empresa, email) VALUES (?, ?, ?)",
            (nombre.strip(), codigo_empresa.strip(), email.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def eliminar_cliente(codigo: str) -> bool:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("DELETE FROM clientes WHERE codigo = ?", (codigo,))
    conn.commit()
    conn.close()
    return True


def eliminar_atencion(nombre: str) -> bool:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("DELETE FROM atenciones WHERE nombre = ?", (nombre,))
    conn.commit()
    conn.close()
    return True


def editar_cliente(codigo: str, nuevo_nombre: str, ruc: str = "", ubicacion: str = "", abreviacion: str = "") -> bool:
    """Actualiza nombre, RUC, dirección y abreviación de un cliente existente."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "UPDATE clientes SET nombre=?, ruc=?, ubicacion=?, abreviacion=? WHERE codigo=?",
            (nuevo_nombre.strip(), ruc.strip(), ubicacion.strip(), abreviacion.strip(), codigo.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def editar_atencion(nombre_actual: str, nuevo_nombre: str, nuevo_codigo_empresa: str, email: str = "") -> bool:
    """Actualiza nombre, código de empresa y email de una atención."""
    try:
        conn = sqlite3.connect(DB_PATH, timeout=10)
        conn.execute(
            "UPDATE atenciones SET nombre=?, codigo_empresa=?, email=? WHERE nombre=?",
            (nuevo_nombre.strip(), nuevo_codigo_empresa.strip(), email.strip(), nombre_actual.strip()),
        )
        conn.commit()
        conn.close()
        return True
    except Exception:
        return False


def exportar_contactos_xlsx() -> bytes:
    """
    Exporta clientes, atenciones y monedas a un .xlsx con 3 hojas.
    Retorna los bytes del archivo para streaming.

    Hojas:
      CLIENTES:   codigo · nombre · ruc · ubicacion
      ATENCIONES: nombre · codigo_empresa · email
      MONEDAS:    nombre
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        raise ImportError("openpyxl no está instalado")

    from io import BytesIO

    catalogo = obtener_catalogo()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # quitar hoja vacía por defecto

    bold = Font(bold=True)

    # Hoja CLIENTES
    ws_cli = wb.create_sheet("CLIENTES")
    headers_cli = ["codigo", "nombre", "ruc", "ubicacion", "abreviacion"]
    for col, h in enumerate(headers_cli, 1):
        cell = ws_cli.cell(row=1, column=col, value=h)
        cell.font = bold
    for row_idx, c in enumerate(catalogo["clientes"], 2):
        ws_cli.cell(row=row_idx, column=1, value=c.get("codigo", ""))
        ws_cli.cell(row=row_idx, column=2, value=c.get("nombre", ""))
        ws_cli.cell(row=row_idx, column=3, value=c.get("ruc", ""))
        ws_cli.cell(row=row_idx, column=4, value=c.get("ubicacion", ""))
        ws_cli.cell(row=row_idx, column=5, value=c.get("abreviacion", ""))

    # Hoja ATENCIONES
    ws_ate = wb.create_sheet("ATENCIONES")
    headers_ate = ["nombre", "codigo_empresa", "email"]
    for col, h in enumerate(headers_ate, 1):
        cell = ws_ate.cell(row=1, column=col, value=h)
        cell.font = bold
    for row_idx, a in enumerate(catalogo["atenciones"], 2):
        ws_ate.cell(row=row_idx, column=1, value=a.get("nombre", ""))
        ws_ate.cell(row=row_idx, column=2, value=a.get("codigo_empresa", ""))
        ws_ate.cell(row=row_idx, column=3, value=a.get("email", ""))

    # Hoja MONEDAS
    ws_mon = wb.create_sheet("MONEDAS")
    cell = ws_mon.cell(row=1, column=1, value="nombre")
    cell.font = bold
    for row_idx, m in enumerate(catalogo["monedas"], 2):
        ws_mon.cell(row=row_idx, column=1, value=m)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def importar_contactos_desde_xlsx(contenido_bytes: bytes) -> Dict[str, int]:
    """
    Lee un .xlsx exportado por exportar_contactos_xlsx() e importa/actualiza
    clientes, atenciones y monedas en SQLite.

    Estructura esperada:
      Hoja CLIENTES:   col A=codigo, B=nombre, C=ruc, D=ubicacion
      Hoja ATENCIONES: col A=nombre, B=codigo_empresa, C=email
      Hoja MONEDAS:    col A=nombre

    Ignora filas donde la columna A esté vacía (encabezados u hojas vacías).
    Retorna conteo de filas procesadas por hoja.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl no está instalado")

    from io import BytesIO

    wb = openpyxl.load_workbook(BytesIO(contenido_bytes), read_only=True, data_only=True)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conteo = {"clientes": 0, "atenciones": 0, "monedas": 0}

    def _s(val) -> str:
        return str(val).strip() if val is not None else ""

    # Hoja CLIENTES
    if "CLIENTES" in wb.sheetnames:
        ws = wb["CLIENTES"]
        for row in ws.iter_rows(min_row=2, max_col=5, values_only=True):
            codigo = _s(row[0]) if len(row) > 0 else ""
            if not codigo:
                continue
            nombre = _s(row[1]) if len(row) > 1 else ""
            ruc = _s(row[2]) if len(row) > 2 else ""
            ubicacion = _s(row[3]) if len(row) > 3 else ""
            abreviacion = _s(row[4]) if len(row) > 4 else ""
            conn.execute(
                """INSERT INTO clientes (codigo, nombre, ruc, ubicacion, abreviacion)
                   VALUES (?, ?, ?, ?, ?)
                   ON CONFLICT(codigo) DO UPDATE SET
                     nombre=excluded.nombre,
                     ruc=excluded.ruc,
                     ubicacion=excluded.ubicacion,
                     abreviacion=excluded.abreviacion""",
                (codigo, nombre or codigo, ruc, ubicacion, abreviacion),
            )
            conteo["clientes"] += 1

    # Hoja ATENCIONES
    if "ATENCIONES" in wb.sheetnames:
        ws = wb["ATENCIONES"]
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            nombre = _s(row[0]) if len(row) > 0 else ""
            if not nombre:
                continue
            codigo_empresa = _s(row[1]) if len(row) > 1 else ""
            email = _s(row[2]) if len(row) > 2 else ""
            conn.execute(
                """INSERT INTO atenciones (nombre, codigo_empresa, email)
                   VALUES (?, ?, ?)
                   ON CONFLICT(nombre) DO UPDATE SET
                     codigo_empresa=excluded.codigo_empresa,
                     email=excluded.email""",
                (nombre, codigo_empresa, email),
            )
            conteo["atenciones"] += 1

    # Hoja MONEDAS
    if "MONEDAS" in wb.sheetnames:
        ws = wb["MONEDAS"]
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = _s(row[0]) if len(row) > 0 else ""
            if val:
                conn.execute("INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (val,))
                conteo["monedas"] += 1

    conn.commit()
    conn.close()
    wb.close()
    return conteo


def importar_catalogo_desde_json(ruta: str) -> Dict[str, int]:
    """
    Lee clientes, atenciones y monedas desde un JSON y los inserta en SQLite (INSERT OR IGNORE).
    Retorna conteo de registros procesados.
    """
    if not os.path.exists(ruta):
        raise FileNotFoundError(f"Archivo no encontrado: {ruta}")
    with open(ruta, encoding="utf-8") as f:
        data = json.load(f)
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conteo = {"clientes": 0, "atenciones": 0, "monedas": 0}
    for c in data.get("clientes", []):
        conn.execute(
            "INSERT OR IGNORE INTO clientes (codigo, nombre) VALUES (?, ?)",
            (c["codigo"], c.get("nombre", c["codigo"])),
        )
        conteo["clientes"] += 1
    for a in data.get("atenciones", []):
        conn.execute(
            "INSERT OR IGNORE INTO atenciones (nombre, codigo_empresa) VALUES (?, ?)",
            (a["nombre"], a["codigo_empresa"]),
        )
        conteo["atenciones"] += 1
    for m in data.get("monedas", []):
        conn.execute(
            "INSERT OR IGNORE INTO monedas (nombre) VALUES (?)", (m,)
        )
        conteo["monedas"] += 1
    conn.commit()
    conn.close()
    return conteo


# ─────────────────────────────────────────────
# Carrito persistente
# ─────────────────────────────────────────────

def get_carrito_db(username: str) -> List[Dict]:
    """Devuelve los items del carrito del usuario desde SQLite."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM carrito_items WHERE username=? ORDER BY COALESCE(posicion, id)",
        (username,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_item_carrito_db(username: str, item: Dict, tapa_para_id: Optional[int] = None) -> int:
    """Inserta un item en el carrito. Retorna el id generado."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    # Determinar posicion: siguiente entero después del máximo actual (solo cuerpos)
    max_pos = c.execute(
        "SELECT COALESCE(MAX(COALESCE(posicion, id)), 0) FROM carrito_items WHERE username=? AND tapa_para_id IS NULL",
        (username,),
    ).fetchone()[0]
    if tapa_para_id is not None:
        # Tapa: posicion = posicion del cuerpo + 0.5
        cuerpo_pos = c.execute(
            "SELECT COALESCE(posicion, id) FROM carrito_items WHERE id=?",
            (tapa_para_id,),
        ).fetchone()
        new_pos = float(cuerpo_pos[0]) + 0.5 if cuerpo_pos else float(int(max_pos) + 1)
    else:
        new_pos = float(int(max_pos) + 1)
    c.execute(
        """INSERT INTO carrito_items
           (username, tipo, descripcion, precio_unitario, peso_unitario,
            cantidad, unidad, tipo_galvanizado, porcentaje_ganancia, descripcion_calculada,
            posicion, tapa_para_id, precio_manual)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (
            username,
            item.get("tipo", ""),
            item.get("descripcion", ""),
            item.get("precio_unitario", 0),
            item.get("peso_unitario", 0),
            item.get("cantidad", 1),
            item.get("unidad", "UND"),
            item.get("tipo_galvanizado", "GO"),
            item.get("porcentaje_ganancia", "30"),
            item.get("descripcion_calculada") or None,
            new_pos,
            tapa_para_id,
            1 if item.get("precio_manual") else 0,
        ),
    )
    item_id = c.lastrowid
    conn.commit()
    conn.close()
    return item_id


def update_cantidad_carrito_db(item_id: int, username: str, cantidad: float) -> bool:
    """Actualiza la cantidad de un item. Retorna True si se actualizó."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        "UPDATE carrito_items SET cantidad=? WHERE id=? AND username=?",
        (cantidad, item_id, username),
    )
    updated = c.rowcount > 0
    conn.commit()
    conn.close()
    return updated


def update_item_precio_carrito_db(
    item_id: int, username: str,
    precio_unitario: float, peso_unitario: float,
    descripcion: str, descripcion_calculada: Optional[str] = None,
) -> bool:
    """Actualiza precio, peso y descripción de un item del carrito.
    Si se pasa descripcion_calculada, actualiza ese campo en vez de descripcion.
    Siempre resetea precio_manual a 0 (precio recalculado automáticamente)."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    if descripcion_calculada is not None:
        c.execute(
            "UPDATE carrito_items SET precio_unitario=?, peso_unitario=?, descripcion_calculada=?, precio_manual=0 WHERE id=? AND username=?",
            (round(precio_unitario, 4), round(peso_unitario, 6), descripcion_calculada, item_id, username),
        )
    else:
        c.execute(
            "UPDATE carrito_items SET precio_unitario=?, peso_unitario=?, descripcion=?, precio_manual=0 WHERE id=? AND username=?",
            (round(precio_unitario, 4), round(peso_unitario, 6), descripcion, item_id, username),
        )
    updated = c.rowcount > 0
    conn.commit()
    conn.close()
    return updated


def update_item_campos_carrito_db(
    item_id: int, username: str,
    descripcion: str, unidad: str,
    precio_unitario: float, descripcion_calculada: Optional[str] = None,
) -> bool:
    """Actualiza descripción, unidad, precio y opcionalmente descripcion_calculada de un item."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        """UPDATE carrito_items
           SET descripcion=?, unidad=?, precio_unitario=?, descripcion_calculada=?
           WHERE id=? AND username=?""",
        (descripcion, unidad, round(precio_unitario, 4),
         descripcion_calculada or None, item_id, username),
    )
    updated = c.rowcount > 0
    conn.commit()
    conn.close()
    return updated


def delete_item_carrito_db(item_id: int, username: str) -> bool:
    """Elimina un item del carrito. Retorna True si se eliminó."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        "DELETE FROM carrito_items WHERE id=? AND username=?",
        (item_id, username),
    )
    deleted = c.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


def clear_carrito_db(username: str):
    """Vacía todo el carrito del usuario."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("DELETE FROM carrito_items WHERE username=?", (username,))
    conn.commit()
    conn.close()


def mover_item_carrito_db(item_id: int, username: str, direccion: str) -> bool:
    """Mueve un item cuerpo arriba o abajo en el carrito. Las tapas vinculadas se mueven junto."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    # Solo cuerpos (no tapas separadas)
    cuerpos = conn.execute(
        """SELECT id, COALESCE(posicion, id) AS pos FROM carrito_items
           WHERE username=? AND tapa_para_id IS NULL ORDER BY COALESCE(posicion, id)""",
        (username,),
    ).fetchall()
    cuerpos = [(r["id"], r["pos"]) for r in cuerpos]

    idx = next((i for i, (cid, _) in enumerate(cuerpos) if cid == item_id), None)
    if idx is None:
        conn.close()
        return False
    if direccion == "arriba" and idx == 0:
        conn.close()
        return False
    if direccion == "abajo" and idx == len(cuerpos) - 1:
        conn.close()
        return False

    target_idx = idx - 1 if direccion == "arriba" else idx + 1
    cur_id, cur_pos = cuerpos[idx]
    tgt_id, tgt_pos = cuerpos[target_idx]

    c = conn.cursor()
    # Swap posicion de los dos cuerpos
    c.execute("UPDATE carrito_items SET posicion=? WHERE id=? AND username=?", (tgt_pos, cur_id, username))
    c.execute("UPDATE carrito_items SET posicion=? WHERE id=? AND username=?", (cur_pos, tgt_id, username))
    # Actualizar tapas vinculadas a cada cuerpo
    c.execute("UPDATE carrito_items SET posicion=? WHERE tapa_para_id=? AND username=?", (tgt_pos + 0.5, cur_id, username))
    c.execute("UPDATE carrito_items SET posicion=? WHERE tapa_para_id=? AND username=?", (cur_pos + 0.5, tgt_id, username))
    conn.commit()
    conn.close()
    return True


def reordenar_carrito_db(username: str, ids_ordenados: list) -> bool:
    """Establece la posición de los cuerpos según el orden indicado. Las tapas siguen a su cuerpo."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    for idx, item_id in enumerate(ids_ordenados):
        posicion = (idx + 1) * 10
        c.execute(
            "UPDATE carrito_items SET posicion=? WHERE id=? AND username=? AND tapa_para_id IS NULL",
            (posicion, item_id, username),
        )
        c.execute(
            "UPDATE carrito_items SET posicion=? WHERE tapa_para_id=? AND username=?",
            (posicion + 0.5, item_id, username),
        )
    conn.commit()
    conn.close()
    return True


def update_item_completo_carrito_db(
    item_id: int, username: str,
    descripcion: str, unidad: str,
    precio_unitario: float, peso_unitario: float,
    tipo_galvanizado: str, porcentaje_ganancia: str,
    descripcion_calculada: Optional[str] = None,
    precio_manual: bool = False,
) -> bool:
    """Actualiza todos los campos calculables de un item."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        """UPDATE carrito_items
           SET descripcion=?, unidad=?, precio_unitario=?, peso_unitario=?,
               tipo_galvanizado=?, porcentaje_ganancia=?, descripcion_calculada=?,
               precio_manual=?
           WHERE id=? AND username=?""",
        (descripcion, unidad, round(precio_unitario, 4), round(peso_unitario, 6),
         tipo_galvanizado, porcentaje_ganancia,
         descripcion_calculada or None, 1 if precio_manual else 0,
         item_id, username),
    )
    updated = c.rowcount > 0
    conn.commit()
    conn.close()
    return updated


def cargar_cotizacion_al_carrito_db(cotizacion_id: int, username: str, require_ownership: bool = True) -> Optional[Dict]:
    """
    Limpia el carrito del usuario y copia los items de una cotización guardada.
    Devuelve los metadatos de la cotización para pre-llenar el formulario del carrito.
    Si require_ownership=False (admin), permite cargar cualquier cotización.
    Retorna None si la cotización no existe o (con require_ownership=True) no pertenece al usuario.
    """
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row

    if require_ownership:
        row = conn.execute(
            "SELECT * FROM cotizaciones WHERE id=? AND username=?",
            (cotizacion_id, username),
        ).fetchone()
    else:
        row = conn.execute(
            "SELECT * FROM cotizaciones WHERE id=?",
            (cotizacion_id,),
        ).fetchone()
    if not row:
        conn.close()
        return None

    # Limpiar carrito actual
    conn.execute("DELETE FROM carrito_items WHERE username=?", (username,))

    # Copiar items de cotizacion_items → carrito_items
    items = conn.execute(
        "SELECT * FROM cotizacion_items WHERE cotizacion_id=? ORDER BY id",
        (cotizacion_id,),
    ).fetchall()
    for item in items:
        conn.execute(
            """INSERT INTO carrito_items
               (username, tipo, descripcion, precio_unitario, peso_unitario,
                cantidad, unidad, tipo_galvanizado, porcentaje_ganancia, precio_manual)
               VALUES (?,?,?,?,?,?,?,?,?,?)""",
            (
                username, item["tipo"], item["descripcion"],
                item["precio_unitario"], item["peso_unitario"],
                item["cantidad"], item["unidad"],
                item["tipo_galvanizado"], item["porcentaje_ganancia"],
                item["precio_manual"] if "precio_manual" in item.keys() else 0,
            ),
        )
    conn.commit()

    meta = {
        "id": row["id"],
        "cliente": row["cliente"],
        "cliente_nombre": row["cliente_nombre"],
        "cliente_ruc": row["cliente_ruc"],
        "cliente_ubicacion": row["cliente_ubicacion"],
        "atencion": row["atencion"],
        "atencion_email": row["atencion_email"],
        "proyecto": row["proyecto"],
        "moneda": row["moneda"],
        "validez": row["validez"],
        "encabezado_tabla": row["encabezado_tabla"],
    }
    conn.close()
    return meta


# ─────────────────────────────────────────────
# Historial de cotizaciones
# ─────────────────────────────────────────────

def guardar_cotizacion_db(
    username: str,
    cliente: str,
    atencion: str,
    proyecto: str,
    moneda: str,
    items: List[Dict],
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar_rate: float = 3.8,
    validez: str = "30 días",
    encabezado_tabla: str = "",
) -> int:
    """Guarda una cotización con sus items. Retorna el id de la cotización."""
    from datetime import datetime
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    total_precio = sum(i.get("precio_unitario", 0) * i.get("cantidad", 1) for i in items)
    total_peso = sum(i.get("peso_unitario", 0) * i.get("cantidad", 1) for i in items)

    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA foreign_keys = ON")
    c = conn.cursor()
    c.execute(
        """INSERT INTO cotizaciones
           (username, fecha, cliente, atencion, proyecto, moneda, total_precio, total_peso,
            cliente_nombre, cliente_ruc, cliente_ubicacion, atencion_email,
            dolar_rate, validez, encabezado_tabla)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (username, fecha, cliente, atencion, proyecto, moneda,
         round(total_precio, 4), round(total_peso, 6),
         cliente_nombre, cliente_ruc, cliente_ubicacion, atencion_email,
         dolar_rate, validez, encabezado_tabla),
    )
    cotizacion_id = c.lastrowid
    for item in items:
        c.execute(
            """INSERT INTO cotizacion_items
               (cotizacion_id, tipo, descripcion, precio_unitario, peso_unitario,
                cantidad, unidad, tipo_galvanizado, porcentaje_ganancia, precio_manual)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                cotizacion_id,
                item.get("tipo", ""),
                item.get("descripcion", ""),
                item.get("precio_unitario", 0),
                item.get("peso_unitario", 0),
                item.get("cantidad", 1),
                item.get("unidad", "UND"),
                item.get("tipo_galvanizado", "N/A"),
                item.get("porcentaje_ganancia", "N/A"),
                1 if item.get("precio_manual") else 0,
            ),
        )
    conn.commit()
    conn.close()
    return cotizacion_id


def guardar_cotizacion_importada_db(
    username: str,
    cliente: str,
    atencion: str,
    proyecto: str,
    moneda: str,
    items: List[Dict],
    cliente_nombre: str = "",
    cliente_ruc: str = "",
    cliente_ubicacion: str = "",
    atencion_email: str = "",
    dolar_rate: float = 3.8,
    validez: str = "30 días",
    encabezado_tabla: str = "",
    fecha: Optional[str] = None,
    origen: str = "pdf_import",
) -> int:
    """Guarda una cotización importada (p. ej. desde PDF) con fecha y origen personalizados."""
    from datetime import datetime
    if not fecha:
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    total_precio = sum(i.get("precio_unitario", 0) * i.get("cantidad", 1) for i in items)
    total_peso   = sum(i.get("peso_unitario", 0)  * i.get("cantidad", 1) for i in items)

    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA foreign_keys = ON")
    c = conn.cursor()
    c.execute(
        """INSERT INTO cotizaciones
           (username, fecha, cliente, atencion, proyecto, moneda, total_precio, total_peso,
            cliente_nombre, cliente_ruc, cliente_ubicacion, atencion_email,
            dolar_rate, validez, encabezado_tabla, origen)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (username, fecha, cliente, atencion, proyecto, moneda,
         round(total_precio, 4), round(total_peso, 6),
         cliente_nombre, cliente_ruc, cliente_ubicacion, atencion_email,
         dolar_rate, validez, encabezado_tabla, origen),
    )
    cotizacion_id = c.lastrowid
    for item in items:
        c.execute(
            """INSERT INTO cotizacion_items
               (cotizacion_id, tipo, descripcion, precio_unitario, peso_unitario,
                cantidad, unidad, tipo_galvanizado, porcentaje_ganancia, precio_manual)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
            (
                cotizacion_id,
                item.get("tipo", "MANUAL"),
                item.get("descripcion", ""),
                item.get("precio_unitario", 0),
                item.get("peso_unitario", 0),
                item.get("cantidad", 1),
                item.get("unidad", "UND"),
                item.get("tipo_galvanizado", "N/A"),
                item.get("porcentaje_ganancia", "N/A"),
                1 if item.get("precio_manual") else 0,
            ),
        )
    conn.commit()
    conn.close()
    return cotizacion_id


def listar_cotizaciones_db(
    username: Optional[str] = None,
    tipos: Optional[List[str]] = None,
    q: str = "",
    galvanizados: Optional[List[str]] = None,
    ganancias: Optional[List[str]] = None,
) -> List[Dict]:
    """Lista cotizaciones guardadas con conteo de items.

    Params:
        username     — filtra por usuario (obligatorio en uso normal)
        tipos        — lista de códigos de tipo (B, CH, …); si no está vacía, sólo
                       devuelve cotizaciones que contengan al menos uno de esos tipos
        q            — texto libre; filtra cotizaciones cuya descripción de ítem contenga
                       este texto (case-insensitive)
        galvanizados — lista de valores de galvanizado (GO, GC, N/A); si no está vacía,
                       sólo devuelve cotizaciones que contengan al menos uno de esos tipos
        ganancias    — lista de valores de porcentaje_ganancia (30, 35, N/A); si no está
                       vacía, sólo devuelve cotizaciones que contengan al menos uno
    """
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row

    conditions: List[str] = []
    params: List = []

    if username:
        conditions.append("c.username=?")
        params.append(username)

    if tipos:
        placeholders = ",".join("?" * len(tipos))
        conditions.append(
            f"EXISTS (SELECT 1 FROM cotizacion_items ci2"
            f" WHERE ci2.cotizacion_id = c.id AND ci2.tipo IN ({placeholders}))"
        )
        params.extend(tipos)

    if q:
        palabras = [p.strip() for p in q.split(",") if p.strip()]
        if palabras:
            like_clauses = " AND ".join(
                "LOWER(ci3.descripcion) LIKE ?" for _ in palabras
            )
            conditions.append(
                f"EXISTS (SELECT 1 FROM cotizacion_items ci3"
                f" WHERE ci3.cotizacion_id = c.id AND {like_clauses})"
            )
            params.extend(f"%{p.lower()}%" for p in palabras)

    if galvanizados:
        placeholders = ",".join("?" * len(galvanizados))
        conditions.append(
            f"EXISTS (SELECT 1 FROM cotizacion_items ci4"
            f" WHERE ci4.cotizacion_id = c.id AND ci4.tipo_galvanizado IN ({placeholders}))"
        )
        params.extend(galvanizados)

    if ganancias:
        placeholders = ",".join("?" * len(ganancias))
        conditions.append(
            f"EXISTS (SELECT 1 FROM cotizacion_items ci5"
            f" WHERE ci5.cotizacion_id = c.id AND ci5.porcentaje_ganancia IN ({placeholders}))"
        )
        params.extend(ganancias)

    where = "WHERE " + " AND ".join(conditions) if conditions else ""

    sql = f"""
        SELECT c.*, COUNT(ci.id) AS total_items
        FROM cotizaciones c
        LEFT JOIN cotizacion_items ci ON ci.cotizacion_id = c.id
        {where}
        GROUP BY c.id
        ORDER BY c.id DESC
    """
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_estadisticas_db(username: Optional[str] = None) -> Dict:
    """Devuelve métricas agregadas del historial de cotizaciones.

    Si username es None (admin) devuelve datos globales; si se indica un
    username devuelve únicamente los datos de ese usuario.
    """
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    if username:
        wp = [username]
        wc = "WHERE c.username=? AND (c.origen = 'web' OR c.origen IS NULL)"
    else:
        wp = []
        wc = "WHERE (c.origen = 'web' OR c.origen IS NULL)"

    # Totales globales
    row = conn.execute(
        f"SELECT COUNT(*) AS n, COALESCE(SUM(total_precio),0) AS total "
        f"FROM cotizaciones c {wc}", wp
    ).fetchone()
    total_cots     = row["n"]
    total_facturado = round(float(row["total"]), 2)

    # Por mes (últimos 12, orden cronológico)
    por_mes = list(reversed(conn.execute(f"""
        SELECT strftime('%Y-%m', fecha) AS mes,
               COUNT(*) AS cantidad,
               ROUND(COALESCE(SUM(total_precio),0), 2) AS total
        FROM cotizaciones c {wc}
        GROUP BY mes ORDER BY mes DESC LIMIT 12
    """, wp).fetchall()))

    # Top 5 clientes por cantidad de cotizaciones (excluye sin cliente)
    top_clientes = conn.execute(f"""
        SELECT COALESCE(NULLIF(TRIM(cliente_nombre),''), NULLIF(TRIM(cliente),''), 'Sin cliente') AS nombre,
               COUNT(*) AS cantidad,
               ROUND(COALESCE(SUM(total_precio),0), 2) AS total
        FROM cotizaciones c {wc}
        GROUP BY nombre HAVING nombre != 'Sin cliente' ORDER BY cantidad DESC LIMIT 5
    """, wp).fetchall()

    # Por tipo de producto
    if username:
        por_tipo = conn.execute("""
            SELECT ci.tipo, COUNT(*) AS cantidad
            FROM cotizacion_items ci
            JOIN cotizaciones c ON c.id = ci.cotizacion_id
            WHERE c.username=? AND (c.origen = 'web' OR c.origen IS NULL)
            GROUP BY ci.tipo ORDER BY cantidad DESC
        """, [username]).fetchall()
    else:
        por_tipo = conn.execute("""
            SELECT ci.tipo, COUNT(*) AS cantidad
            FROM cotizacion_items ci
            JOIN cotizaciones c ON c.id = ci.cotizacion_id
            WHERE (c.origen = 'web' OR c.origen IS NULL)
            GROUP BY ci.tipo ORDER BY cantidad DESC
        """).fetchall()

    # Por moneda
    por_moneda = conn.execute(f"""
        SELECT moneda, COUNT(*) AS cantidad, ROUND(COALESCE(SUM(total_precio),0),2) AS total
        FROM cotizaciones c {wc}
        GROUP BY moneda
    """, wp).fetchall()

    conn.close()
    return {
        "total_cotizaciones": total_cots,
        "total_facturado_soles": total_facturado,
        "por_mes":       [dict(r) for r in por_mes],
        "top_clientes":  [dict(r) for r in top_clientes],
        "por_tipo":      [dict(r) for r in por_tipo],
        "por_moneda":    [dict(r) for r in por_moneda],
    }


def _parse_espesor(desc: str) -> str:
    """Extrae el espesor nominal (1.2, 1.5, 2.0) de la descripción de un ítem."""
    d = desc.upper()
    if "1/20" in d:
        return "1.2"
    if "1/16" in d:
        return "1.5"
    m = _re.search(r'\b([12]\.\d)\s*MM', d)
    if m:
        v = m.group(1)
        if v in ("1.2", "1.5", "2.0"):
            return v
    return ""


def get_items_frecuentes_db(
    username: Optional[str] = None,
    cliente: str = "",
    proyecto: str = "",
    limit: int = 40,
) -> List[Dict]:
    """Devuelve ítems del historial ordenados por frecuencia de aparición."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row

    where_parts: List[str] = []
    params: List = []

    if username:
        where_parts.append("c.username = ?")
        params.append(username)
    if cliente.strip():
        like = f"%{cliente.strip()}%"
        where_parts.append("(c.cliente LIKE ? OR c.cliente_nombre LIKE ?)")
        params.extend([like, like])
    if proyecto.strip():
        where_parts.append("c.proyecto LIKE ?")
        params.append(f"%{proyecto.strip()}%")

    where = "WHERE " + " AND ".join(where_parts) if where_parts else ""
    sql = f"""
        SELECT ci.descripcion, ci.tipo, COUNT(*) AS count
        FROM cotizacion_items ci
        JOIN cotizaciones c ON c.id = ci.cotizacion_id
        {where}
        GROUP BY ci.descripcion
        ORDER BY count DESC
        LIMIT ?
    """
    params.append(limit)
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    result = []
    for r in rows:
        d = dict(r)
        d["espesor"] = _parse_espesor(d["descripcion"])
        result.append(d)
    return result


def get_tendencias_items_db(
    clientes: List[str],
    proyecto: str = "",
    q: str = "",
    username: Optional[str] = None,
    tipos: Optional[List[str]] = None,
    galvanizados: Optional[List[str]] = None,
    ganancias: Optional[List[str]] = None,
    monedas: Optional[List[str]] = None,
    espesores: Optional[List[str]] = None,
) -> List[Dict]:
    """Devuelve puntos de precio por ítem para graficar tendencias.

    Para cada cliente en `clientes` ejecuta una consulta y etiqueta los
    resultados con `cliente_idx` (0 = primer cliente, 1 = segundo).
    El precio se normaliza a S/ usando el `dolar_rate` de cada cotización.
    """
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    results: List[Dict] = []

    for idx, cli in enumerate(clientes):
        cli = (cli or "").strip()
        if not cli:
            continue

        where_parts: List[str] = []
        params: List = []

        like_cli = f"%{cli}%"
        where_parts.append("(c.cliente LIKE ? OR c.cliente_nombre LIKE ?)")
        params.extend([like_cli, like_cli])

        if username:
            where_parts.append("c.username = ?")
            params.append(username)

        if proyecto.strip():
            where_parts.append("c.proyecto LIKE ?")
            params.append(f"%{proyecto.strip()}%")

        if q.strip():
            palabras = [p.strip() for p in q.split(",") if p.strip()]
            for p in palabras:
                where_parts.append("LOWER(ci.descripcion) LIKE ?")
                params.append(f"%{p.lower()}%")

        if tipos:
            placeholders = ",".join("?" * len(tipos))
            where_parts.append(f"ci.tipo IN ({placeholders})")
            params.extend(tipos)

        if galvanizados:
            placeholders = ",".join("?" * len(galvanizados))
            where_parts.append(f"ci.tipo_galvanizado IN ({placeholders})")
            params.extend(galvanizados)

        if ganancias:
            placeholders = ",".join("?" * len(ganancias))
            where_parts.append(f"ci.porcentaje_ganancia IN ({placeholders})")
            params.extend(ganancias)

        if monedas:
            placeholders = ",".join("?" * len(monedas))
            where_parts.append(f"c.moneda IN ({placeholders})")
            params.extend(monedas)

        sql = f"""
            SELECT
                c.id AS cotizacion_id,
                c.fecha,
                c.cliente,
                c.cliente_nombre,
                c.proyecto,
                c.moneda,
                c.dolar_rate,
                ci.descripcion,
                ci.precio_unitario,
                ci.tipo,
                ci.tipo_galvanizado
            FROM cotizaciones c
            JOIN cotizacion_items ci ON ci.cotizacion_id = c.id
            {"WHERE " + " AND ".join(where_parts) if where_parts else ""}
            ORDER BY c.fecha ASC, c.id ASC
        """
        rows = conn.execute(sql, params).fetchall()
        for row in rows:
            d = dict(row)
            moneda = d.get("moneda", "SOLES")
            dolar_rate = float(d.get("dolar_rate") or 3.8)
            precio_soles = (
                d["precio_unitario"] if moneda == "SOLES"
                else d["precio_unitario"] * dolar_rate
            )
            d["precio_soles"] = round(precio_soles, 2)
            d["espesor"] = _parse_espesor(d["descripcion"])
            d["galvanizado"] = d.get("tipo_galvanizado", "") or ""
            d["cliente_idx"] = idx
            d["cliente_label"] = d["cliente_nombre"] or d["cliente"] or cli
            if espesores and d["espesor"] not in espesores:
                continue
            results.append(d)

    conn.close()
    return results


def get_cotizacion_db(cotizacion_id: int) -> Optional[Dict]:
    """Devuelve cabecera + items de una cotización guardada."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT * FROM cotizaciones WHERE id=?", (cotizacion_id,)
    ).fetchone()
    if not row:
        conn.close()
        return None
    cotizacion = dict(row)
    items = conn.execute(
        "SELECT * FROM cotizacion_items WHERE cotizacion_id=? ORDER BY id",
        (cotizacion_id,),
    ).fetchall()
    cotizacion["items"] = [dict(i) for i in items]
    conn.close()
    return cotizacion


def eliminar_cotizacion_db(cotizacion_id: int, username: Optional[str]) -> bool:
    """Elimina una cotización. Si username es None (admin), elimina sin restricción de propiedad."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA foreign_keys = ON")
    c = conn.cursor()
    if username is None:
        c.execute("DELETE FROM cotizaciones WHERE id=?", (cotizacion_id,))
    else:
        c.execute(
            "DELETE FROM cotizaciones WHERE id=? AND username=?",
            (cotizacion_id, username),
        )
    deleted = c.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


def eliminar_cotizaciones_bulk_db(ids: List[int]) -> int:
    """Elimina múltiples cotizaciones (admin). Retorna cantidad eliminada."""
    if not ids:
        return 0
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("PRAGMA foreign_keys = ON")
    placeholders = ",".join("?" * len(ids))
    c = conn.cursor()
    c.execute(f"DELETE FROM cotizaciones WHERE id IN ({placeholders})", ids)
    deleted = c.rowcount
    conn.commit()
    conn.close()
    return deleted


def _fp_items(items: list) -> str:
    """Fingerprint estable de una lista de ítems: hash MD5 de (descripcion, cantidad) ordenados."""
    import hashlib
    normalized = sorted(
        (str(i.get("descripcion", "")).strip().lower(),
         round(float(i.get("cantidad") or 1), 4))
        for i in items
    )
    raw = "|".join(f"{d}:{c}" for d, c in normalized)
    return hashlib.md5(raw.encode()).hexdigest()[:12]


def detectar_duplicados_db() -> List[Dict]:
    """
    Agrupa cotizaciones por (cliente_norm, proyecto_norm, total_redondeado, items_fp)
    y retorna solo los grupos con 2+ entradas. Ignora filas sin cliente y sin proyecto.
    """
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row

    rows = conn.execute("""
        SELECT c.id, c.username, c.fecha, c.cliente, c.proyecto, c.moneda,
               c.cliente_nombre, c.total_precio, c.dolar_rate, c.origen,
               COUNT(ci.id) AS total_items
        FROM cotizaciones c
        LEFT JOIN cotizacion_items ci ON ci.cotizacion_id = c.id
        GROUP BY c.id
        ORDER BY c.id ASC
    """).fetchall()

    item_rows = conn.execute("""
        SELECT cotizacion_id, descripcion, cantidad
        FROM cotizacion_items
    """).fetchall()
    conn.close()

    from collections import defaultdict
    items_by_cot: Dict[int, list] = defaultdict(list)
    for ir in item_rows:
        items_by_cot[ir["cotizacion_id"]].append({
            "descripcion": ir["descripcion"],
            "cantidad":    ir["cantidad"],
        })

    grupos: Dict[tuple, list] = defaultdict(list)
    for r in rows:
        cliente_n  = (r["cliente_nombre"] or r["cliente"] or "").strip().lower()
        proyecto_n = (r["proyecto"] or "").strip().lower()
        if not cliente_n and not proyecto_n:
            continue
        total_norm = round(float(r["total_precio"] or 0), 2)
        items_fp   = _fp_items(items_by_cot.get(r["id"], []))
        clave = (cliente_n, proyecto_n, total_norm, items_fp)
        grupos[clave].append(dict(r))

    return [
        {
            "cliente":  grupo[0].get("cliente_nombre") or grupo[0].get("cliente") or "",
            "proyecto": grupo[0].get("proyecto") or "",
            "total":    round(float(grupo[0].get("total_precio") or 0), 2),
            "cotizaciones": grupo,
        }
        for grupo in grupos.values()
        if len(grupo) >= 2
    ]


def fingerprints_cotizaciones_db() -> Dict[str, List[int]]:
    """
    Retorna dict fingerprint → [ids] para todas las cotizaciones.
    Fingerprint = 'cliente_norm|proyecto_norm|total_redondeado|items_fp'.
    Usado para detectar posibles duplicados durante la importación de PDFs.
    """
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row

    rows = conn.execute("""
        SELECT id, cliente_nombre, cliente, proyecto, total_precio
        FROM cotizaciones
    """).fetchall()

    item_rows = conn.execute("""
        SELECT cotizacion_id, descripcion, cantidad
        FROM cotizacion_items
    """).fetchall()
    conn.close()

    from collections import defaultdict
    items_by_cot: Dict[int, list] = defaultdict(list)
    for ir in item_rows:
        items_by_cot[ir["cotizacion_id"]].append({
            "descripcion": ir["descripcion"],
            "cantidad":    ir["cantidad"],
        })

    result: Dict[str, List[int]] = {}
    for r in rows:
        cliente  = (r["cliente_nombre"] or r["cliente"] or "").strip().lower()
        proyecto = (r["proyecto"] or "").strip().lower()
        total    = round(float(r["total_precio"] or 0), 2)
        items_fp = _fp_items(items_by_cot.get(r["id"], []))
        fp = f"{cliente}|{proyecto}|{total}|{items_fp}"
        result.setdefault(fp, []).append(r["id"])
    return result


# ─────────────────────────────────────────────
# Proyectos (Kanban de obras)
# ─────────────────────────────────────────────

ESTADOS_KANBAN = ['APROBADO', 'EN_PRODUCCION', 'DESPACHADO']
ESTADO_LABELS = {
    'APROBADO':      'Aprobado con OC',
    'EN_PRODUCCION': 'En Producción',
    'DESPACHADO':    'Despachado',
}

ADJUNTOS_DIR = BASE_DIR / "data" / "adjuntos"


def init_proyectos():
    """Crea tablas proyectos + proyecto_adjuntos + proyecto_oc_items."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS proyectos (
            nombre TEXT PRIMARY KEY,
            estado TEXT NOT NULL DEFAULT 'APROBADO',
            updated_at TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS proyecto_adjuntos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto TEXT NOT NULL,
            filename TEXT NOT NULL,
            filepath TEXT NOT NULL,
            content_type TEXT NOT NULL DEFAULT '',
            uploaded_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS proyecto_oc_items (
            id                  INTEGER PRIMARY KEY AUTOINCREMENT,
            proyecto            TEXT NOT NULL,
            descripcion         TEXT NOT NULL DEFAULT '',
            unidad              TEXT NOT NULL DEFAULT 'UND',
            cantidad_pedida     REAL NOT NULL DEFAULT 0,
            cantidad_despachada REAL NOT NULL DEFAULT 0,
            orden               INTEGER NOT NULL DEFAULT 0
        )
    """)
    # Migraciones de columnas
    _add_column_if_missing(conn, "proyectos", "direccion",     "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyectos", "cliente",       "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyectos", "numero_oc",     "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyectos", "created_at",    "TEXT")
    _add_column_if_missing(conn, "proyectos", "contacto",      "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyectos", "lugar_entrega", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyectos", "fecha_entrega", "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyectos", "fecha_oc",      "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyectos", "notas",         "TEXT NOT NULL DEFAULT ''")
    _add_column_if_missing(conn, "proyecto_adjuntos", "categoria", "TEXT NOT NULL DEFAULT 'oc'")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_proyectos_estado ON proyectos(estado)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_proyecto_adjuntos_proyecto ON proyecto_adjuntos(proyecto)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_proyecto_oc_items_proyecto ON proyecto_oc_items(proyecto)")
    conn.commit()
    conn.close()
    ADJUNTOS_DIR.mkdir(parents=True, exist_ok=True)


def get_proyectos_con_stats() -> List[Dict]:
    """Lista todos los proyectos con estadísticas, conteo de adjuntos e items OC."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    rows = conn.execute("""
        SELECT
            p.nombre,
            p.estado,
            p.cliente,
            p.numero_oc,
            p.direccion,
            p.updated_at,
            p.created_at,
            p.contacto,
            p.lugar_entrega,
            p.fecha_entrega,
            p.fecha_oc,
            COALESCE(p.notas, '') AS notas,
            COALESCE(SUM(c.total_precio), 0) AS total_cotizado,
            MAX(c.fecha) AS ultima_fecha,
            (SELECT c2.cliente_nombre FROM cotizaciones c2
             WHERE c2.proyecto = p.nombre ORDER BY c2.fecha DESC LIMIT 1) AS cliente_nombre,
            (SELECT COUNT(*) FROM proyecto_adjuntos pa
             WHERE pa.proyecto = p.nombre AND pa.categoria = 'oc') AS n_oc_docs,
            (SELECT COUNT(*) FROM proyecto_adjuntos pa
             WHERE pa.proyecto = p.nombre AND pa.categoria = 'ev') AS n_evidencia,
            (SELECT COUNT(*) FROM proyecto_oc_items oi
             WHERE oi.proyecto = p.nombre) AS n_items,
            (SELECT COUNT(*)
             FROM proyecto_oc_items oi2
             WHERE oi2.proyecto = p.nombre
               AND oi2.cantidad_pedida > oi2.cantidad_despachada) AS items_pendientes
        FROM proyectos p
        LEFT JOIN cotizaciones c ON c.proyecto = p.nombre
        GROUP BY p.nombre, p.estado, p.cliente, p.direccion
        ORDER BY p.nombre
    """).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def set_proyecto_estado(nombre: str, estado: str):
    """Actualiza el estado Kanban de un proyecto."""
    from datetime import datetime
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        "UPDATE proyectos SET estado=?, updated_at=? WHERE nombre=?",
        (estado, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), nombre),
    )
    conn.commit()
    conn.close()


def get_kpis_proyectos() -> Dict:
    """KPIs del pipeline: proyectos aprobados, en producción, despachados e items pendientes."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    row = conn.execute("""
        SELECT
            COUNT(CASE WHEN p.estado = 'APROBADO'      THEN 1 END) AS n_aprobado,
            COUNT(CASE WHEN p.estado = 'EN_PRODUCCION' THEN 1 END) AS n_produccion,
            COUNT(CASE WHEN p.estado = 'DESPACHADO'    THEN 1 END) AS n_despachado,
            COALESCE(SUM(CASE WHEN p.estado = 'APROBADO'
                THEN COALESCE(c_sum.total, 0) ELSE 0 END), 0) AS total_aprobado,
            (SELECT COUNT(*)
             FROM proyecto_oc_items oi
             JOIN proyectos pj ON pj.nombre = oi.proyecto
             WHERE pj.estado != 'DESPACHADO'
               AND oi.cantidad_pedida > oi.cantidad_despachada) AS items_pendientes
        FROM proyectos p
        LEFT JOIN (
            SELECT proyecto, SUM(total_precio) AS total
            FROM cotizaciones GROUP BY proyecto
        ) c_sum ON c_sum.proyecto = p.nombre
    """).fetchone()
    conn.close()
    return dict(row) if row else {
        "n_aprobado": 0, "total_aprobado": 0.0, "n_produccion": 0, "n_despachado": 0,
        "items_pendientes": 0,
    }


def update_proyecto_direccion(nombre: str, direccion: str):
    """Actualiza la dirección manual de la obra."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        "UPDATE proyectos SET direccion=? WHERE nombre=?",
        (direccion.strip(), nombre),
    )
    conn.commit()
    conn.close()


def update_proyecto_notas(nombre: str, notas: str):
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute("UPDATE proyectos SET notas=? WHERE nombre=?", (notas.strip(), nombre))
    conn.commit()
    conn.close()


def update_proyecto_contacto(nombre: str, contacto: str):
    """Actualiza el contacto de la obra."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        "UPDATE proyectos SET contacto=? WHERE nombre=?",
        (contacto.strip(), nombre),
    )
    conn.commit()
    conn.close()


def update_proyecto_numero_oc(nombre: str, numero_oc: str):
    """Actualiza el número de orden de compra de la obra."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        "UPDATE proyectos SET numero_oc=? WHERE nombre=?",
        (numero_oc.strip(), nombre),
    )
    conn.commit()
    conn.close()


# ── Adjuntos ──

def add_adjunto(proyecto: str, filename: str, filepath: str, content_type: str = "", categoria: str = "oc") -> int:
    from datetime import datetime
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        """INSERT INTO proyecto_adjuntos (proyecto, filename, filepath, content_type, uploaded_at, categoria)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (proyecto, filename, filepath, content_type, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), categoria),
    )
    adj_id = c.lastrowid
    conn.commit()
    conn.close()
    return adj_id


def list_adjuntos(proyecto: str) -> List[Dict]:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT id, proyecto, filename, content_type, uploaded_at, categoria FROM proyecto_adjuntos"
        " WHERE proyecto=? ORDER BY uploaded_at DESC",
        (proyecto,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_adjunto_filepath(adjunto_id: int, proyecto: str) -> Optional[dict]:
    """Devuelve el filepath de un adjunto verificando que pertenezca al proyecto."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    row = conn.execute(
        "SELECT filepath, filename, content_type FROM proyecto_adjuntos WHERE id=? AND proyecto=?",
        (adjunto_id, proyecto),
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def delete_adjunto(adjunto_id: int) -> Optional[str]:
    """Elimina el registro y retorna el filepath para borrar el archivo físico."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    row = conn.execute(
        "SELECT filepath FROM proyecto_adjuntos WHERE id=?", (adjunto_id,)
    ).fetchone()
    if not row:
        conn.close()
        return None
    filepath = row[0]
    conn.execute("DELETE FROM proyecto_adjuntos WHERE id=?", (adjunto_id,))
    conn.commit()
    conn.close()
    return filepath


# ── Proyectos: creación/eliminación manual ──

def renombrar_proyecto(nombre: str, nuevo_nombre: str, nuevo_cliente: str) -> bool:
    """Renombra un proyecto y actualiza el cliente. Cascada manual a tablas relacionadas."""
    nuevo_nombre = nuevo_nombre.strip()
    nuevo_cliente = nuevo_cliente.strip()
    if not nuevo_nombre:
        return False
    conn = sqlite3.connect(DB_PATH, timeout=10)
    try:
        conn.execute("BEGIN")
        if nuevo_nombre != nombre:
            # Verificar que el nuevo nombre no exista ya
            existe = conn.execute(
                "SELECT 1 FROM proyectos WHERE nombre=?", (nuevo_nombre,)
            ).fetchone()
            if existe:
                conn.rollback()
                conn.close()
                return False
            conn.execute("UPDATE cotizaciones SET proyecto=? WHERE proyecto=?", (nuevo_nombre, nombre))
            conn.execute("UPDATE proyecto_adjuntos SET proyecto=? WHERE proyecto=?", (nuevo_nombre, nombre))
            conn.execute("UPDATE proyecto_oc_items SET proyecto=? WHERE proyecto=?", (nuevo_nombre, nombre))
            conn.execute("UPDATE proyectos SET nombre=?, cliente=? WHERE nombre=?", (nuevo_nombre, nuevo_cliente, nombre))
        else:
            conn.execute("UPDATE proyectos SET cliente=? WHERE nombre=?", (nuevo_cliente, nombre))
        conn.commit()
    except Exception:
        conn.rollback()
        conn.close()
        return False
    conn.close()
    return True


def crear_proyecto(
    nombre: str,
    cliente: str,
    lugar_entrega: str = "",
    fecha_entrega: str = "",
    fecha_oc: str = "",
) -> bool:
    """Crea un proyecto en estado APROBADO. Retorna False si ya existe."""
    from datetime import datetime
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        """INSERT OR IGNORE INTO proyectos
           (nombre, cliente, estado, created_at, lugar_entrega, fecha_entrega, fecha_oc)
           VALUES (?, ?, 'APROBADO', ?, ?, ?, ?)""",
        (nombre.strip(), cliente.strip(), datetime.now().strftime("%Y-%m-%d"),
         lugar_entrega.strip(), fecha_entrega.strip(), fecha_oc.strip()),
    )
    created = c.rowcount > 0
    conn.commit()
    conn.close()
    return created


def proyecto_existe(nombre: str) -> bool:
    """Retorna True si ya hay un proyecto con ese nombre exacto."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    row = conn.execute("SELECT 1 FROM proyectos WHERE nombre=?", (nombre.strip(),)).fetchone()
    conn.close()
    return row is not None


def eliminar_proyecto(nombre: str) -> bool:
    """Elimina un proyecto, sus adjuntos físicos y sus items OC."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    adjuntos = conn.execute(
        "SELECT filepath FROM proyecto_adjuntos WHERE proyecto=?", (nombre,)
    ).fetchall()
    for adj in adjuntos:
        try:
            Path(adj["filepath"]).unlink(missing_ok=True)
        except Exception:
            pass
    conn.execute("DELETE FROM proyecto_adjuntos WHERE proyecto=?", (nombre,))
    conn.execute("DELETE FROM proyecto_oc_items WHERE proyecto=?", (nombre,))
    conn.execute("DELETE FROM email_importados WHERE proyecto=?", (nombre,))
    c_del = conn.execute("DELETE FROM proyectos WHERE nombre=?", (nombre,))
    deleted = c_del.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


# ── OC Items CRUD ──

def get_oc_items(proyecto: str) -> List[Dict]:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    rows = conn.execute(
        "SELECT * FROM proyecto_oc_items WHERE proyecto=? ORDER BY orden, id",
        (proyecto,),
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_oc_item(
    proyecto: str,
    descripcion: str,
    unidad: str,
    cantidad_pedida: float,
    cantidad_despachada: float,
    orden: int = 0,
) -> int:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        """INSERT INTO proyecto_oc_items
           (proyecto, descripcion, unidad, cantidad_pedida, cantidad_despachada, orden)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (proyecto, descripcion, unidad, cantidad_pedida, cantidad_despachada, orden),
    )
    item_id = c.lastrowid
    conn.commit()
    conn.close()
    return item_id


def update_oc_item(
    item_id: int,
    proyecto: str,
    descripcion: str,
    unidad: str,
    cantidad_pedida: float,
    cantidad_despachada: float,
) -> bool:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        """UPDATE proyecto_oc_items
           SET descripcion=?, unidad=?, cantidad_pedida=?, cantidad_despachada=?
           WHERE id=? AND proyecto=?""",
        (descripcion, unidad, cantidad_pedida, cantidad_despachada, item_id, proyecto),
    )
    updated = c.rowcount > 0
    conn.commit()
    conn.close()
    return updated


def delete_oc_item(item_id: int, proyecto: str) -> bool:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        "DELETE FROM proyecto_oc_items WHERE id=? AND proyecto=?",
        (item_id, proyecto),
    )
    deleted = c.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


# ─────────────────────────────────────────────
# Reportes de asistencia
# ─────────────────────────────────────────────

def guardar_reporte_asistencia(
    periodo: str,
    periodo_label: str,
    subido_por: str,
    nombre_archivo: str,
    num_empleados: int,
    datos_json_str: str,
) -> int:
    """Inserta un reporte procesado. Devuelve el id insertado."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute(
        """INSERT INTO reportes_asistencia
           (periodo, periodo_label, fecha_subida, subido_por, nombre_archivo, num_empleados, datos_json)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (periodo, periodo_label, _dt.now().isoformat(timespec="seconds"),
         subido_por, nombre_archivo, num_empleados, datos_json_str),
    )
    nuevo_id = c.lastrowid
    conn.commit()
    conn.close()
    return nuevo_id


def listar_reportes_asistencia() -> List[Dict]:
    """Devuelve todos los reportes sin datos_json, ordenados por fecha_subida DESC."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute(
        """SELECT id, periodo, periodo_label, fecha_subida, subido_por, nombre_archivo, num_empleados
           FROM reportes_asistencia ORDER BY periodo ASC"""
    )
    rows = [dict(r) for r in c.fetchall()]
    conn.close()
    return rows


def obtener_reporte_asistencia(reporte_id: int) -> Optional[Dict]:
    """Devuelve el reporte completo incluyendo datos_json, o None si no existe."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT * FROM reportes_asistencia WHERE id=?", (reporte_id,))
    row = c.fetchone()
    conn.close()
    return dict(row) if row else None


def eliminar_reporte_asistencia(reporte_id: int) -> bool:
    """Elimina el reporte de DB. Devuelve True si existía."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute("DELETE FROM reportes_asistencia WHERE id=?", (reporte_id,))
    deleted = c.rowcount > 0
    conn.commit()
    conn.close()
    return deleted


def periodo_existe(periodo: str) -> Optional[int]:
    """Devuelve el id del reporte si el período ya existe, o None."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    c = conn.cursor()
    c.execute("SELECT id FROM reportes_asistencia WHERE periodo=?", (periodo,))
    row = c.fetchone()
    conn.close()
    return row[0] if row else None


def buscar_reporte_por_mes(anio: int, mes: int) -> Optional[Dict]:
    """Busca un reporte existente que cubra el mismo mes/año (quincena o mes completo)."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    patron = f"{anio:04d}-{mes:02d}-%"
    c.execute("SELECT * FROM reportes_asistencia WHERE periodo LIKE ?", (patron,))
    row = c.fetchone()
    conn.close()
    return dict(row) if row else None


def actualizar_reporte_asistencia(
    reporte_id: int,
    nombre_archivo: str,
    subido_por: str,
    num_empleados: int,
    datos_json_str: str,
):
    """Sobreescribe datos de un reporte existente (para duplicados confirmados)."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        """UPDATE reportes_asistencia
           SET nombre_archivo=?, subido_por=?, fecha_subida=?, num_empleados=?, datos_json=?
           WHERE id=?""",
        (nombre_archivo, subido_por, _dt.now().isoformat(timespec="seconds"),
         num_empleados, datos_json_str, reporte_id),
    )
    conn.commit()
    conn.close()


def fusionar_reporte_asistencia(
    reporte_id: int,
    periodo: str,
    periodo_label: str,
    subido_por: str,
    num_empleados: int,
    datos_json_str: str,
) -> None:
    """Actualiza un reporte con datos fusionados de dos quincenas del mismo mes."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        """UPDATE reportes_asistencia
           SET periodo=?, periodo_label=?, subido_por=?, fecha_subida=?,
               num_empleados=?, datos_json=?
           WHERE id=?""",
        (periodo, periodo_label, subido_por, _dt.now().isoformat(timespec="seconds"),
         num_empleados, datos_json_str, reporte_id),
    )
    conn.commit()
    conn.close()


# ── IMAP / correo electrónico ────────────────────────────────────────────────

def get_email_imap_config() -> Optional[Dict]:
    """Devuelve la configuración IMAP guardada, o None si aún no está configurada."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    row = conn.execute("SELECT * FROM email_imap_config WHERE id=1").fetchone()
    conn.close()
    if not row or not row["host"]:
        return None
    return dict(row)


def save_email_imap_config(host: str, port: int, username: str, password: str,
                            folder: str, days_back: int) -> None:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        """INSERT INTO email_imap_config (id, host, port, username, password, folder, days_back)
           VALUES (1, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(id) DO UPDATE SET
               host=excluded.host, port=excluded.port, username=excluded.username,
               password=excluded.password, folder=excluded.folder, days_back=excluded.days_back""",
        (host, port, username, password, folder, days_back),
    )
    conn.commit()
    conn.close()


def email_ya_importado(message_id: str) -> bool:
    """True si el message_id ya fue importado y el proyecto todavía existe."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    row = conn.execute(
        """SELECT ei.proyecto FROM email_importados ei
           JOIN proyectos p ON p.nombre = ei.proyecto
           WHERE ei.message_id = ?""",
        (message_id,),
    ).fetchone()
    conn.close()
    return row is not None


def pdf_hash_ya_importado(pdf_hash: str) -> bool:
    """True si ya se importó un PDF con ese hash SHA-256 y el proyecto aún existe.
    Permite detectar reenvíos/respuestas que adjuntan el mismo PDF."""
    if not pdf_hash:
        return False
    conn = sqlite3.connect(DB_PATH, timeout=10)
    row = conn.execute(
        """SELECT ei.proyecto FROM email_importados ei
           JOIN proyectos p ON p.nombre = ei.proyecto
           WHERE ei.pdf_hash = ?""",
        (pdf_hash,),
    ).fetchone()
    conn.close()
    return row is not None


def registrar_email_importado(message_id: str, proyecto: str, pdf_hash: str = "") -> None:
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.execute(
        """INSERT OR IGNORE INTO email_importados (message_id, proyecto, importado_at, pdf_hash)
           VALUES (?, ?, ?, ?)""",
        (message_id, proyecto, _dt.now().strftime("%Y-%m-%d %H:%M:%S"), pdf_hash),
    )
    conn.commit()
    conn.close()


def get_dominios_clientes() -> set:
    """Devuelve el conjunto de dominios de correo de las atenciones registradas (ej. {'empresa.com'})."""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    rows = conn.execute("SELECT email FROM atenciones WHERE email != ''").fetchall()
    conn.close()
    dominios = set()
    for (email_addr,) in rows:
        if "@" in email_addr:
            domain = email_addr.split("@", 1)[1].strip().lower()
            if domain:
                dominios.add(domain)
    return dominios


def get_cliente_nombre_por_dominio(domain: str) -> str:
    """Devuelve el nombre del cliente registrado con ese dominio de correo, o '' si no existe."""
    if not domain:
        return ""
    conn = sqlite3.connect(DB_PATH, timeout=10)
    row = conn.execute(
        """SELECT c.nombre FROM atenciones a
           JOIN clientes c ON c.codigo = a.codigo_empresa
           WHERE lower(a.email) LIKE ? LIMIT 1""",
        (f"%@{domain.lower()}",),
    ).fetchone()
    conn.close()
    return row[0] if row else ""


# Inicializar al importar
init_db()
init_proyectos()
