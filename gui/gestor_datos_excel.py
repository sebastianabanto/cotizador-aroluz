# gestor_datos_excel.py
"""
Sistema completo de gestión de clientes y atenciones para Excel
Permite crear, leer, actualizar y eliminar (CRUD) registros directamente desde la GUI
"""

import os
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from typing import List, Dict, Optional, Tuple
import re


class GestorDatosExcel:
    """Clase principal para gestionar datos de clientes y atenciones en Excel"""
    
    def __init__(self, ruta_excel: str):
        self.ruta_excel = ruta_excel
        self.datos_clientes = []  # Lista de diccionarios con datos de clientes
        self.datos_atenciones = []  # Lista de diccionarios con datos de atenciones
        
    def verificar_archivo(self) -> bool:
        """Verifica que el archivo Excel exista"""
        return os.path.exists(self.ruta_excel)
    
    def cargar_datos_completos(self) -> bool:
        """Carga todos los datos de clientes y atenciones desde Excel"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Se requiere openpyxl. Instale con: pip install openpyxl")
        
        if not self.verificar_archivo():
            raise FileNotFoundError(f"No se encontró el archivo Excel en: {self.ruta_excel}")
        
        try:
            # Abrir el libro
            workbook = openpyxl.load_workbook(self.ruta_excel, data_only=True)
            
            # Cargar clientes
            self._cargar_clientes(workbook)
            
            # Cargar atenciones
            self._cargar_atenciones(workbook)
            
            workbook.close()
            return True
            
        except Exception as e:
            raise Exception(f"Error cargando datos: {str(e)}")
    
    def _cargar_clientes(self, workbook):
        """Carga datos de la hoja CLIENTES"""
        if "CLIENTES" not in workbook.sheetnames:
            raise ValueError("La hoja 'CLIENTES' no existe")
        
        worksheet = workbook["CLIENTES"]
        self.datos_clientes = []
        
        # Leer desde la fila 2 (asumiendo encabezados en fila 1)
        for fila in range(2, worksheet.max_row + 1):
            codigo = self._obtener_valor_celda(worksheet, f"A{fila}")
            razon_social = self._obtener_valor_celda(worksheet, f"B{fila}")
            ruc = self._obtener_valor_celda(worksheet, f"C{fila}")
            ubicacion = self._obtener_valor_celda(worksheet, f"D{fila}")
            
            # Solo agregar si al menos tiene código o razón social
            if codigo or razon_social:
                cliente = {
                    'fila': fila,
                    'codigo': codigo,
                    'razon_social': razon_social,
                    'ruc': ruc,
                    'ubicacion': ubicacion
                }
                self.datos_clientes.append(cliente)
    
    def _cargar_atenciones(self, workbook):
        """Carga datos de la hoja ATENCIÓN"""
        if "ATENCIÓN" not in workbook.sheetnames:
            raise ValueError("La hoja 'ATENCIÓN' no existe")
        
        worksheet = workbook["ATENCIÓN"]
        self.datos_atenciones = []
        
        # Leer desde la fila 2 (asumiendo encabezados en fila 1)
        for fila in range(2, worksheet.max_row + 1):
            codigo = self._obtener_valor_celda(worksheet, f"A{fila}")
            nombres = self._obtener_valor_celda(worksheet, f"B{fila}")
            correo = self._obtener_valor_celda(worksheet, f"C{fila}")
            celular = self._obtener_valor_celda(worksheet, f"D{fila}")
            razon_social = self._obtener_valor_celda(worksheet, f"E{fila}")
            
            # Solo agregar si al menos tiene código o nombres
            if codigo or nombres:
                atencion = {
                    'fila': fila,
                    'codigo': codigo,
                    'nombres': nombres,
                    'correo': correo,
                    'celular': celular,
                    'razon_social': razon_social
                }
                self.datos_atenciones.append(atencion)
    
    def _obtener_valor_celda(self, worksheet, celda) -> str:
        """Obtiene el valor de una celda de forma segura"""
        try:
            valor = worksheet[celda].value
            return str(valor).strip() if valor is not None else ""
        except:
            return ""
    
    def guardar_cambios(self) -> bool:
        """Guarda todos los cambios en el Excel"""
        try:
            import openpyxl
            
            # Hacer backup antes de guardar
            self._crear_backup()
            
            # Abrir el libro para escritura
            workbook = openpyxl.load_workbook(self.ruta_excel)
            
            # Guardar clientes
            self._guardar_clientes(workbook)
            
            # Guardar atenciones
            self._guardar_atenciones(workbook)
            
            # Guardar archivo
            workbook.save(self.ruta_excel)
            workbook.close()
            
            return True
            
        except Exception as e:
            raise Exception(f"Error guardando cambios: {str(e)}")
    
    def _crear_backup(self):
        """Crea un backup del archivo Excel antes de modificarlo"""
        import shutil
        from datetime import datetime
        
        nombre_base = os.path.splitext(self.ruta_excel)[0]
        extension = os.path.splitext(self.ruta_excel)[1]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_backup = f"{nombre_base}_backup_{timestamp}{extension}"
        
        shutil.copy2(self.ruta_excel, ruta_backup)
        print(f"📄 Backup creado: {ruta_backup}")
    
    def _guardar_clientes(self, workbook):
        """Guarda datos de clientes en la hoja CLIENTES"""
        worksheet = workbook["CLIENTES"]
        
        # Limpiar datos existentes (desde fila 2)
        for fila in range(2, worksheet.max_row + 1):
            for col in ['A', 'B', 'C', 'D']:
                worksheet[f"{col}{fila}"].value = None
        
        # Escribir nuevos datos
        for i, cliente in enumerate(self.datos_clientes, start=2):
            worksheet[f"A{i}"].value = cliente['codigo']
            worksheet[f"B{i}"].value = cliente['razon_social']
            worksheet[f"C{i}"].value = cliente['ruc']
            worksheet[f"D{i}"].value = cliente['ubicacion']
    
    def _guardar_atenciones(self, workbook):
        """Guarda datos de atenciones en la hoja ATENCIÓN"""
        worksheet = workbook["ATENCIÓN"]
        
        # Limpiar datos existentes (desde fila 2)
        for fila in range(2, worksheet.max_row + 1):
            for col in ['A', 'B', 'C', 'D', 'E']:
                worksheet[f"{col}{fila}"].value = None
        
        # Escribir nuevos datos
        for i, atencion in enumerate(self.datos_atenciones, start=2):
            worksheet[f"A{i}"].value = atencion['codigo']
            worksheet[f"B{i}"].value = atencion['nombres']
            worksheet[f"C{i}"].value = atencion['correo']
            worksheet[f"D{i}"].value = atencion['celular']
            worksheet[f"E{i}"].value = atencion['razon_social']
    
    # MÉTODOS CRUD PARA CLIENTES
    def agregar_cliente(self, codigo: str, razon_social: str, ruc: str = "", ubicacion: str = "") -> bool:
        """Agrega un nuevo cliente"""
        # Validar que no exista el código
        if any(c['codigo'].upper() == codigo.upper() for c in self.datos_clientes):
            raise ValueError(f"Ya existe un cliente con código '{codigo}'")
        
        # Validar RUC si se proporciona
        if ruc and not self._validar_ruc(ruc):
            raise ValueError("El RUC no tiene un formato válido")
        
        nuevo_cliente = {
            'fila': 0,  # Se asignará al guardar
            'codigo': codigo.strip(),
            'razon_social': razon_social.strip(),
            'ruc': ruc.strip(),
            'ubicacion': ubicacion.strip()
        }
        
        self.datos_clientes.append(nuevo_cliente)
        return True
    
    def modificar_cliente(self, indice: int, codigo: str, razon_social: str, ruc: str = "", ubicacion: str = "") -> bool:
        """Modifica un cliente existente"""
        if not (0 <= indice < len(self.datos_clientes)):
            raise ValueError("Índice de cliente inválido")
        
        cliente_actual = self.datos_clientes[indice]
        
        # Validar que el código no esté duplicado (excepto si es el mismo)
        if codigo.upper() != cliente_actual['codigo'].upper():
            if any(c['codigo'].upper() == codigo.upper() for c in self.datos_clientes):
                raise ValueError(f"Ya existe otro cliente con código '{codigo}'")
        
        # Validar RUC si se proporciona
        if ruc and not self._validar_ruc(ruc):
            raise ValueError("El RUC no tiene un formato válido")
        
        # Actualizar datos
        self.datos_clientes[indice].update({
            'codigo': codigo.strip(),
            'razon_social': razon_social.strip(),
            'ruc': ruc.strip(),
            'ubicacion': ubicacion.strip()
        })
        
        # Actualizar referencias en atenciones
        self._actualizar_referencias_atencion(cliente_actual['codigo'], codigo.strip())
        
        return True
    
    def eliminar_cliente(self, indice: int) -> bool:
        """Elimina un cliente (y sus atenciones asociadas)"""
        if not (0 <= indice < len(self.datos_clientes)):
            raise ValueError("Índice de cliente inválido")
        
        cliente = self.datos_clientes[indice]
        codigo_cliente = cliente['codigo']
        
        # Verificar si tiene atenciones asociadas
        atenciones_asociadas = [a for a in self.datos_atenciones if a['razon_social'] == codigo_cliente]
        
        if atenciones_asociadas:
            # Preguntar qué hacer con las atenciones
            return False, f"El cliente tiene {len(atenciones_asociadas)} atención(es) asociada(s)"
        
        # Eliminar cliente
        del self.datos_clientes[indice]
        return True
    
    def _actualizar_referencias_atencion(self, codigo_anterior: str, codigo_nuevo: str):
        """Actualiza las referencias en atenciones cuando se cambia el código de un cliente"""
        for atencion in self.datos_atenciones:
            if atencion['razon_social'] == codigo_anterior:
                atencion['razon_social'] = codigo_nuevo
    
    # MÉTODOS CRUD PARA ATENCIONES
    def agregar_atencion(self, codigo: str, nombres: str, correo: str = "", celular: str = "", razon_social: str = "") -> bool:
        """Agrega una nueva atención"""
        # Validar que no exista el código
        if any(a['codigo'].upper() == codigo.upper() for a in self.datos_atenciones):
            raise ValueError(f"Ya existe una atención con código '{codigo}'")
        
        # Validar que el cliente existe
        if razon_social and not any(c['codigo'] == razon_social for c in self.datos_clientes):
            raise ValueError(f"No existe el cliente '{razon_social}'")
        
        # Validar email si se proporciona
        if correo and not self._validar_email(correo):
            raise ValueError("El email no tiene un formato válido")
        
        nueva_atencion = {
            'fila': 0,  # Se asignará al guardar
            'codigo': codigo.strip(),
            'nombres': nombres.strip(),
            'correo': correo.strip(),
            'celular': celular.strip(),
            'razon_social': razon_social.strip()
        }
        
        self.datos_atenciones.append(nueva_atencion)
        return True
    
    def modificar_atencion(self, indice: int, codigo: str, nombres: str, correo: str = "", celular: str = "", razon_social: str = "") -> bool:
        """Modifica una atención existente"""
        if not (0 <= indice < len(self.datos_atenciones)):
            raise ValueError("Índice de atención inválido")
        
        atencion_actual = self.datos_atenciones[indice]
        
        # Validar que el código no esté duplicado (excepto si es el mismo)
        if codigo.upper() != atencion_actual['codigo'].upper():
            if any(a['codigo'].upper() == codigo.upper() for a in self.datos_atenciones):
                raise ValueError(f"Ya existe otra atención con código '{codigo}'")
        
        # Validar que el cliente existe
        if razon_social and not any(c['codigo'] == razon_social for c in self.datos_clientes):
            raise ValueError(f"No existe el cliente '{razon_social}'")
        
        # Validar email si se proporciona
        if correo and not self._validar_email(correo):
            raise ValueError("El email no tiene un formato válido")
        
        # Actualizar datos
        self.datos_atenciones[indice].update({
            'codigo': codigo.strip(),
            'nombres': nombres.strip(),
            'correo': correo.strip(),
            'celular': celular.strip(),
            'razon_social': razon_social.strip()
        })
        
        return True
    
    def eliminar_atencion(self, indice: int) -> bool:
        """Elimina una atención"""
        if not (0 <= indice < len(self.datos_atenciones)):
            raise ValueError("Índice de atención inválido")
        
        del self.datos_atenciones[indice]
        return True
    
    # MÉTODOS DE UTILIDAD
    def _validar_ruc(self, ruc: str) -> bool:
        """Valida formato básico de RUC peruano"""
        # RUC peruano: 11 dígitos
        return bool(re.match(r'^\d{11}$', ruc.strip()))
    
    def _validar_email(self, email: str) -> bool:
        """Valida formato básico de email"""
        patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        return bool(re.match(patron, email.strip()))
    
    def obtener_lista_clientes(self) -> List[str]:
        """Obtiene lista de códigos de clientes para combobox"""
        return [cliente['codigo'] for cliente in self.datos_clientes if cliente['codigo']]
    
    def obtener_cliente_por_codigo(self, codigo: str) -> Optional[Dict]:
        """Obtiene un cliente por su código"""
        for cliente in self.datos_clientes:
            if cliente['codigo'] == codigo:
                return cliente
        return None
    
    def obtener_atenciones_por_cliente(self, codigo_cliente: str) -> List[Dict]:
        """Obtiene todas las atenciones de un cliente específico"""
        return [a for a in self.datos_atenciones if a['razon_social'] == codigo_cliente]
    
    def eliminar_cliente_con_atenciones(self, indice: int, eliminar_atenciones: bool = False) -> bool:
        """Elimina un cliente y opcionalmente sus atenciones"""
        if not (0 <= indice < len(self.datos_clientes)):
            raise ValueError("Índice de cliente inválido")
        
        cliente = self.datos_clientes[indice]
        codigo_cliente = cliente['codigo']
        
        if eliminar_atenciones:
            # Eliminar todas las atenciones asociadas
            self.datos_atenciones = [a for a in self.datos_atenciones if a['razon_social'] != codigo_cliente]
        
        # Eliminar cliente
        del self.datos_clientes[indice]
        return True


## Las ventanas Tkinter (VentanaGestionDatos, VentanaEditorCliente, VentanaEditorAtencion)
## y la función abrir_gestion_datos() están en gestion_ventanas.py


# Función de prueba independiente — mantiene acceso standalone al módulo
def _main_prueba():
    """Función principal para prueba independiente del gestor de datos"""
    import tkinter as tk
    from tkinter import ttk, messagebox
    root = tk.Tk()
    root.withdraw()  # Ocultar ventana principal

    # Configurar estilos
    style = ttk.Style()
    try:
        style.theme_use('clam')
        style.configure("Accent.TButton", foreground="white", background="#0078d4")
    except:
        pass

    # Ruta de prueba (cambiar por tu ruta)
    ruta_excel = r"D:\Sebastian_2\Escritorio\ARCHIVOS AROLUZ\APLICACION\Cotizador_App\V1\COTIZACIÓN v1.2     12-07-2023.xlsm"

    # Verificar que el archivo existe
    if not os.path.exists(ruta_excel):
        messagebox.showerror("Error", f"No se encontró el archivo Excel en:\n{ruta_excel}")
        return

    # Abrir gestión
    from .gestion_ventanas import VentanaGestionDatos
    VentanaGestionDatos(root, ruta_excel)

    root.mainloop()


if __name__ == "__main__":
    _main_prueba()


