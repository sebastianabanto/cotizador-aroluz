# lector_excel.py - VERSIÓN ACTUALIZADA
import os
from typing import List, Tuple, Dict

# Ruta del archivo Excel (la misma que en exportar_excel.py)
def obtener_ruta_excel():
    """Obtiene ruta del Excel desde configuración"""
    import os  # ← MOVER IMPORT AL INICIO
    
    try:
        from .configuracion import ConfiguracionManager
        config_manager = ConfiguracionManager()
        
        # Primero intentar obtener de la configuración
        config = config_manager.config
        if "plantilla_excel" in config.get("rutas", {}):
            ruta_config = config["rutas"]["plantilla_excel"]
            if ruta_config and os.path.exists(ruta_config):
                return ruta_config
        
        # Si no, usar get_ruta_plantilla() (plantillas/NOMBRE.xlsm)
        ruta_plantilla = config_manager.get_ruta_plantilla()
        if os.path.exists(ruta_plantilla):
            return ruta_plantilla
            
    except Exception as e:
        print(f"⚠️ Error obteniendo ruta de plantilla: {e}")
    
    # Último recurso: buscar en carpeta plantillas local
    base_dir = os.path.dirname(os.path.abspath(__file__))
    ruta_local = os.path.join(os.path.dirname(base_dir), "plantillas", "COTIZACIÓN v1.2     12-07-2023.xlsm")
    
    if os.path.exists(ruta_local):
        return ruta_local
    
    # Si nada funciona, devolver la ruta esperada (aunque no exista)
    return os.path.join(os.path.dirname(base_dir), "plantillas", "COTIZACIÓN v1.2     12-07-2023.xlsm")

RUTA_EXCEL = obtener_ruta_excel()

def verificar_archivo_excel():
    """Verifica que el archivo Excel exista"""
    global RUTA_EXCEL
    RUTA_EXCEL = obtener_ruta_excel()  # Actualizar ruta dinámicamente
    return os.path.exists(RUTA_EXCEL)

def leer_columna_excel(nombre_hoja: str, columna: str = "A", max_filas: int = 100, fila_inicio: int = 2) -> List[str]:
    """
    Lee una columna completa de una hoja específica del Excel
    
    Args:
        nombre_hoja: Nombre de la hoja (ej: "CLIENTES", "ATENCIÓN", "MONEDA")
        columna: Letra de la columna a leer (por defecto "A")
        max_filas: Número máximo de filas a leer
        fila_inicio: Fila desde donde empezar a leer (por defecto 2, saltando encabezados)
    
    Returns:
        Lista de valores encontrados (sin valores vacíos)
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Se requiere openpyxl. Instale con: pip install openpyxl")
    
    if not verificar_archivo_excel():
        raise FileNotFoundError(f"No se encontró el archivo Excel en: {RUTA_EXCEL}")
    
    try:
        # Abrir el libro en modo solo lectura
        workbook = openpyxl.load_workbook(RUTA_EXCEL, read_only=True, data_only=True)
        
        # Verificar que la hoja exista
        if nombre_hoja not in workbook.sheetnames:
            available_sheets = ", ".join(workbook.sheetnames)
            raise ValueError(f"La hoja '{nombre_hoja}' no existe. Hojas disponibles: {available_sheets}")
        
        worksheet = workbook[nombre_hoja]
        valores = []
        
        print(f"📖 Leyendo {nombre_hoja} desde fila {fila_inicio} hasta {max_filas + fila_inicio - 1}")
        
        # Leer la columna especificada DESDE fila_inicio
        for fila in range(fila_inicio, max_filas + fila_inicio):  # ✅ Empieza desde fila_inicio (2)
            celda = worksheet[f"{columna}{fila}"]
            valor = celda.value
            
            # Solo agregar valores no vacíos y convertir a string
            if valor is not None and str(valor).strip():
                valores.append(str(valor).strip())
                print(f"  🔍 {columna}{fila}: {str(valor).strip()}")
        
        workbook.close()
        print(f"✅ {nombre_hoja}: {len(valores)} elementos encontrados")
        return valores
        
    except Exception as e:
        raise Exception(f"Error leyendo la hoja '{nombre_hoja}': {str(e)}")

def leer_todos_los_datos() -> Dict[str, List[str]]:
    """
    Lee todos los datos necesarios del Excel de una sola vez
    VERSIÓN ACTUALIZADA: Compatible con gestión de datos
    """
    try:
        datos = {}
        
        # 📋 Leer códigos de clientes desde columna A
        print("🏢 Leyendo códigos de clientes...")
        datos['razones_sociales'] = leer_columna_excel("CLIENTES", "A", 100, 2)
        
        # 👥 Leer atenciones con códigos de empresa (columna E)
        print("👥 Leyendo atenciones con empresas...")
        datos['atenciones_completas'] = leer_atenciones_con_empresas_actualizado()
        
        # 💰 Leer monedas
        print("💰 Leyendo monedas...")
        datos['monedas'] = leer_columna_excel("MONEDA", "A", 100, 2)
        
        # Crear lista de todas las atenciones (para compatibilidad)
        datos['atenciones'] = list(datos['atenciones_completas'].keys())
        
        return datos
        
    except Exception as e:
        print(f"❌ Error leyendo datos del Excel: {e}")
        return {
            'razones_sociales': ['CLIENTE EJEMPLO 1', 'CLIENTE EJEMPLO 2'],
            'atenciones': ['JUAN PÉREZ', 'MARÍA GARCÍA'],
            'atenciones_completas': {'JUAN PÉREZ': 'CLIENTE EJEMPLO 1'},
            'monedas': ['SOLES', 'DÓLARES']
        }
    
def leer_atenciones_con_empresas_actualizado() -> Dict[str, str]:
    """
    📋 VERSIÓN ACTUALIZADA: Lee las atenciones con sus empresas usando códigos
    
    Lee NOMBRES (columna B) y el CÓDIGO de empresa (columna E)
    La columna E tiene los mismos códigos que columna A de CLIENTES
    
    Returns:
        Diccionario donde key=nombre_persona, value=codigo_empresa
        Ejemplo: {'Juan Pérez': 'CONKRETO', 'María García': 'EFR1'}
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Se requiere openpyxl. Instale con: pip install openpyxl")
    
    if not verificar_archivo_excel():
        raise FileNotFoundError(f"No se encontró el archivo Excel en: {RUTA_EXCEL}")
    
    try:
        workbook = openpyxl.load_workbook(RUTA_EXCEL, read_only=True, data_only=True)
        
        if "ATENCIÓN" not in workbook.sheetnames:
            raise ValueError("La hoja 'ATENCIÓN' no existe")
        
        worksheet_atenciones = workbook["ATENCIÓN"]
        atenciones_empresas = {}
        
        print(f"👥 Leyendo atenciones con códigos de empresa...")
        for fila in range(2, 102):
            nombres = worksheet_atenciones[f"A{fila}"].value  # NOMBRES
            codigo_empresa = worksheet_atenciones[f"E{fila}"].value  # CÓDIGO EMPRESA
            
            if nombres is not None and codigo_empresa is not None:
                nombres_str = str(nombres).strip()
                codigo_empresa_str = str(codigo_empresa).strip()
                
                if nombres_str and codigo_empresa_str:
                    atenciones_empresas[nombres_str] = codigo_empresa_str
                    print(f"  👥 {nombres_str} → {codigo_empresa_str}")
        
        workbook.close()
        print(f"✅ Relaciones encontradas: {len(atenciones_empresas)}")
        return atenciones_empresas
        
    except Exception as e:
        raise Exception(f"Error leyendo atenciones con empresas: {str(e)}")

def verificar_dependencias() -> List[str]:
    """Verifica si están instaladas las dependencias necesarias"""
    dependencias_faltantes = []
    
    try:
        import openpyxl
    except ImportError:
        dependencias_faltantes.append("openpyxl")
    
    return dependencias_faltantes

def instalar_dependencias() -> Tuple[bool, str]:
    """Instala las dependencias faltantes"""
    import subprocess
    import sys
    
    dependencias = verificar_dependencias()
    if not dependencias:
        return True, "Todas las dependencias ya están instaladas."
    
    try:
        for dep in dependencias:
            print(f"Instalando {dep}...")
            subprocess.check_call([sys.executable, "-m", "pip", "install", dep])
        
        return True, f"Dependencias instaladas correctamente: {', '.join(dependencias)}"
    except Exception as e:
        return False, f"Error instalando dependencias: {str(e)}"

# 🔄 NUEVA FUNCIÓN: Refrescar datos después de gestión
def refrescar_datos_gui(gui_instance):
    """
    Función para refrescar los datos en la GUI después de gestionar datos
    
    Args:
        gui_instance: Instancia de la clase CotizadorAroluz
    """
    try:
        print("🔄 Refrescando datos en GUI después de gestión...")
        gui_instance.actualizar_datos_excel()
        print("✅ Datos refrescados correctamente")
    except Exception as e:
        print(f"❌ Error refrescando datos: {e}")

# Función de prueba
def test_lectura():
    """Función para probar la lectura de datos"""
    print("🧪 Probando lectura de datos del Excel...")
    print(f"📁 Archivo: {RUTA_EXCEL}")
    print(f"📂 Existe: {'✅' if verificar_archivo_excel() else '❌'}")
    
    if not verificar_archivo_excel():
        print("❌ No se puede continuar sin el archivo Excel")
        return
    
    try:
        datos = leer_todos_los_datos()
        
        print("\n📊 Resultados:")
        print(f"🏢 Razones sociales encontradas: {len(datos['razones_sociales'])}")
        for i, rs in enumerate(datos['razones_sociales'][:5]):  # Mostrar solo primeros 5
            print(f"   {i+1}. {rs}")
        if len(datos['razones_sociales']) > 5:
            print(f"   ... y {len(datos['razones_sociales']) - 5} más")
        
        print(f"\n👥 Atenciones encontradas: {len(datos['atenciones'])}")
        for i, at in enumerate(datos['atenciones'][:5]):
            print(f"   {i+1}. {at}")
        if len(datos['atenciones']) > 5:
            print(f"   ... y {len(datos['atenciones']) - 5} más")
        
        print(f"\n💰 Monedas encontradas: {len(datos['monedas'])}")
        for i, mon in enumerate(datos['monedas']):
            print(f"   {i+1}. {mon}")
        
        print(f"\n🔗 Relaciones Atención-Empresa:")
        for persona, empresa in list(datos['atenciones_completas'].items())[:3]:
            print(f"   👥 {persona} → {empresa}")
        if len(datos['atenciones_completas']) > 3:
            print(f"   ... y {len(datos['atenciones_completas']) - 3} más relaciones")
        
        print("\n✅ Lectura completada exitosamente!")
        
    except Exception as e:
        print(f"❌ Error durante la prueba: {e}")

if __name__ == "__main__":
    # Verificar dependencias
    deps = verificar_dependencias()
    if deps:
        print(f"⚠️ Dependencias faltantes: {deps}")
        print("Ejecute: pip install " + " ".join(deps))
    else:
        print("✅ Todas las dependencias están disponibles")
        test_lectura()